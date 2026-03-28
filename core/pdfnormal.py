
# Copyright (c) 2026 Bhanusri mattey
# Licensed under the Business Source License 1.1
# See LICENSE file in the project root for details.
# Commercial use prohibited until 2030-03-01

from PIL import ImageDraw
import bisect
from collections import defaultdict
import copy
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import range_boundaries
from openpyxl.styles import Border, Side
from pdfminer.high_level import extract_pages
from pdfminer.layout import LTLine, LTRect
from pdfminer.layout import LTTextLine, LTTextBox, LTTextContainer, LTTextLineVertical, LTChar
from pdfminer.layout import LAParams
from typing import List, Dict
from core.pdftext import*
from openpyxl.utils import get_column_letter

pdf_path = "test6.pdf"
on_page_done = None




def snap_close_values(values, tolerance=2.0):

    # Snaps edges with values within tolerance
    values = sorted(values)
    snapped = []
    group = []

    for i, val in enumerate(values):
        if not group:
            group.append(val)
        elif abs(val - group[-1]) <= tolerance:
            group.append(val)
        else:
            avg = sum(group) / len(group)
            snapped.append(avg)
            group = [val]

    # handle last group
    if group:
        avg = sum(group) / len(group)
        snapped.append(avg)

    return snapped

class EdgeIndex:
    """EdgeIndex provides efficient spatial lookup for edge objects based on
    their geometric properties (x0, x1, top, bottom)."""
    def __init__(self, edges):
        # Store lists of edges per value
        self.by_x1 = defaultdict(list)
        self.by_top = defaultdict(list)
        self.by_x0 = defaultdict(list)
        self.by_bottom = defaultdict(list)

        for edge in edges:
            self.by_x1[edge["x1"]].append(edge)
            self.by_top[edge["top"]].append(edge)
            self.by_x0[edge["x0"]].append(edge)
            self.by_bottom[edge["bottom"]].append(edge)

        # Sorted unique keys for bisect
        self.x1_keys = sorted(self.by_x1.keys())
        self.top_keys = sorted(self.by_top.keys())
        self.x0_keys = sorted(self.by_x0.keys())
        self.bottom_keys = sorted(self.by_bottom.keys())

    def get_closest_key(self, val, sorted_keys, tolerance=1.0):
    # Use bisect to find the insertion point
        idx = bisect.bisect_left(sorted_keys, val)

        candidates = []
        if idx > 0:
            candidates.append(sorted_keys[idx - 1])
        if idx < len(sorted_keys):
            candidates.append(sorted_keys[idx])

    # Find the closest candidate within tolerance
        closest = min(
        (key for key in candidates if abs(key - val) <= tolerance),
        key=lambda k: abs(k - val),
        default=None
        )
        return closest
    # These now return a LIST of matching edges
    def find_by_x1(self, x1_val, tolerance=1.0):
        key = self.get_closest_key(x1_val, self.x1_keys, tolerance)
        return self.by_x1.get(key, [])

    def find_by_x0(self, x0_val, tolerance=1.0):
        key = self.get_closest_key(x0_val, self.x0_keys, tolerance)
        return self.by_x0.get(key, [])

    def find_by_top(self, top_val, tolerance=1.0):
        key = self.get_closest_key(top_val, self.top_keys, tolerance)
        return self.by_top.get(key, [])

    def find_by_bottom(self, bottom_val, tolerance=1.0):
        key = self.get_closest_key(bottom_val, self.bottom_keys, tolerance)
        return self.by_bottom.get(key, [])
    
    def find_by_x1_and_bottom(self, x1_val, bottom_val, x1_tol=2.0, bottom_tol=2.0):
        x1_key = self.get_closest_key(x1_val, self.x1_keys, x1_tol)
        x1_matches = self.by_x1.get(x1_key, [])

    # Now filter x1_matches by top tolerance manually
        filtered = [
            edge for edge in x1_matches
            if abs(edge["bottom"] - bottom_val) <= bottom_tol
        ]

        return filtered
    
    def find_by_x0_and_top(self, x0_val, top_val, x0_tol=6.0, top_tol=6.0):
        x0_key = self.get_closest_key(x0_val, self.x0_keys, x0_tol)
        x0_matches = self.by_x0.get(x0_key, [])

    # Now filter x1_matches by top tolerance manually
        filtered = [
            edge for edge in x0_matches
            if abs(edge["top"] - top_val) <= top_tol
        ]

        return filtered
    
def Intersection(h_edge, v_edge):
    # Checks whether Horizontal edge and Vertical edge are intersecting
    h0, h1 = h_edge["x0"], h_edge["x1"]
    hy = h_edge["top"]  # or bottom — they should be the same for horizontal
    vx = v_edge["x0"]   # or x1 — same for vertical
    vtop, vbottom = v_edge["top"], v_edge["bottom"]
   
    # Check if vx lies between h0 and h1
    # AND hy lies between vtop and vbottom
    if (h0 <= vx <= h1 or nearedges(h1,vx) or nearedges(h0,vx)) and (nearedges(hy,vbottom) or nearedges(hy,vtop) or vtop <= hy <= vbottom):
        return True
    return False
    
def nearedges(a,b):
    # Checks if both values are within tolerance
    tolerance = 2
    if abs(a-b) <= tolerance:
        return True
    return False



def SnapHEdges(h_edges,hsnappedx0,hsnappedx1,hsnappedtop):
    # Snaps all Horizontal Edges
    for edge in h_edges:
        for x0 in hsnappedx0:
            if nearedges(edge["x0"],x0):
                edge["x0"] = x0
        for x1 in hsnappedx1:
            if nearedges(edge["x1"],x1):
                edge["x1"] = x1
        for top in hsnappedtop:
            if nearedges(edge["top"],top):
                edge["top"] = top

def SnapHEdge(h_edges,hsnappedtop):
    # Snaps Horizontal Edge
    for edge in h_edges:
        for top in hsnappedtop:
            if nearedges(edge["top"],top):
                edge["top"] = top

def SnapVEdges(v_edges,vsnappedx0,vsnappedtop,vsnappedbottom):
    # Snaps all Vertical Edges
    for edge in v_edges:
        for x0 in vsnappedx0:
            if nearedges(edge["x0"],x0):
                edge["x0"] = x0
        for top in vsnappedtop:
            if nearedges(edge["top"],top):
                edge["top"] = top
        for bottom in vsnappedbottom:
            if nearedges(edge["bottom"],bottom):
                edge["bottom"] = bottom

def SnapVEdge(v_edges,vsnappedx0):
    # Snaps Vertical Edge
    for edge in v_edges:
        for x0 in vsnappedx0:
            if nearedges(edge["x0"],x0):
                edge["x0"] = x0


def FindIntersections(h_edges,v_edges):
    # Finds intersections of Horizontal edges and Vertical edges
    for h_edge in h_edges:
        for v_edge in v_edges:
            if Intersection(h_edge,v_edge):
                x = v_edge["x0"]
                y = h_edge["top"]
                h_edge.setdefault("intersections", []).append({
                    "edge": v_edge,
                    "point": (x, y)
                })
                v_edge.setdefault("intersections", []).append({
                    "edge": h_edge,
                    "point": (x, y)
                })
    return h_edges,v_edges



def RemoveIntersections(h_edges, v_edges):
    # Removes intersections of edges
    for h_edge in h_edges:
        if "intersections" in h_edge:
            del h_edge["intersections"]
    for v_edge in v_edges:
        if "intersections" in v_edge:
            del v_edge["intersections"]
    return h_edges, v_edges

def RemoveIndexing(h_edges, v_edges):
    # Resets indexes of edges
    for h_edge in h_edges:
        if "index" in h_edge:
            del h_edge["index"]
    for v_edge in v_edges:
        if "index" in v_edge:
            del v_edge["index"]
    return h_edges, v_edges



def LeftIntersection(h_edge):
    # Finds Left Intersection of Horizontal edge
    dist = 0
    minDist = float('inf')
    minEdge = None
    v_int_List = h_edge["intersections"]
    #v_int_List = h_edge.get("intersections", [])
    for v_int in v_int_List:
        dist =  abs(h_edge["x0"]- v_int["edge"]["x0"])
        if dist < minDist:
            minDist = dist
            minEdge = v_int["edge"]
    return minEdge


def RightIntersection(h_edge):
    # Finds Right Intersection of Horizontal edge
    dist = 0
    minDist = float('inf')
    minEdge = None
    v_int_List = h_edge.get("intersections", [])
    for v_int in v_int_List:
        dist =  abs(h_edge["x1"]- v_int["edge"]["x1"])
        if dist < minDist:
            minDist = dist
            minEdge = v_int["edge"]

    return minEdge

def TopIntersection(v_edge):
    # Finds Top Intersection of Vertical Edge
    dist = 0
    minDist = float('inf')
    minEdge = None
    h_int_List = v_edge["intersections"]
    for h_int in h_int_List:
        dist =  abs(v_edge["top"]- h_int["edge"]["top"])
        if dist < minDist:
            minDist = dist
            minEdge = h_int["edge"]
    return minEdge

def BottomIntersection(v_edge):
    # Finds Bottom Intersection of Vertical Edge
    dist = 0
    minDist = float('inf')
    minEdge = None
    h_int_List = v_edge["intersections"]
    if h_int_List is not None:
        for h_int in h_int_List:
            dist =  abs(v_edge["bottom"]- h_int["edge"]["bottom"])
            if dist < minDist:
                minDist = dist
                minEdge = h_int["edge"]
    return minEdge

def FindHorizontalEdges(table,h_edges):
    # Finds horizontal edges within boundary of table formed
    
    h_edges_tb = []
    left_edge = table[0]
    right_edge = table[1]
    leftPos = table[0]["x0"]
    rightPos = table[1]["x1"]
    topPos = table[2]["top"]
    bottomPos = table[3]["bottom"]
    for h_edge in h_edges:
        if (
            (nearedges(h_edge["x0"], leftPos) or leftPos <= h_edge["x0"] <= rightPos) and
            (nearedges(h_edge["x1"], rightPos) or leftPos <= h_edge["x1"] <= rightPos) and
            (nearedges(h_edge["top"], topPos) or topPos <= h_edge["top"] <= bottomPos) and
            (nearedges(h_edge["bottom"], bottomPos) or topPos <= h_edge["bottom"] <= bottomPos)
        ):
            if h_edge not in h_edges_tb:
                h_edges_tb.append(h_edge)

    for h_edge in h_edges:
        if(Intersection(h_edge,left_edge) and Intersection(h_edge,right_edge) and topPos <= h_edge["top"] <= bottomPos):
            if h_edge not in h_edges_tb:
                h_edges_tb.append(h_edge)

    for h_edge in h_edges:
        if(leftPos <= h_edge["x0"] <= rightPos and Intersection(h_edge,right_edge) and topPos <= h_edge["top"] <= bottomPos):
            if h_edge not in h_edges_tb:
                h_edges_tb.append(h_edge)

    for h_edge in h_edges:
        if(Intersection(h_edge,left_edge) and leftPos <= h_edge["x1"] <= rightPos and topPos <= h_edge["top"] <= bottomPos):
            if h_edge not in h_edges_tb:
                h_edges_tb.append(h_edge)


    return h_edges_tb

def FindVerticalEdges(table,v_edges):
    # Finds vertical edges within boundary of table formed
    v_edges_tb = []
    top_edge = table[2]
    bottom_edge = table[3]
    leftPos = table[0]["x0"]
    rightPos = table[1]["x1"]
    topPos = table[2]["top"]
    bottomPos = table[3]["bottom"]
    for v_edge in v_edges:
        if (
            (nearedges(v_edge["x0"], leftPos) or leftPos <= v_edge["x0"] <= rightPos) and
            (nearedges(v_edge["x1"], rightPos) or leftPos <= v_edge["x1"] <= rightPos) and
            (nearedges(v_edge["top"], topPos) or  topPos <= v_edge["top"] <= bottomPos) and
            (nearedges(v_edge["bottom"], bottomPos) or topPos <= v_edge["bottom"] <= bottomPos )
        ):
            if v_edge not in v_edges_tb:
                v_edges_tb.append(v_edge)

    for v_edge in v_edges:
        if(Intersection(top_edge,v_edge) and Intersection(bottom_edge,v_edge) and leftPos <= v_edge["x0"] <= rightPos):
            if v_edge not in v_edges_tb:
                v_edges_tb.append(v_edge)

    for v_edge in v_edges:
        if topPos <= v_edge["top"] <= bottomPos and Intersection(bottom_edge,v_edge) and leftPos <= v_edge["x0"] <= rightPos:
            if v_edge not in v_edges_tb:
                v_edges_tb.append(v_edge)

    for v_edge in v_edges:
        if Intersection(top_edge,v_edge) and topPos <= v_edge["bottom"] <= bottomPos and leftPos <= v_edge["x0"] <= rightPos:
            if v_edge not in v_edges_tb:
                v_edges_tb.append(v_edge)
    return v_edges_tb






def FilterLooseEdges(h_edges, v_edges):
    # Removes edges with less than 2 intersections
    v_edges = [v for v in v_edges if len(v.get("intersections", [])) >= 2]
    h_edges = [h for h in h_edges if len(h.get("intersections", [])) >= 2]
    v_edges = [v for v in v_edges if not nearedges(v["top"],v["bottom"])]
    h_edges = [h for h in h_edges if not nearedges(h["x0"],h["x1"])]
    return h_edges, v_edges  




def ModifyHEdge(edge1,edge2):
    # Modifies Horizontal edge
    edge1["x1"] = edge2["x1"]

def ModifyVEdge(edge1,edge2):
    # Modifies Vertical Edge
    edge1["bottom"] = edge2["bottom"]


def FindTable(index,h_edges,v_edges):
    # Finds table with respect to top horizontal edge
    ed = []
    
    top_edge = h_edges[index]
    left_edge_t = LeftIntersection(top_edge)
    right_edge_t = RightIntersection(top_edge)
    bottom_edge_l = BottomIntersection(left_edge_t)
    bottom_edge_r = BottomIntersection(right_edge_t)
    

    if bottom_edge_r != bottom_edge_l:
        bottom_edge = bottom_edge_l if bottom_edge_l["top"] > bottom_edge_r["top"] else bottom_edge_r
    if bottom_edge_r != bottom_edge_l:
        if nearedges(bottom_edge_l["top"],bottom_edge_r["top"]):
            ModifyHEdge(bottom_edge_l,bottom_edge_r)
            bottom_edge = bottom_edge_l
    else:
        bottom_edge = bottom_edge_r

   

    left_edge_b = LeftIntersection(bottom_edge)
    left_edge_t = LeftIntersection(top_edge)
    
    if left_edge_t != left_edge_b:
        left_edge = left_edge_t if left_edge_t["x0"] < left_edge_b["x0"] else left_edge_b
    if left_edge_t != left_edge_b:
        if nearedges(left_edge_t["x0"],left_edge_b["x0"]):
            
            ModifyVEdge(left_edge_t,left_edge_b)
            left_edge = left_edge_t
    else:
        left_edge = left_edge_b

    


    right_edge_b = RightIntersection(bottom_edge)
    right_edge_t = RightIntersection(top_edge)

   

    if right_edge_t != right_edge_b:
        right_edge = right_edge_t if right_edge_t["x0"] > right_edge_b["x0"] else right_edge_b
    if right_edge_t != right_edge_b:
        if nearedges(right_edge_t["x0"],right_edge_b["x0"]):
            
            ModifyVEdge(right_edge_t,right_edge_b)
            right_edge = right_edge_t
    else:
        right_edge = right_edge_b

    ed.append(left_edge)
    ed.append(right_edge)
    ed.append(top_edge)  
    ed.append(bottom_edge)

    ed,h_edges,v_edges = FixOverlappingEdges(ed,h_edges,v_edges)

    return ed,h_edges,v_edges




def isTableConnected(tb):

    # Checks if the boundary of table is connected or not
    if (Intersection(tb[2],tb[0]) and nearedges(tb[2]["top"],tb[0]["top"]))  and (Intersection(tb[2],tb[1]) and nearedges(tb[2]["top"],tb[1]["top"]))  and (Intersection(tb[3],tb[0]) and nearedges(tb[3]["bottom"],tb[0]["bottom"]))  and (Intersection(tb[3],tb[1]) and nearedges(tb[3]["bottom"],tb[1]["bottom"])):
        return True
    return False



def AddVLeftEdge(edge1,edge2,v_edges,min_prop):
    # Adds Left Edge
    edge = copy.deepcopy(v_edges[0])
    edge["top"] = edge1["top"]
    edge["bottom"] = edge2["top"]
    
    edge["x0"] = edge2[min_prop] if edge1[min_prop] < edge2[min_prop] else edge1[min_prop]
    edge["x1"] = edge2[min_prop] if edge1[min_prop] < edge2[min_prop] else edge1[min_prop]
    v_edges.append(edge)
    return v_edges,edge

                    
def AddVRightEdge(edge1,edge2,v_edges,min_prop):
    # Adds Right Edge
    edge = copy.deepcopy(v_edges[0])
    edge["top"] = edge1["top"]
    edge["bottom"] = edge2["top"]
    
    edge["x0"] = edge1[min_prop] if edge1[min_prop] > edge2[min_prop] else edge2[min_prop]
    edge["x1"] = edge1[min_prop] if edge1[min_prop] > edge2[min_prop] else edge2[min_prop]
    v_edges.append(edge)
    return v_edges,edge

def AddHTopEdge(edge1,edge2,h_edges,min_prop):
    # Adds Top Edge
    edge = copy.deepcopy(h_edges[0])
    edge["x0"] = edge1["x0"]
    edge["x1"] = edge2["x0"]
    
    
    edge["top"] = edge2[min_prop] if edge1[min_prop] < edge2[min_prop] else edge1[min_prop]
    edge["bottom"] = edge2[min_prop] if edge1[min_prop] < edge2[min_prop] else edge1[min_prop]
    
    h_edges.append(edge)
    return h_edges,edge

def AddHBottomEdge(edge1,edge2,h_edges,min_prop):
    # Adds Bottom edge
    edge = copy.deepcopy(h_edges[0])
    edge["x0"] = edge1["x0"]
    edge["x1"] = edge2["x0"]
    
    edge["top"] = edge2[min_prop] if edge1[min_prop] > edge2[min_prop] else edge1[min_prop]
    edge["bottom"] = edge2[min_prop] if edge1[min_prop] > edge2[min_prop] else edge1[min_prop]
    h_edges.append(edge)
    return h_edges,edge

def ReconstructEdges(tb,h_edges,v_edges):
    # Reconstructs boundary edges of table
    if tb[1]["top"] < tb[0]["top"]:
        tb[0]["top"] = tb[1]["top"]
        h_edges,edge = AddHTopEdge(tb[0],tb[1],h_edges,"top")
        tb[2] = edge
    else:
        tb[1]["top"] = tb[0]["top"]
        h_edges,edge = AddHTopEdge(tb[0],tb[1],h_edges,"top")
        tb[2] = edge
    
    if tb[0]["bottom"] < tb[1]["bottom"]:
        tb[0]["bottom"] = tb[1]["bottom"]
        h_edges,edge = AddHBottomEdge(tb[0],tb[1],h_edges,"bottom")
        tb[3] = edge
    else:
        tb[1]["bottom"] = tb[0]["bottom"]
        h_edges,edge = AddHBottomEdge(tb[0],tb[1],h_edges,"bottom")
        tb[3] = edge

    if tb[2]["x1"] > tb[3]["x1"]:
        tb[3]["x1"] = tb[2]['x1']
        v_edges,edge = AddVRightEdge(tb[2],tb[3],v_edges,"x1")
        tb[1] = edge
    else:
        tb[2]["x1"] = tb[3]['x1']
        v_edges,edge = AddVRightEdge(tb[2],tb[3],v_edges,"x1")
        tb[1] = edge

    if tb[2]["x0"] < tb[3]["x0"]:
        tb[3]["x0"] = tb[2]['x0']
        v_edges,edge = AddVLeftEdge(tb[2],tb[3],v_edges,"x0")
        tb[0] = edge
    else:
        tb[2]["x0"] = tb[3]['x0']
        v_edges,edge = AddVLeftEdge(tb[2],tb[3],v_edges,"x0")
        tb[0] = edge
    return h_edges,v_edges




def FixOverlappingEdges(table,h_edges,v_edges):
    # Fixes overlapping edges in boundary of table
    v_edges = [
    v for v in v_edges
    if not (
        (nearedges(v["x0"],table[0]["x0"])) 
        and
        ((nearedges(v["top"],table[0]["top"]) and table[0]["top"] < v["bottom"] < table[0]["bottom"])
        or
        (nearedges(v["bottom"],table[0]["bottom"]) and table[0]["top"] < v["top"] < table[0]["bottom"]))
    )
    ]

    v_edges = [
    v for v in v_edges
    if not (
        (nearedges(v["x0"],table[1]["x0"])) 
        and
        ((nearedges(v["top"],table[1]["top"]) and table[1]["top"] < v["bottom"] < table[1]["bottom"])
        or
        (nearedges(v["bottom"],table[1]["bottom"]) and table[1]["top"] < v["top"] < table[1]["bottom"]))
    )
    ]

    h_edges = [
    h for h in h_edges
    if not (
        (nearedges(h["top"],table[2]["top"])) 
        and
        ((nearedges(h["x0"],table[2]["x0"]) and table[2]["x0"] < h["x1"] < table[2]["x1"])
        or
        (nearedges(h["x1"],table[2]["x1"]) and table[2]["x0"] < h["x0"] < table[2]["x1"]))
    )
    ]

    h_edges = [
    h for h in h_edges
    if not (
        (nearedges(h["top"],table[3]["top"])) 
        and
        ((nearedges(h["x0"],table[3]["x0"]) and table[3]["x0"] < h["x1"] < table[3]["x1"])
        or
        (nearedges(h["x1"],table[3]["x1"]) and table[3]["x0"] < h["x0"] < table[3]["x1"]))
    )
    ]

    return table,h_edges,v_edges

          
 
def snap_edges(edges,x_tolerance,y_tolerance):
    # Snaps edges within the same tolerance
    v_edges = [e for e in edges if e["orientation"] == "v"]
    h_edges = [e for e in edges if e["orientation"] == "h"]
    
    h_edges_dict = EdgeIndex(h_edges)

    # Snapping close values of Horizontal edges
    hsnappedtop = snap_close_values(h_edges_dict.top_keys,y_tolerance)
    SnapHEdge(h_edges,hsnappedtop) 

    v_edges_dict = EdgeIndex(v_edges)

    # Snapping close values of Vertical edges
    vsnappedx0 = snap_close_values(v_edges_dict.x0_keys,x_tolerance)
    SnapVEdge(v_edges,vsnappedx0)

    return h_edges+v_edges



def ResizeVEdge(last,edge):
    # Resizes vertical edge
    last["bottom"] = edge["bottom"]
    return last

def ResizeHEdge(last,edge):
    # Resizes horizontal edge
    last["x1"] = edge["x1"]
    return last

def join_edge_group(edges, orientation, threshold=1):
    """
    Merges nearby edges (lines) on a given axis.
    Example: horizontal lines (axis="y"), vertical lines (axis="x").
    """
   

    if orientation == "h":
        group_key = lambda e: (round(e["top"], 1))
        sort_key = lambda e: e["x0"]
    else:
        group_key = lambda e: (round(e["x0"], 1))
        sort_key = lambda e: e["top"]

    from collections import defaultdict

    grouped = defaultdict(list)
    for edge in edges:
        grouped[group_key(edge)].append(edge)

    merged = []
    for group in grouped.values():
        group = list(sorted(group, key=sort_key))
        
        merged_line = [group[0]]
        
        for edge in group[1:]:
            last = merged_line[-1]
            if orientation == "v":
                if edge["top"] - last["bottom"] < threshold:
                    if edge["bottom"] > last["bottom"]:
                        
                        merged_line[-1] = ResizeVEdge(last,edge)
                        
                else:
                    merged_line.append(edge)
               

                
            else:
                if edge["x0"] - last["x1"] < threshold:
                    if edge["x1"] > last["x1"]:
                        merged_line[-1] = ResizeHEdge(last,edge)
                       
                else :
                    merged_line.append(edge)
                
                
        merged.append(merged_line)
                                     
    flat = [item for sublist in merged for item in sublist]
    return flat




def merge_edges_1(edges,snap_x_tolerance,snap_y_tolerance,join_x_tolerance,join_y_tolerance):
    """
    Using the `snap_edges` and `join_edge_group` methods above,
    merge a list of edges into a more "seamless" list.
    """

    

    if snap_x_tolerance > 0 or snap_y_tolerance > 0:
        edges = snap_edges(edges, snap_x_tolerance, snap_y_tolerance)

    v_edges = [e for e in edges if e["orientation"] == "v"]
    h_edges = [e for e in edges if e["orientation"] == "h"]

    v_edges = join_edge_group(v_edges,"v",join_y_tolerance)
    h_edges = join_edge_group(h_edges,"h",join_x_tolerance)
    
    
    
    return h_edges+v_edges
    
    

def PdfCleaner(x):
    """merges edges and finds vertical intersections of horizontal edges and finds 
       horizontal intersections of vertical edges removes edges with less than 
       2 intersections"""
    total_edges = merge_edges_1(x,5,5,2,2)
    
   
    v_edges = [e for e in total_edges if e["orientation"] == "v"]
    h_edges = [e for e in total_edges if e["orientation"] == "h"]

   
    
    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)
    h_edges,v_edges = FilterLooseEdges(h_edges,v_edges)
    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)
    
    
    h_edges_dict = EdgeIndex(h_edges)

    # Snapping close values of Horizontal edges
    hsnappedx0 = snap_close_values(h_edges_dict.x0_keys)
    hsnappedx1 = snap_close_values(h_edges_dict.x1_keys)
    hsnappedtop = snap_close_values(h_edges_dict.top_keys)

    SnapHEdges(h_edges,hsnappedx0,hsnappedx1,hsnappedtop)
    
    v_edges_dict = EdgeIndex(v_edges)

    # Snapping close values of Vertical edges
    vsnappedx0 = snap_close_values(v_edges_dict.x0_keys)
    vsnappedtop = snap_close_values(v_edges_dict.top_keys)
    vsnappedbottom = snap_close_values(v_edges_dict.bottom_keys)

    SnapVEdges(v_edges,vsnappedx0,vsnappedtop,vsnappedbottom)

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)
   
    return h_edges,v_edges


def CheckTableExtended(table,h_edges,v_edges):
    """Checks if top edge and bottom edge are extended i.e left intersection of 
        both edges does not match with x0 or x1 of those edges
        Simialrly left edge and right edge"""
    isExtended = False
    
    if not (nearedges(table[0]["x0"],table[2]["x0"]) or nearedges(table[0]["x0"],table[3]["x0"])):
        v_edges,edge =  AddVLeftEdge(table[2],table[3],v_edges,"x0")
        table[0] = edge
        isExtended = True
        
    if not (nearedges(table[1]["x1"],table[2]["x1"]) or nearedges(table[1]["x1"],table[3]["x1"])):
        v_edges,edge =  AddVRightEdge(table[2],table[3],v_edges,"x1")
        table[1] = edge
        isExtended = True
        
    if not (nearedges(table[2]["top"],table[0]["top"]) or nearedges(table[2]["top"],table[1]["top"])):
        
        h_edges,edge =  AddHTopEdge(table[0],table[1],h_edges,"top")
        table[2] = edge
        isExtended = True
        
    if not (nearedges(table[3]["bottom"],table[0]["bottom"]) or nearedges(table[3]["bottom"],table[1]["bottom"])):
        h_edges,edge =  AddHBottomEdge(table[0],table[1],h_edges,"bottom")
        table[3] = edge
        isExtended = True
        
    return isExtended,h_edges,v_edges 

    



def FindBottomEdge(top,left,right,h_edges_tb):
    # Finds the bottom edge of a cell in a table
    filtered_edges = [h for h in h_edges_tb if h["top"] > top["top"]]
    filtered_edges_1 = [h for h in filtered_edges if Intersection(h,left) and Intersection(h,right)]
    edges_sorted_by_top = sorted(filtered_edges_1, key=lambda e: e["top"])
    bottom = edges_sorted_by_top[0]
    return bottom

def LoopFindTable(h_edges,v_edges):
     # Loops through edges and returns a table
     i = 0
     while(i<len(h_edges)-1):
        try:
            table,h_edges,v_edges = FindTable(i,h_edges,v_edges)
            if table:
                return table,h_edges,v_edges
                
        except:
            i = i+1

     return None,h_edges,v_edges


def Indexing(h_edges_tb,v_edges_tb):
    """Indexes the edges if two horizontal edges have same y position they get same index
        Similarly if tow vertical edges have same x position they get same index"""
    h_index = 0
    v_index = 0
    for j,h_edge in enumerate(h_edges_tb):
        if j == 0:
            h_edge["index"] = h_index
            last = h_edge
           
        if j > 0:
            if last["top"] == h_edge["top"]:
                h_edge["index"] = h_index
                last = h_edge
                
            elif last["top"] < h_edge["top"]:
                h_index = h_index + 1
                h_edge["index"] = h_index
                last = h_edge

    for k,v_edge in enumerate(v_edges_tb):
        if k == 0:
            v_edge["index"] = v_index
            last = v_edge
           
        if k > 0:
            if last["x0"] == v_edge["x0"]:
                v_edge["index"] = v_index
                last = v_edge
                
            elif last["x0"] < v_edge["x0"]:
                v_index = v_index + 1
                v_edge["index"] = v_index
                last = v_edge
                
    return h_edges_tb,v_edges_tb

def is_same_edge(e1, e2, tol=2):
    # Checks if two edges are same
    return (
        abs(e1["x0"] - e2["x0"]) < tol and
        abs(e1["x1"] - e2["x1"]) < tol and
        abs(e1["top"] - e2["top"]) < tol and
        abs(e1["bottom"] - e2["bottom"]) < tol and
        e1["orientation"] == e2["orientation"]
    )

def CheckMisAlignedTableX(h_edges_tb, v_edges_tb, h_edges,v_edges):
    # Checks if top edge left intersection is same as second top edge similarly right intersection
    right_1 = RightIntersection(h_edges_tb[0])
    right_2 = RightIntersection(h_edges_tb[1])
    right_3 = RightIntersection(h_edges_tb[-1])
    right_4 = RightIntersection(h_edges_tb[-2])
    left_1  = LeftIntersection(h_edges_tb[0])
    left_2  = LeftIntersection(h_edges_tb[1])
    left_3 = LeftIntersection(h_edges_tb[-1])
    left_4 = LeftIntersection(h_edges_tb[-2])
    width_1 = h_edges_tb[0]["x1"] - h_edges_tb[0]["x0"]
    width_2 = h_edges_tb[1]["x1"] - h_edges_tb[1]["x0"]
    width_3 = h_edges_tb[-1]["x1"] - h_edges_tb[-1]["x0"]
    width_4 = h_edges_tb[-2]["x1"] - h_edges_tb[-2]["x0"]
    isMisAligned = False
    
    if not is_same_edge(right_1, right_2) and width_1 < width_2:
        
        new_x1 = right_2["x1"]
        new_top = h_edges_tb[0]["top"]
       
        # Reflect the same update in the original h_edges
        for edge in h_edges:
            if edge is h_edges_tb[0]:  # same object reference
                edge["x1"] = new_x1
                break
        for edge in v_edges:
            if edge is right_2:  # same object reference
                edge["top"] = new_top
                break
        
        isMisAligned = True

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    if not is_same_edge(right_3,right_4) and width_3 < width_4:
        
        new_x1 = right_4["x1"]
        new_bottom = h_edges_tb[-1]["bottom"]
       
        # Reflect the same update in the original h_edges
        for edge in h_edges:
            if edge is h_edges_tb[-1]:  # same object reference
                edge["x1"] = new_x1
                break
        for edge in v_edges:
            if edge is right_4:  # same object reference
                edge["bottom"] = new_bottom
                break
        
        isMisAligned = True
    
    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    if not is_same_edge(left_1, left_2) and width_1 < width_2:
        
        new_x0 = left_2["x0"]
        new_top = h_edges_tb[0]["top"]
       
        # Reflect the same update in the original h_edges
        for edge in h_edges:
            if edge is h_edges_tb[0]:  # same object reference
                edge["x0"] = new_x0
                break
        for edge in v_edges:
            if edge is left_2:  # same object reference
                edge["top"] = new_top
                break
        
        isMisAligned = True

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    if not is_same_edge(left_3,left_4) and width_3 < width_4:
        
        new_x0 = left_4["x0"]
        new_bottom = h_edges_tb[-1]["bottom"]
       
        # Reflect the same update in the original h_edges
        for edge in h_edges:
            if edge is h_edges_tb[-1]:  # same object reference
                edge["x0"] = new_x0
                break
        for edge in v_edges:
            if edge is left_4:  # same object reference
                edge["bottom"] = new_bottom
                break
        
        isMisAligned = True
    
    return isMisAligned,h_edges,v_edges

def CheckMisAlignedTableY(h_edges_tb, v_edges_tb, h_edges,v_edges):
    # Checks if left edge top intersection is same as second left edge similarly bottom intersection
    bottom_1 = BottomIntersection(v_edges_tb[0])
    bottom_2 = BottomIntersection(v_edges_tb[1])
    bottom_3 = BottomIntersection(v_edges_tb[-1])
    bottom_4 = BottomIntersection(v_edges_tb[-2])
    top_1 = TopIntersection(v_edges_tb[0])
    top_2 = TopIntersection(v_edges_tb[1])
    top_3 = TopIntersection(v_edges_tb[-1])
    top_4 = TopIntersection(v_edges_tb[-2])
    height_1 = v_edges_tb[0]["bottom"] - v_edges_tb[0]["top"]
    height_2 = v_edges_tb[1]["bottom"] - v_edges_tb[1]["top"]
    height_3 = v_edges_tb[-1]["bottom"] - v_edges_tb[-1]["top"]
    height_4 = v_edges_tb[-2]["bottom"] - v_edges_tb[-2]["top"]
    isMisAligned = False
    
    if not is_same_edge(bottom_1, bottom_2) and height_1 < height_2:
        
        new_bottom = bottom_2["bottom"]
        new_x0 = v_edges_tb[0]["x0"]
       
        # Reflect the same update in the original h_edges
        for edge in v_edges:
            if edge is v_edges_tb[0]:  # same object reference
                edge["bottom"] = new_bottom
                break
        for edge in h_edges:
            if edge is bottom_2:  # same object reference
                edge["x0"] = new_x0
                break
        
        isMisAligned = True

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    if not is_same_edge(bottom_3,bottom_4) and height_3 < height_4:
        
        new_bottom = bottom_4["bottom"]
        new_x1 = v_edges_tb[-1]["x1"]
       
        # Reflect the same update in the original h_edges
        for edge in v_edges:
            if edge is v_edges_tb[-1]:  # same object reference
                edge["bottom"] = new_bottom
                break
        for edge in h_edges:
            if edge is bottom_4:  # same object reference
                edge["x1"] = new_x1
                break
        
        isMisAligned = True
    
    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    if not is_same_edge(top_1, top_2) and height_1 < height_2:
        
        new_top = top_2["top"]
        new_x0 = v_edges_tb[0]["top"]
       
        # Reflect the same update in the original h_edges
        for edge in v_edges:
            if edge is v_edges_tb[0]:  # same object reference
                edge["top"] = new_top
                break
        for edge in h_edges:
            if edge is top_2:  # same object reference
                edge["x0"] = new_x0
                break
        
        isMisAligned = True

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    if not is_same_edge(top_3,top_4) and height_3 < height_4:
        
        new_top = top_4["top"]
        new_x1 = v_edges_tb[-1]["x1"]
       
        # Reflect the same update in the original h_edges
        for edge in v_edges:
            if edge is v_edges_tb[-1]:  # same object reference
                edge["top"] = new_top
                break
        for edge in h_edges:
            if edge is top_4:  # same object reference
                edge["x1"] = new_x1
                break
        
        isMisAligned = True
    
    return isMisAligned,h_edges,v_edges

def edge_exists(edge, edge_list):
    # Check if edge exists in a list of edges
    for e in edge_list:
        if is_same_edge(edge, e):
            return True
    return False

def FormCells(h_edges_tb,v_edges_tb):
    # Constrcuting cells from the table formed
    v_edges_tb.sort(key=lambda e: e["x0"])
    h_edges_tb.sort(key=lambda e: e["top"])
    h_edges_tb,v_edges_tb = RemoveIndexing(h_edges_tb,v_edges_tb)
    h_edges_tb, v_edges_tb = Indexing(h_edges_tb,v_edges_tb)
    maxRowIndex = h_edges_tb[-1]["index"]
    maxColumnIndex = v_edges_tb[-1]["index"]
    table = []
    pts1 = []
    max_dim = (maxRowIndex,maxColumnIndex)
    
    for j,h_edge in enumerate(h_edges_tb):
        
        if j < len(h_edges_tb) - 1:
            
            v_int = [v.copy() for v in h_edge["intersections"]]

            v_int = [e for e in v_int if not nearedges(e["edge"]["bottom"],h_edge["bottom"])]
            filtered_v_int = []
            

            for v in v_int:
                if edge_exists(v["edge"], v_edges_tb):
                    filtered_v_int.append(v)

            for i,v in enumerate(filtered_v_int):
               
                if i < len(filtered_v_int) - 1: 
                    start_row = 0
                    end_row = 0
                    start_col = 0
                    end_col = 0
                    cell = []
                    pts = []
                    a = []
                    b = []
                    c = []
                    d = []
                    left = filtered_v_int[i]["edge"]
                    right = filtered_v_int[i+1]["edge"]
                    bottom = FindBottomEdge(h_edge,left,right,h_edges_tb)
                    
                    start_row = h_edge["index"]
                    end_row = bottom["index"] 
                   
                    start_col = left["index"]
                    end_col = right["index"] 
                    if end_row - start_row == 0:
                        return None,None,None
                    if end_col - start_col == 0:
                        return None,None,None
                   
                    cell.append((start_row,end_row))
                    cell.append((start_col,end_col))
                    cell.append((left["x0"],right["x0"]))
                    cell.append((h_edge["top"],bottom["top"]))
                    table.append(cell)
                    a.append(left["x0"])
                    a.append(h_edge["top"])
                    b.append(right["x0"])
                    b.append(h_edge["top"])
                    c.append(right["x0"])
                    c.append(bottom["top"])
                    d.append(left["x0"])
                    d.append(bottom["top"])
                    pts.append(a)
                    pts.append(b)
                    pts.append(c)
                    pts.append(d)
                    pts1.append(pts)
            
    
    
    return pts1,table,max_dim                     

def Tableextractor(x):
    # Extracts table from edges
    isMisAligned = False
    h_edges,v_edges = PdfCleaner(x)
    
    table,h_edges,v_edges = LoopFindTable(h_edges,v_edges)

    if table:
        h_edges_tb = FindHorizontalEdges(table,h_edges)
        v_edges_tb = FindVerticalEdges(table,v_edges)

       

        isMisAligned,h_edges,v_edges = CheckMisAlignedTableX(h_edges_tb,v_edges_tb,h_edges,v_edges)
        isMisAligned,h_edges,v_edges = CheckMisAlignedTableY(h_edges_tb,v_edges_tb,h_edges,v_edges)

        if isMisAligned:
            table,h_edges,v_edges = LoopFindTable(h_edges,v_edges)
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)

    isTableExtended,h_edges,v_edges = CheckTableExtended(table,h_edges,v_edges)


        
    if isTableExtended:
        
        v_edges.sort(key=lambda e: e["x0"])
        h_edges.sort(key=lambda e: (e["top"], e["x0"]))
        table,h_edges,v_edges = FixOverlappingEdges(table,h_edges,v_edges)
        h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
        h_edges,v_edges = FindIntersections(h_edges,v_edges)
        table,h_edges,v_edges = LoopFindTable(h_edges,v_edges)
        h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
        h_edges,v_edges = FindIntersections(h_edges,v_edges)

    if not isTableConnected(table):
        h_edges,v_edges = ReconstructEdges(table,h_edges,v_edges)
        
        v_edges.sort(key=lambda e: e["x0"])
        h_edges.sort(key=lambda e: (e["top"], e["x0"]))
        table, h_edges, v_edges = FixOverlappingEdges(table,h_edges,v_edges)
        h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
        h_edges,v_edges = FindIntersections(h_edges,v_edges)
        table,h_edges,v_edges = LoopFindTable(h_edges,v_edges)
        h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
        h_edges,v_edges = FindIntersections(h_edges,v_edges)

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    if table:
        h_edges_tb = FindHorizontalEdges(table,h_edges)
        v_edges_tb = FindVerticalEdges(table,v_edges)

        
        
        v_edges_tb.sort(key=lambda e: e["x0"])
        h_edges_tb.sort(key=lambda e: e["top"])
        
        pts,tableComp,max_dim = FormCells (h_edges_tb,v_edges_tb)

        
        

        return table,h_edges_tb,v_edges_tb,h_edges,v_edges,pts,tableComp,max_dim
    return None,None,None,h_edges,v_edges,None,None,None

def group_text_by_top(text: list, tolerance: int = 2):
    # Groups text by its y position
    text_group = defaultdict(list)
    for t in text:
        top = t["top"] + (t["bottom"]-t["top"])/2
        matched = False
        for key in text_group:
            if abs(top - key) <= tolerance:
                text_group[key].append(t)
                matched = True
                break
        if not matched:
            text_group[top].append(t)
    return text_group

def vertical_overlap(a_top, a_bottom, b_top, b_bottom, tolerance=2):
    # Checks if char is overlapping with the word
    return not (
        a_bottom < b_top - tolerance or
        a_top > b_bottom + tolerance
    )


def map_chars_to_line(line, chars):
    
    # returns chars that are matching with the word
    
    matched_chars = []

    for ch in chars:
        if ch["page"] != line["page"]:
            continue

        if vertical_overlap(
            ch["top"], ch["bottom"],
            line["top"], line["bottom"]
        ):
            matched_chars.append(ch)

    # sort left → right
    matched_chars.sort(key=lambda c: c["x0"])

    

    return matched_chars

def filter_chars_by_x(chars, x0, x1):
    # returns chars between position x0 and x1
    return [
        ch for ch in chars
        if x0 <= ch["x0"] and ch["x1"] <= x1
    ]

def chars_to_text(chars, space_factor=0.5):
    # converts char to word
    if not chars:
        return ""

    text = ""
    prev = None

    # estimate average character width
    avg_width = sum(c["width"] for c in chars) / len(chars)
    space_threshold = avg_width * space_factor

    for ch in chars:
        if prev:
            gap = ch["x0"] - prev["x1"]
            if gap > space_threshold:
                text += " "

        text += ch["char"]
        prev = ch

    return text.strip()


def FindTextinCells(table,textGrouped_o,chars_p):
    # returns text in each cell of table formed
    textCells = []
    last = []
    textGrouped = copy.deepcopy(textGrouped_o)
    for i, tb in enumerate(table):
        textCell = []
        x0 , x1 = table[i][2]
        top, bottom = table[i][3]
        start_row , end_row = table[i][0]
        start_col, end_col = table[i][1]
        for t in textGrouped:
            if top <= t <= bottom:
                to_remove = []
                for j in textGrouped[t]:
                   if x0 <= j["x0"] + (j["x1"]-j["x0"])/2 <= x1 and x0 <= j["x0"] <= x1 and x0 <= j["x1"] <= x1:
                       textCell.append(j["text"])
                       textCell.append(end_row-start_row)
                       textCell.append(end_col-start_col)
                       textCell.append(j["top"]+(j["bottom"]-j["top"])/2)
                       textCell.append(j["x0"] + (j["x1"]-j["x0"])/2)
                       to_remove.append(j)
                      
                       
                   elif x0 <= j["x0"] <= x1:
                       
                       matched_chars = map_chars_to_line(j,chars_p)
                       truncated_chars = filter_chars_by_x(matched_chars, x0, x1)
                       if truncated_chars:
                            last_x1 = truncated_chars[-1]["x1"]
                            last_x0 = truncated_chars[0]["x0"]
                       truncated_text = chars_to_text(truncated_chars)  
                       textCell.append(truncated_text)
                       textCell.append(end_row-start_row)
                       textCell.append(end_col-start_col)
                       textCell.append(j["top"]+(j["bottom"]-j["top"])/2)
                       textCell.append(last_x0 + (last_x1-last_x0)/2)

                   elif  x0 <= j["x0"] + (j["x1"]-j["x0"])/2 <= x1:
                       matched_chars = map_chars_to_line(j,chars_p)
                       truncated_chars = filter_chars_by_x(matched_chars, x0, x1)
                       if truncated_chars:
                            last_x1 = truncated_chars[-1]["x1"]
                            last_x0 = truncated_chars[0]["x0"]
                       truncated_text = chars_to_text(truncated_chars)  
                       textCell.append(truncated_text)
                       textCell.append(end_row-start_row)
                       textCell.append(end_col-start_col)
                       textCell.append(j["top"]+(j["bottom"]-j["top"])/2)
                       textCell.append(last_x0 + (last_x1-last_x0)/2)      
                   
                       
                   elif x0 <= j["x1"] <= x1:
                       
                        matched_chars = map_chars_to_line(j,chars_p)
                        truncated_chars = filter_chars_by_x(matched_chars, x0, x1)
                        if truncated_chars:
                            last_x1 = truncated_chars[-1]["x1"]
                            last_x0 = truncated_chars[0]["x0"]
                        truncated_text = chars_to_text(truncated_chars) 

                        textCell.append(truncated_text)
                        textCell.append(end_row-start_row)
                        textCell.append(end_col-start_col)
                        textCell.append(j["top"]+(j["bottom"]-j["top"])/2)
                        textCell.append(last_x0 + (last_x1-last_x0)/2)
                        to_remove.append(j)
                        
                   
                        

                   

                for k in to_remove:
                    textGrouped[t].remove(k)
        
        if textCell == []:
            textCell.append("")
            textCell.append(end_row-start_row)
            textCell.append(end_col-start_col)
        textCells.append(textCell)
        
        


    return textCells,textGrouped   

def FindTextClusters(textCells,max_dim):
    # returns text of each row of table formed
    maxRowIndex , maxColumnIndex = max_dim
  
    textRows = []
    totalCol = 0
   
    total = 0
    offset = {}
    start = 0
    end = 0
    isOffset = False
    colsCheck = []
    row = 1
    addT = 0
    col = 0
    isrowStart = False
    for textCell in textCells:
       
        rows = textCell[1]
        columns = textCell[2]
        totalCol = columns + totalCol
        
        end = end + 1
        if columns == 1:
            col = col+1
        else:
            col = col+columns
        
         
        if isrowStart == True:
            for i in range(1,maxColumnIndex+1):
                
                if i in offset:
                    for l in offset[i]:
                        (r,c,m) = l
                       
                        addT = c + addT
                    
            isrowStart = False
       
        
        if rows > 1:
           
            if columns == 1:
                colCheck = totalCol
            elif columns > 1:
                colCheck = totalCol - columns
            colsCheck.append(colCheck)
            if colCheck not in offset:
                offset[colCheck] = []

            offset[colCheck].append((rows - 1, columns, row))
            
                
        
        if maxColumnIndex == totalCol + addT:
            
           
            # Decrease first element of tuple by 1
            for k in list(offset.keys()):
                updated_tuples = []
                for v in offset[k]:
                    if v[2] < row:
                        new_v0 = v[0] - 1
                        if new_v0 != 0:
                            updated_tuples.append((new_v0, v[1], v[2]))
                    else:
                        updated_tuples.append(v)

                if updated_tuples:
                    offset[k] = updated_tuples
                else:
                    del offset[k]
                    if k in colsCheck:
                        colsCheck.remove(k)
            totalCol = 0
            row = row + 1
            addT = 0      
            textRows.append(textCells[start:end])
            start = end
            col = 0
            isrowStart = True
       
    return textRows


def FindDataFrame(textClusters):
    # returns textrows with text and centered x and y positions
    textRows = []
    for cluster in textClusters:
        row = []
        for group in cluster:
            cell = []
            for i in range(0, len(group), 5):
                if i + 5 <= len(group):
                    text, rspan, cspan, top, center = group[i:i+5]
                    cell.append((str(text), round(top, 1), round(center, 1)))
                else:
                    pass
            row.append(cell)
        textRows.append(row)
    return textRows


def ModifyDataFrame(textDataFrame):
    """Arrange and join cell texts based on their positions.
    If x positions differ, separate by space.
    If y positions differ, separate by newline.
    """
    processed_rows = []
    tolerance = 2

    for row in textDataFrame:
        new_row = []
        for cell in row:
            # Sort by y (top), then by x (center)
            sorted_cell = sorted(cell, key=lambda t: (t[1], (round(t[2] / tolerance) * tolerance)))
            
            output = []
            current_y = None
            line = []

            for i, (text, y, x) in enumerate(sorted_cell):
                y_group = round(y / tolerance) * tolerance
                if current_y is None:
                    current_y = y_group
                if y_group != current_y:
                    output.append("".join(line))
                    line = [text]
                    current_y = y_group
                else:
                    if line:
                        prev_x = sorted_cell[i - 1][2]
                        space_count = max(1, round((x - prev_x) / 25))  # adjustable spacing factor
                        line.append(" " * space_count + text)
                    else:
                        line.append(text)
            if line:
                output.append("".join(line))

            new_row.append("\n".join(output))
        processed_rows.append(new_row)

    return processed_rows


def FindMergeData(textClusters,max_dim):
   # Extract (second, third) from each cell, preserving row structure
    merge_data = {}
    offset = {}
    maxRow, maxCol = max_dim
    
    colIndex = 0
    for row in range(1,maxRow):
        colsSkip = []
        colPassed = 0
        rowIndex = row-1
        for col in range(1,maxCol):
            isCol = False
            if col == 1:
                colIndex = col-1
            
            if col in colsSkip:
                continue
           
            if row in offset:
                columns = offset[row]
               
                if col in columns:
                    isCol = True
                    continue
            
            if isCol == False:
                cell = textClusters[rowIndex][colIndex]
                rowSpan = cell[1]
                colSpan = cell[2]
                
                if not (rowSpan == 1 and colSpan == 1):
                    if row not in merge_data:
                        merge_data[row] = []
                    merge_data[row].append((col,rowSpan,colSpan))
                
                if rowSpan > 1:
                    for ii in range(row+1,rowSpan+row):
                        if ii not in offset:
                            offset[ii] = []
                        for jj in range(col,col+colSpan):
                            offset[ii].append(jj)
                if colSpan > 1:
                    for jk in range(col,col+colSpan):
                        colsSkip.append(jk)

                colIndex = colIndex+1
                
    return merge_data

def FindMergeDataModified(mergeRowsData,index1,index2):
    # updated merged cells data of tables except first table
    updated_data = {}
    for key,tuples in mergeRowsData.items():
        for t in tuples:
            new_key = key + index1-1 # increase key by index1 value
            new_tuple = (t[0] + index2-1, t[1], t[2])  # increase first value by index2 value
            updated_data.setdefault(new_key, []).append(new_tuple)

    return updated_data




   
def mergeCells(index1,index2,mergeData,ws,maxRow,maxCol):
    # merge cells and writes test value in top left cell
    for row in range(index1,maxRow+index1-1):
        if not row in mergeData:
            continue
        else:
            start_row = row
            row1 = row
            for col in range(index2,maxCol+index2-1):
                for m in mergeData[row]:
                    column, rowSpan, colSpan = m
                    col1 = col
                    
                    start_row = row1
                    if col == column:
                       start_col = col1
                       end_row = start_row + rowSpan-1
                       end_col = start_col + colSpan-1
                       if rowSpan > 1 and colSpan > 1:
                            
                            ws.cell(row = start_row, column = start_col, value = "Bhanu")
                            ws.merge_cells(start_row = start_row, start_column = start_col, end_row = end_row, end_column = end_col) 
                       elif rowSpan > 1 :
                            
                            ws.cell(row = start_row, column = start_col, value = "Bhanu")
                            ws.merge_cells(start_row = start_row, start_column = start_col, end_row = end_row, end_column = start_col)
                       elif colSpan > 1:
                            
                            ws.cell(row = start_row, column = start_col, value = "Bhanu")
                            ws.merge_cells(start_row = start_row, start_column = start_col, end_row = start_row, end_column = end_col)
    return ws
    
   
    
   


def get_merged_cell_map(ws):
    # Returns merged_cells and top_left_cells in merged cell range to write the merged data to top_left_cells 
    merged_cells = set()
    top_left_cells = set()
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_cells.add((r, c))
        top_left_cells.add((min_row, min_col))
    return merged_cells, top_left_cells   



def write_to_excel(index1,index2,wb,ws,data,mergeRowsData,max_dim):
    # writing cell data to excel using openpyxl
    maxRow,maxColumn = max_dim
    maxRow = maxRow + 1
    maxColumn = maxColumn + 1
    
    ws = mergeCells(index1,index2,mergeRowsData,ws,maxRow,maxColumn)
    values = []
    valueTo = 0
    # Write each cell and apply wrap_text where needed
    merged_cells, top_left_cells = get_merged_cell_map(ws)
    
    
        
    for row_idx in range(index1,maxRow+index1-1):
        rowData = data[row_idx-index1]
        if rowData:
            i = 0
            for col_idx in range(index2,maxColumn+index2-1):
                if i < len(rowData):
                    colValue = rowData[i]
                
                if (row_idx, col_idx) in merged_cells:
                        if (row_idx, col_idx) in top_left_cells:
                            start_row, start_col = row_idx, col_idx
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.value = colValue
                            cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                        
                            i = i + 1
                        continue
                        
                else:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = colValue
                        cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                        i = i + 1
                
    
    thin_border = Border(
        left = Side(style='thin'),
        right = Side(style='thin'),
        top = Side(style='thin'),
        bottom = Side(style='thin')
    )

    for row in ws.iter_rows(min_row=index1,
    max_row=maxRow+index1-2, min_col=index2, max_col=maxColumn+index2-2):
        for cell in row:
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                
                cell_text = str(cell.value)

                lines = cell_text.split("\n")
                max_len = max(len(line) for line in lines)

                max_length = max(max_length, max_len)

        ws.column_dimensions[col_letter].width = max_length + 2


    
def extract_pdf_edges(pdf_path: str, tolerance: float = 1.5):
    """
    Extract horizontal and vertical lines from a PDF using pdfminer,
    correctly ordered from top-to-bottom (y descending) and left-to-right (x ascending).
    """
    h_edges: List[Dict] = []
    v_edges: List[Dict] = []

    for pg_num, page_layout in enumerate(extract_pages(pdf_path), start=1):
        for element in page_layout:
            if isinstance(element, (LTLine, LTRect)):
                x0, y0, x1, y1 = element.x0, element.y0, element.x1, element.y1
                height = page_layout.height

                if abs(y1 - y0) < tolerance:
                    # Horizontal line
                    h_edges.append({
                        "x0": min(x0, x1),
                        "x1": max(x0, x1),
                        "bottom": height - round(min(y0, y1), 2),      
                        "top": height - round(max(y0, y1), 2),
                        "orientation": "h",
                        
                        "page": pg_num
                    })
                elif abs(x1 - x0) < tolerance:
                    # Vertical line
                    v_edges.append({
                        "x0": round(min(x0, x1), 2),
                        "x1": round(max(x0, x1), 2),
                        "bottom": height - min(y0, y1),
                        "top": height - max(y0, y1),
                        "orientation": "v",
                        "page" : pg_num
                    })

    # Sort: top-to-bottom for horizontal, left-to-right for vertical
    h_edges.sort(key=lambda e: (e["top"], e["x0"]))
    v_edges.sort(key=lambda e: e["x0"])

    return h_edges, v_edges





def IndexCalculation1(present_edge,tableFactor,base_edge_t,base_edge_v):
    # Caluclates index based on horizontal edge alignment
    tableFactor1 = tableFactor[:]
    tableFactor1.reverse()
    index1 = 0
    index2 = 0
    for table in tableFactor1:
        isHFound = False
        isVFound = False
        (h_edges_tb,v_edges_tb,max_dim) = table
        maxRow, lindex1,maxColumn, lindex2 = max_dim
        for i, h_edge in enumerate(h_edges_tb):
            if nearedges(h_edge["top"],present_edge["top"]):
               
                index2 = maxColumn + 3
                
                index1 = lindex1 + i 
                isHFound = True
                break
        if isHFound == True:
            break

    if index1 == 0 and index2 == 0:
        return -1,-1
    else:
        return index1,index2

def IndexCalculation2(present_edge,tableFactor,base_edge_t,base_edge_v):
    # Caluclates index based on vertical edge alignment
    tableFactor1 = tableFactor[:]
    tableFactor1.reverse()
    index1 = 0
    index2 = 0
    for table in tableFactor1:
        isHFound = False
        isVFound = False
        (h_edges_tb,v_edges_tb,max_dim) = table
        maxRow, lindex1,maxColumn, lindex2 = max_dim
        for j, v_edge in enumerate(v_edges_tb):
            if nearedges(v_edge["x0"],present_edge["x0"]):
                index1 = maxRow + 3
               
                index2 = lindex2 + j 
                isVFound = True
                break
        if isVFound == True:
            break

    if index1 == 0 and index2 == 0:
        return -1,-1
    else:
        return index1,index2
    
def IndexCalculation3(present_edge,tableFactor,base_edge_t,base_edge_v):
    # Calculates index if none of horizontal edges or vertical edges align
    tableFactor1 = tableFactor[:]
    tableFactor1.reverse()
    index1 = 0
    index2 = 0
    for table in tableFactor1:
        isHFound = False
        isVFound = False
        (h_edges_tb,v_edges_tb,max_dim) = table
        maxRow, lindex1,maxColumn, lindex2 = max_dim
        index1 = maxRow + 3
        if v_edges_tb[-1]["x0"] < base_edge_v["x0"]:
            index2 = lindex2 + 2
            break
        else:
            (h_edges_tb,v_edges_tb,max_dim) = tableFactor[0]
            maxRow, lindex1,maxColumn, lindex2 = max_dim
            index2 = lindex2 - 2
            break
    return index1,index2


def IndexCalculation(h_edges_tb,v_edges_tb,h_edges,v_edges,tableFactor,first):
   # Calculates row-index and column index of table to be placed in excel
   if first:
     present_left = v_edges_tb[0]
     v_edges_left = [e for e in v_edges if e["x0"] < present_left["x0"]]
     unique_x0 = []
     for e in v_edges_left:
        if e["x0"] not in unique_x0:
            unique_x0.append(e["x0"])
     spaces = len(unique_x0)
    
     index1 = 1
     index2 = 1+spaces
    
     return index1,index2

   else:
    base_edge_t = h_edges_tb[0]
    base_edge_v = v_edges_tb[0]
    for i, h_edge in enumerate(h_edges_tb):
        index1, index2 = IndexCalculation1(h_edge,tableFactor,base_edge_t,base_edge_v)
        if index1 != -1 and index2 != -1:
            return index1+i , index2
    for j, v_edge in enumerate(v_edges_tb):
        index1, index2 = IndexCalculation2(v_edge,tableFactor,base_edge_t,base_edge_v)
        if index1 != -1 and index2 != -1:
            return index1, index2+j 
        
    if index1 == -1 and index2 == -1:
        index1, index2 = IndexCalculation3(h_edges_tb[0],tableFactor,base_edge_t,base_edge_v)
        return index1, index2
        
    
def SameLine(h_edges_tb_1,h_edges_tb_2):
    """Checks whether two tables align horizontally """
    top = h_edges_tb_1[0]["top"]
    bottom = h_edges_tb_1[-1]["top"]
    for h_edge in h_edges_tb_2:
        if top <= h_edge["top"] <= bottom:
            return True
    return False




# Open the PDF
def ExcelWriter(h_edges,v_edges,text,chars_p):
    """ Exracts tables from pdf"""
    try:
        textGrouped_o = group_text_by_top(text)
        isStart = True 
        tableList = []
        tableData = []
        if isStart == True:
            i = 0
            
            table,h_edges_tb,v_edges_tb,h_edges,v_edges,pts,tableComp,max_dim = Tableextractor(h_edges+v_edges)

            if table:
                    
                    textCells,textGrouped = FindTextinCells(tableComp,textGrouped_o,chars_p)
                    
                    textClusters = FindTextClusters(textCells,max_dim)
                    
                    mergeRowsData = FindMergeData(textClusters,max_dim)
                    
                    
                    textDataFrame = FindDataFrame(textClusters)
                    
                    excelData = ModifyDataFrame(textDataFrame)
                    
                    
                    
                    
                    tableData.append((h_edges_tb,v_edges_tb,h_edges,v_edges,excelData,mergeRowsData,max_dim))
                    h_edges_t1 = [e for e in h_edges if e not in h_edges_tb]
                    v_edges_t1 = [e for e in v_edges if e not in v_edges_tb]
                    tableList.append(table)
                    isStart = False
                    
            

        
        if isStart == False:
            
            while len(h_edges_t1) != 0 and len(v_edges_t1) != 0:
                    i = i+1
                    table,h_edges_tb,v_edges_tb,h_edges,v_edges,pts,tableComp,max_dim = Tableextractor(h_edges_t1+v_edges_t1)
                    
                    if table:
                
                        textCells,textGrouped = FindTextinCells(tableComp,textGrouped,chars_p)
                
                        textClusters = FindTextClusters(textCells,max_dim)
                    
                        mergeRowsData = FindMergeData(textClusters,max_dim)
                    
                        
                        textDataFrame = FindDataFrame(textClusters)
                    
                        excelData = ModifyDataFrame(textDataFrame)
                    
                        
                        
                        tableData.append((h_edges_tb,v_edges_tb,h_edges,v_edges,excelData,mergeRowsData,max_dim))
                        h_edges_t1 = [e for e in h_edges if e not in h_edges_tb]
                        v_edges_t1 = [e for e in v_edges if e not in v_edges_tb]
                        tableList.append(table)

        return tableData,tableList,textGrouped
    except:
        return [],[],textGrouped_o
                
       
def CheckRearrangeTables(tables):
    """As the tables are extracted based on h_edge["top"]
       it will rearrange tables based on table position in pdf"""
    tablesOrdered = []
    if len(tables) == 1:
        return tables
    for i,table in enumerate(tables):
        if i < len(tables)-1:
           
            t1 = tables[i]
            t2 = tables[i+1]
            (h_edges_tb_1, v_edges_tb_1, h_edges_1,v_edges_1, excelData_1, mergeRowsData_1, max_dim_1) = t1
            (h_edges_tb_2, v_edges_tb_2, h_edges_2,v_edges_2, excelData_2, mergeRowsData_2, max_dim_2) = t2
           
            
            if h_edges_tb_1[0]["top"] < h_edges_tb_2[0]["top"] and SameLine(h_edges_tb_1,h_edges_tb_2)  :
                if v_edges_tb_2[0]["x0"] < v_edges_tb_1[0]["x0"] :
                    
                    tablesOrdered.append(t2)
                    if i == len(tables) - 2:
                        tablesOrdered.append(t1)
                    
                    
                else:
                    
                    tablesOrdered.append(t1)
                    if i == len(tables) - 2:
                        tablesOrdered.append(t2)
                
                    
                    
            else:
                tablesOrdered.append(t1)
                if i == len(tables) - 2:
                        tablesOrdered.append(t2)
            
            
                
                
    return tablesOrdered

def ArrangeTables(tables,wb,ws):
    """Arranges tables based on their positions in pdf"""
    tableFactor = []
    lindex1 = 1
    lindex2 = 1
    try:
        for i,table in enumerate(tables):
            (h_edges_tb, v_edges_tb, h_edges, v_edges, excelData, mergeRowsData, max_dim) = table
            
            if i==0:
                index1, index2 = IndexCalculation(h_edges_tb,v_edges_tb,h_edges,v_edges,tableFactor,True)
                maxRow,maxColumn = max_dim
                max_dim_1 = (maxRow+index1,index1,maxColumn+index2,index2)
            
                tableFactor.append((h_edges_tb,v_edges_tb,max_dim_1))
                mergeRowsModified = FindMergeDataModified(mergeRowsData,index1,index2)
                write_to_excel(index1,index2,wb,ws,excelData,mergeRowsModified,max_dim)
                lindex1 = index1
                lindex2 = index2
            else:
                index1, index2 = IndexCalculation(h_edges_tb,v_edges_tb,h_edges,v_edges,tableFactor,False)
                maxRow,maxColumn = max_dim
                max_dim_1 = (maxRow+index1,index1,maxColumn+index2,index2)
                
                tableFactor.append((h_edges_tb,v_edges_tb,max_dim_1))
                mergeRowsModified = FindMergeDataModified(mergeRowsData,index1,index2)
                write_to_excel(index1,index2,wb,ws,excelData,mergeRowsModified,max_dim)
                lindex1 = index1
                lindex2 = index2
    except:
        return lindex1+maxRow+1,lindex2,False

    return index1+maxRow+1,index2,True
       


def extract_lines_topwise(pdf_path):
    # Extracts words per page
    lines = []

    for page_no, page_layout in enumerate(extract_pages(pdf_path), start=1):
        for element in page_layout:
            height = page_layout.height
            if isinstance(element, LTTextContainer):
                for line in element:

                    # 🔹 HORIZONTAL TEXT
                    if isinstance(line, LTTextLine):
                        orientation = "h"

                    # 🔹 VERTICAL TEXT
                    elif isinstance(line, LTTextLineVertical):
                        orientation = "v"

                    else:
                        continue

                    text = line.get_text().strip()
                    if not text:
                        continue

                    lines.append({
                        "page": page_no,
                        "x0": line.x0,
                        "x1": line.x1,
                        "top": height - line.y1,
                        "bottom": height - line.y0,
                        "orientation": orientation,
                        "width": line.x1 - line.x0,
                        "height": line.y1 - line.y0,
                        "text": text
                    })

    return sorted(lines, key=lambda x: (x["page"], x["top"] + (x["bottom"]-x["top"]/2), x["x0"]))

def extract_characters_topwise(pdf_path):
    # Extracts Characters per page
    chars = []
    global_index = 0  # character index across pages

    for page_no, page_layout in enumerate(extract_pages(pdf_path), start=1):
        page_height = page_layout.height

        for element in page_layout:
            if not isinstance(element, LTTextContainer):
                continue

            for line in element:

                # 🔹 HORIZONTAL TEXT
                if isinstance(line, LTTextLine):
                    orientation = "h"

                # 🔹 VERTICAL TEXT
                elif isinstance(line, LTTextLineVertical):
                    orientation = "v"

                else:
                    continue

                for obj in line:
                    if not isinstance(obj, LTChar):
                        continue

                    char_text = obj.get_text()
                    if not char_text.strip():
                        global_index += 1
                        continue

                    chars.append({
                        "page": page_no,
                        "index": global_index,
                        "char": char_text,
                        "x0": obj.x0,
                        "x1": obj.x1,
                        "top": page_height - obj.y1,
                        "bottom": page_height - obj.y0,
                        "width": obj.x1 - obj.x0,
                        "height": obj.y1 - obj.y0,
                        "font": obj.fontname,
                        "size": obj.size,
                        "orientation": orientation
                    })

                    global_index += 1

    # 🔹 Top-wise sorting (same as your line logic)
    return sorted(
        chars,
        key=lambda x: (x["page"], x["top"], x["x0"])
    )


def TableWriterNormal(pdf_path,on_page_done):
    # Writes tables in pdf to excel

    h_edges, v_edges = extract_pdf_edges(pdf_path)
    
    pages = list(extract_pages(pdf_path))
    num_pages = len(pages)
   
    #text = extract_words_with_bbox(pdf_path)
    text = extract_lines_topwise(pdf_path)
    pageChars = extract_characters_topwise(pdf_path)
    pageNo = 1
    wb = Workbook()
    while(pageNo <= num_pages):
        v_edges_p = [e for e in v_edges if e["page"] == pageNo]
        h_edges_p = [e for e in h_edges if e["page"] == pageNo]
        
        text_p = [t for t in text if t["page"] == pageNo]
        text_o = copy.deepcopy(text_p)
        chars_p = [t for t in pageChars if t["page"] == pageNo]
        lindex1 = 1
        lindex2 = 1
        allText = []
        textGrouped = {}
        isFormed = True
        if pageNo == 1:
            ws = wb.active
            
            try:
                tableData,tableList,textGrouped = ExcelWriter(h_edges_p,v_edges_p,text_p,chars_p)
                
                if tableList:
                    
                    tables = CheckRearrangeTables(tableData)
                    lindex1,lindex2,isFormed= ArrangeTables(tables,wb,ws)
                    
            except:
                textGrouped = group_text_by_top(text_o)
            finally:
                if not isFormed:
                    textGrouped = group_text_by_top(text_o)
                allText = NewText(text_o,textGrouped)
                
                tablesText = TableDetectionText(allText)
                TableWriterText(tablesText,wb,ws,lindex1,lindex2)
        else:
            ws = wb.create_sheet()
            try:
                tableData,tableList,textGrouped = ExcelWriter(h_edges_p,v_edges_p,text_p,chars_p)
                
                if tableList:
                    tables = CheckRearrangeTables(tableData)
                    lindex1,lindex2,isFormed = ArrangeTables(tables,wb,ws)
            except:
                
                textGrouped = group_text_by_top(text_o)
                
            finally:
                try:
                    if not isFormed:
                        textGrouped = group_text_by_top(text_o)
                    allText = NewText(text_o,textGrouped)
                    
                    tablesText = TableDetectionText(allText)
                    TableWriterText(tablesText,wb,ws,lindex1,lindex2)
                except:
                    pass

        
       

        if on_page_done:
            on_page_done(pageNo,num_pages)

        pageNo = pageNo + 1

    return wb
            

def run(pdf_path,on_page_done):
    
    wb = TableWriterNormal(pdf_path,on_page_done)
    return wb
    
    

if __name__ == "__main__":
    
    run(pdf_path,on_page_done)

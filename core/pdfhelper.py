
# Copyright (c) 2024 Bhanusri mattey
# Licensed under the Business Source License 1.1
# See LICENSE file in the project root for details.
# Commercial use prohibited until 2030-03-01

from PIL import ImageDraw
import bisect
from collections import defaultdict
import copy
import cv2
import numpy as np
import os
import pypdfium2 as pdfium

os.environ["FLAGS_use_mkldnn"] = "0"
os.environ["FLAGS_enable_pir_api"] = "0"
from paddleocr import PaddleOCR
import paddle





from matplotlib import pyplot as plt

from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import range_boundaries
from openpyxl.styles import Border, Side
import math
from typing import Tuple, Union


#import subprocess

from core.deskew import*
from openpyxl.utils import get_column_letter

class Config:
    DEFAULT_TOLERANCE = None
    GAP_TOLERANCE = None
    EDGES_TOLERANCE = None
    EXTENSION_TOLERANCE = None
    SNAP_X_TOLERANCE = None
    SNAP_Y_TOLERANCE = None
    JOIN_X_TOLERANCE = None
    JOIN_Y_TOLERANCE = None

class Page:
    TEXT_TOLERANCE = None
    PAGE_TOLERANCE = None
class Table:
    TABLE_SNAP_X_TOLERANCE = None
    TABLE_SNAP_Y_TOLERANCE = None
    TABLE_JOIN_X_TOLERANCE = None
    TABLE_JOIN_Y_TOLERANCE = None
    EDGE_SNAP_X_TOLERANCE = None
    EDGE_SNAP_Y_TOLERANCE = None
    EDGE_JOIN_X_TOLERANCE = None
    EDGE_JOIN_Y_TOLERANCE = None


config = Config()
page = Page()
tableP = Table()

def configure(**kwargs):
    """Set module-level config safely from outside."""
    globals().update(kwargs)


import os
import shutil
import tempfile

def pdf_to_images(pdf_path, dpi=300):
    pdf = pdfium.PdfDocument(pdf_path)
    images = []

    scale = dpi / 72  # VERY IMPORTANT

    for page in pdf:
        bitmap = page.render(scale=scale)
        pil_image = bitmap.to_pil()
        images.append(pil_image)

    pdf.close()
    return images

def CheckMisAlignedTableX(h_edges_tb, v_edges_tb, h_edges,v_edges):
    # Checks if top edge left intersection is same as second top edge similarly right intersection
    right_1 = RightIntersection(h_edges_tb[0])
    right_2 = RightIntersection(h_edges_tb[1])
    right_3 = RightIntersection(h_edges_tb[-1])
    right_4 = RightIntersection(h_edges_tb[-2])
    left_1  = LeftIntersection(h_edges_tb[0])
    left_2  = LeftIntersection(h_edges_tb[1])
    left_3 = LeftIntersection(h_edges_tb[-1])
    left_4 = LeftIntersection(h_edges_tb[-2])
    width_1 = h_edges_tb[0]["x1"] - h_edges_tb[0]["x0"]
    width_2 = h_edges_tb[1]["x1"] - h_edges_tb[1]["x0"]
    width_3 = h_edges_tb[-1]["x1"] - h_edges_tb[-1]["x0"]
    width_4 = h_edges_tb[-2]["x1"] - h_edges_tb[-2]["x0"]
    isMisAligned = False
    
    if not is_same_edge(right_1, right_2) and width_1 < width_2:
        
        new_x1 = right_2["x1"]
        new_top = h_edges_tb[0]["top"]
       
        # Reflect the same update in the original h_edges
        for edge in h_edges:
            if edge is h_edges_tb[0]:  # same object reference
                edge["x1"] = new_x1
                break
        for edge in v_edges:
            if edge is right_2:  # same object reference
                edge["top"] = new_top
                break
        
        isMisAligned = True

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    if not is_same_edge(right_3,right_4) and width_3 < width_4:
        
        new_x1 = right_4["x1"]
        new_bottom = h_edges_tb[-1]["bottom"]
       
        # Reflect the same update in the original h_edges
        for edge in h_edges:
            if edge is h_edges_tb[-1]:  # same object reference
                edge["x1"] = new_x1
                break
        for edge in v_edges:
            if edge is right_4:  # same object reference
                edge["bottom"] = new_bottom
                break
        
        isMisAligned = True
    
    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    if not is_same_edge(left_1, left_2) and width_1 < width_2:
        
        new_x0 = left_2["x0"]
        new_top = h_edges_tb[0]["top"]
       
        # Reflect the same update in the original h_edges
        for edge in h_edges:
            if edge is h_edges_tb[0]:  # same object reference
                edge["x0"] = new_x0
                break
        for edge in v_edges:
            if edge is left_2:  # same object reference
                edge["top"] = new_top
                break
        
        isMisAligned = True

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    if not is_same_edge(left_3,left_4) and width_3 < width_4:
        
        new_x0 = left_4["x0"]
        new_bottom = h_edges_tb[-1]["bottom"]
       
        # Reflect the same update in the original h_edges
        for edge in h_edges:
            if edge is h_edges_tb[-1]:  # same object reference
                edge["x0"] = new_x0
                break
        for edge in v_edges:
            if edge is left_4:  # same object reference
                edge["bottom"] = new_bottom
                break
        
        isMisAligned = True
    
    return isMisAligned,h_edges,v_edges

def CheckMisAlignedTableY(h_edges_tb, v_edges_tb, h_edges,v_edges):
    # Checks if left edge top intersection is same as second left edge similarly bottom intersection
    bottom_1 = BottomIntersection(v_edges_tb[0])
    bottom_2 = BottomIntersection(v_edges_tb[1])
    bottom_3 = BottomIntersection(v_edges_tb[-1])
    bottom_4 = BottomIntersection(v_edges_tb[-2])
    top_1 = TopIntersection(v_edges_tb[0])
    top_2 = TopIntersection(v_edges_tb[1])
    top_3 = TopIntersection(v_edges_tb[-1])
    top_4 = TopIntersection(v_edges_tb[-2])
    height_1 = v_edges_tb[0]["bottom"] - v_edges_tb[0]["top"]
    height_2 = v_edges_tb[1]["bottom"] - v_edges_tb[1]["top"]
    height_3 = v_edges_tb[-1]["bottom"] - v_edges_tb[-1]["top"]
    height_4 = v_edges_tb[-2]["bottom"] - v_edges_tb[-2]["top"]
    isMisAligned = False
    
    if not is_same_edge(bottom_1, bottom_2) and height_1 < height_2:
        
        new_bottom = bottom_2["bottom"]
        new_x0 = v_edges_tb[0]["x0"]
       
        # Reflect the same update in the original h_edges
        for edge in v_edges:
            if edge is v_edges_tb[0]:  # same object reference
                edge["bottom"] = new_bottom
                break
        for edge in h_edges:
            if edge is bottom_2:  # same object reference
                edge["x0"] = new_x0
                break
        
        isMisAligned = True

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    if not is_same_edge(bottom_3,bottom_4) and height_3 < height_4:
        
        new_bottom = bottom_4["bottom"]
        new_x1 = v_edges_tb[-1]["x1"]
       
        # Reflect the same update in the original h_edges
        for edge in v_edges:
            if edge is v_edges_tb[-1]:  # same object reference
                edge["bottom"] = new_bottom
                break
        for edge in h_edges:
            if edge is bottom_4:  # same object reference
                edge["x1"] = new_x1
                break
        
        isMisAligned = True
    
    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    if not is_same_edge(top_1, top_2) and height_1 < height_2:
        
        new_top = top_2["top"]
        new_x0 = v_edges_tb[0]["top"]
       
        # Reflect the same update in the original h_edges
        for edge in v_edges:
            if edge is v_edges_tb[0]:  # same object reference
                edge["top"] = new_top
                break
        for edge in h_edges:
            if edge is top_2:  # same object reference
                edge["x0"] = new_x0
                break
        
        isMisAligned = True

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    if not is_same_edge(top_3,top_4) and height_3 < height_4:
        
        new_top = top_4["top"]
        new_x1 = v_edges_tb[-1]["x1"]
       
        # Reflect the same update in the original h_edges
        for edge in v_edges:
            if edge is v_edges_tb[-1]:  # same object reference
                edge["top"] = new_top
                break
        for edge in h_edges:
            if edge is top_4:  # same object reference
                edge["x1"] = new_x1
                break
        
        isMisAligned = True
    
    return isMisAligned,h_edges,v_edges



def deskew_box(pts):

    # deskews rotated box given by paddleocr
    """
    pts: list or array of 4 points [[x,y], ...] in clockwise order
         (top-left, top-right, bottom-right, bottom-left)

    returns:
        rotated_pts : points rotated to be straight
        aligned_box : perfectly axis-aligned rectangle
        angle_deg   : detected rotation angle (degrees)
    """

    pts = np.array(pts, dtype=float)

    # 1️⃣ compute angle from top edge (p0 -> p1)
    dx = pts[1][0] - pts[0][0]
    dy = pts[1][1] - pts[0][1]
    angle = np.arctan2(dy, dx)
    angle_deg = np.degrees(angle)

    # 2️⃣ rotation matrix (deskew → rotate back)
    cos_a = np.cos(-angle)
    sin_a = np.sin(-angle)
    R = np.array([
        [cos_a, -sin_a],
        [sin_a,  cos_a]
    ])

    # 3️⃣ rotate around center
    center = pts.mean(axis=0)
    rotated_pts = (pts - center) @ R.T + center

    # 4️⃣ axis-aligned bounding box
    xmin, ymin = rotated_pts.min(axis=0)
    xmax, ymax = rotated_pts.max(axis=0)

    aligned_box = np.array([
        [xmin, ymin],
        [xmax, ymin],
        [xmax, ymax],
        [xmin, ymax]
    ])

    return rotated_pts,aligned_box,-angle


def TextDetection(img,tolerance=None):
    # Detects text given by paddleocr and returns text grouped by y-position,xpositions
    if tolerance == None:
        tolerance = page.TEXT_TOLERANCE

    MAX_SIDE = 4000
    
    x0, top, bottom, x1 = 0,0,0,0
    allText = []

    image_bgr = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)

    h, w = image_bgr.shape[:2]
    max_side = max(h, w)

    if max_side > MAX_SIDE:
        scale = MAX_SIDE / max_side
        new_w = int(w * scale)
        new_h = int(h * scale)

        image_bgr = cv2.resize(
            image_bgr,
            (new_w, new_h),
            interpolation=cv2.INTER_AREA
    )

    #ocr = PaddleOCR(use_angle_cls = True, lang='en',use_gpu=True)
    ocr = PaddleOCR(use_angle_cls=True, lang='en',rec_batch_limit = 16)   # 👈 increase limit)

    result = ocr.ocr(image_bgr)
    
    
   
    
    for line in result:
        for box, (text, score) in line:
            pts = []
            
                
            x0 = box[0][0]
            x1 = box[1][0]
            y0 = box[0][1]
            y1 = box[2][1]

            

            
            pts.append((x0,y1))
            pts.append((x1,y1))
            pts.append((x1,y0))
            pts.append((x0,y0))
            rotated_pts, aligned_box, angle_deg = deskew_box(pts)
            x0, y0 = aligned_box[0]  # top-left
            x1, y1 = aligned_box[2]  # bottom-right
            
            
            

            w = x1-x0
            h = y1-y0
            if w > h:
                text_start_idx = 0
                text_end_idx   = len(text.strip())   # trims right padding
                box_height = h
                box_width = w
                n = len(text)
                char_w = box_width*0.9 / n
                # assume text occupies ~70–80% vertically
                

                text_height = box_height * 0.685

                texty0 = y0 + (box_height - text_height) / 1.5
                texty1 = texty0 + text_height
                
                textx0 = x0 + text_start_idx * char_w + box_width*0.08
                textx1 =  textx0 + text_end_idx * char_w

                
                text = {
                "x0": textx0,
                "x1": textx1,
                "top": texty0,
                "bottom": texty1,
                "orientation": "h",
                "width": textx1 - textx0,
                "height": texty1 - texty0,
                "text" : text,
                "charWidth":char_w
                }
                allText.append(text)
                
            elif h > w:
                text_start_idx = 0
                text_end_idx   = len(text.strip())   # trims right padding
                box_width = w
                box_height = h
                n = len(text)
                char_w = box_height*0.9 / n
                # assume text occupies ~70–80% vertically
                
                

                text_width = box_width * 0.685

                textx0 = x0 + (box_width - text_width) / 1.5
                textx1 = textx0 + text_width
                
                texty0 = y0 + text_start_idx * char_w + box_height*0.08
                texty1 = texty0 + text_end_idx   * char_w
                text = {
                "x0": textx0,
                "x1": textx1,
                "top": texty0,
                "bottom": texty1,
                "orientation": "v",
                "width": w,
                "height": texty1 - texty0,
                "text" : text,
                "charWidth":char_w
                }
                allText.append(text)
            else:
                text_start_idx = 0
                text_end_idx   = len(text.strip())   # trims right padding
                box_height = h
                box_width = w
                n = len(text)
                char_w = box_width*0.9 / n
                # assume text occupies ~70–80% vertically
                

                text_height = box_height * 0.685

                texty0 = y0 + (box_height - text_height) / 1.5
                texty1 = texty0 + text_height

                
                textx0 = x0 + text_start_idx * char_w + box_width*0.10
                textx1 =  textx0 + text_end_idx * char_w
                text = {
                "x0": textx0,
                "x1": textx1,
                "top": texty0,
                "bottom": texty1,
                "orientation": "h",
                "width": textx1 - textx0,
                "height": texty1 - texty0,
                "text" : text,
                "charWidth":char_w
                }
                allText.append(text)


    def SnapRows_top(text,textsnapped):
        for t in text:
            for top in textsnapped:
                if nearedges(t["top"],top,10):
                    t["top"] = top
                    t["bottom"] = top
    
    def SnapRows_x0(text,textsnapped):
        for t in text:
            for x0 in textsnapped:
                if nearedges(t["x0"],x0,10):
                    t["x0"] = x0
                   
    def SnapRows_x1(text,textsnapped):
        for t in text:
            for x1 in textsnapped:
                if nearedges(t["x1"],x1,10):
                    t["x1"] = x1

    
    text_dict = EdgeIndex(allText)
    textsnapped_top = snap_close_values(text_dict.top_keys,10)
    SnapRows_top(allText,textsnapped_top)
    textsnapped_x0 = snap_close_values(text_dict.x0_keys,10)
    SnapRows_x0(allText,textsnapped_x0)
    textsnapped_x1 = snap_close_values(text_dict.x1_keys,10)
    SnapRows_x1(allText,textsnapped_x1)
    rawText = copy.deepcopy(allText)
    allText = sorted(allText, key=lambda e: e["top"])
    
    def group_text_by_top(text: list, tolerance: int = 10):
        text_group = defaultdict(list)
        for t in text:
            top = t["top"] + (t["bottom"]-t["top"])/2
            matched = False
            for key in text_group:
                if abs(top - key) <= tolerance:
                    text_group[key].append(t)
                    matched = True
                    break
            if not matched:
                text_group[top].append(t)
        return text_group

    textGroupedY = group_text_by_top(allText)

    allText = sorted(allText, key=lambda e: e["x0"])
    

    def group_text_by_xm(text: list, tolerance: int = 10):
        text_group = defaultdict(list)
        for t in text:
            top =  t["x0"] + (t["x1"] - t["x0"])/2
            matched = False
            for key in text_group:
                if abs(top - key) <= tolerance:
                    text_group[key].append(t)
                    matched = True
                    break
            if not matched:
                text_group[top].append(t)
        return text_group
    
    
    
    def group_text_by_xs(text: list, tolerance: int = 10):
        text_group = defaultdict(list)
        for t in text:
            top =  t["x0"]
            matched = False
            for key in text_group:
                if abs(top - key) <= tolerance:
                    text_group[key].append(t)
                    matched = True
                    break
            if not matched:
                text_group[top].append(t)
        return text_group
    
    
    
    def group_text_by_xe(text: list, tolerance: int = 10):
        text_group = defaultdict(list)
        for t in text:
            top =  t["x1"]
            matched = False
            for key in text_group:
                if abs(top - key) <= tolerance:
                    text_group[key].append(t)
                    matched = True
                    break
            if not matched:
                text_group[top].append(t)
        return text_group

    textGroupedXm = group_text_by_xm(allText[:])
    allText = sorted(allText, key=lambda e: e["x0"])
    textGroupedXs = group_text_by_xs(allText[:])
    allText = sorted(allText, key=lambda e: e["x0"])
    textGroupedXe = group_text_by_xe(allText[:])
    allText = sorted(allText, key=lambda e: e["x0"])
    
    return rawText,textGroupedXm,textGroupedXs,textGroupedXe,textGroupedY,image_bgr


   

# Show using matplotlib

# Step 3: Detect horizontal and vertical lines


def remove_page_border_edges(h_edges, v_edges, page_width, page_height, tolerance=None):
    # Removes page border edges detected by opencv

    if tolerance == None:
        tolerance = page.PAGE_TOLERANCE
    # Remove vertical edges near left or right page edges
    filtered_v_edges = [
        e for e in v_edges if tolerance < e["x0"] < (page_width - tolerance)
    ]
    
    # Remove horizontal edges near top or bottom page edges
    filtered_h_edges = [
        e for e in h_edges if tolerance < e["top"] < (page_height - tolerance)
    ]
    
    
    return filtered_h_edges, filtered_v_edges

def snap_close_values(values, tolerance=None):
    # Snaps edges with values within tolerance
    if tolerance == None:
        tolerance = config.DEFAULT_TOLERANCE
    values = sorted(values)
    snapped = []
    group = []

    for i, val in enumerate(values):
        if not group:
            group.append(val)
        elif abs(val - group[-1]) <= tolerance:
            group.append(val)
        else:
            avg = sum(group) / len(group)
            snapped.append(avg)
            group = [val]

    # handle last group
    if group:
        avg = sum(group) / len(group)
        snapped.append(avg)

    return snapped

class EdgeIndex:
    """EdgeIndex provides efficient spatial lookup for edge objects based on
    their geometric properties (x0, x1, top, bottom)."""
    def __init__(self, edges):
        # Store lists of edges per value
        self.by_x1 = defaultdict(list)
        self.by_top = defaultdict(list)
        self.by_x0 = defaultdict(list)
        self.by_bottom = defaultdict(list)

        for edge in edges:
            self.by_x1[edge["x1"]].append(edge)
            self.by_top[edge["top"]].append(edge)
            self.by_x0[edge["x0"]].append(edge)
            self.by_bottom[edge["bottom"]].append(edge)

        # Sorted unique keys for bisect
        self.x1_keys = sorted(self.by_x1.keys())
        self.top_keys = sorted(self.by_top.keys())
        self.x0_keys = sorted(self.by_x0.keys())
        self.bottom_keys = sorted(self.by_bottom.keys())

    def get_closest_key(self, val, sorted_keys, tolerance=None):
        if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    # Use bisect to find the insertion point
        idx = bisect.bisect_left(sorted_keys, val)

        candidates = []
        if idx > 0:
            candidates.append(sorted_keys[idx - 1])
        if idx < len(sorted_keys):
            candidates.append(sorted_keys[idx])

    # Find the closest candidate within tolerance
        closest = min(
        (key for key in candidates if abs(key - val) <= tolerance),
        key=lambda k: abs(k - val),
        default=None
        )
        return closest
    # These now return a LIST of matching edges
    def find_by_x1(self, x1_val, tolerance=None):
        if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
        key = self.get_closest_key(x1_val, self.x1_keys, tolerance)
        return self.by_x1.get(key, [])

    def find_by_x0(self, x0_val, tolerance=None):
        if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
        key = self.get_closest_key(x0_val, self.x0_keys, tolerance)
        return self.by_x0.get(key, [])

    def find_by_top(self, top_val, tolerance=None):
        if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
        key = self.get_closest_key(top_val, self.top_keys, tolerance)
        return self.by_top.get(key, [])

    def find_by_bottom(self, bottom_val, tolerance=None):
        if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
        key = self.get_closest_key(bottom_val, self.bottom_keys, tolerance)
        return self.by_bottom.get(key, [])
    
    def find_by_x1_and_bottom(self, x1_val, bottom_val, x1_tol=None, bottom_tol=None):
        if x1_tol == None:
            tolerance = config.DEFAULT_TOLERANCE
        if bottom_tol == None:
            tolerance = config.DEFAULT_TOLERANCE
        x1_key = self.get_closest_key(x1_val, self.x1_keys, x1_tol)
        x1_matches = self.by_x1.get(x1_key, [])

    # Now filter x1_matches by top tolerance manually
        filtered = [
            edge for edge in x1_matches
            if abs(edge["bottom"] - bottom_val) <= bottom_tol
        ]

        return filtered
    
    def find_by_x0_and_top(self, x0_val, top_val, x0_tol=None, top_tol=None):
        if x0_tol == None:
            tolerance = config.DEFAULT_TOLERANCE
        if top_tol == None:
            tolerance = config.DEFAULT_TOLERANCE
        x0_key = self.get_closest_key(x0_val, self.x0_keys, x0_tol)
        x0_matches = self.by_x0.get(x0_key, [])

    # Now filter x1_matches by top tolerance manually
        filtered = [
            edge for edge in x0_matches
            if abs(edge["top"] - top_val) <= top_tol
        ]

        return filtered
    
def Intersection(h_edge, v_edge,tolerance):
    # Checks whether Horizontal edge and Vertical edge are intersecting
    h0, h1 = h_edge["x0"], h_edge["x1"]
    hy = h_edge["top"]  # or bottom — they should be the same for horizontal
    vx = v_edge["x0"]   # or x1 — same for vertical
    vtop, vbottom = v_edge["top"], v_edge["bottom"]
   
    # Check if vx lies between h0 and h1
    # AND hy lies between vtop and vbottom
    if (h0 <= vx <= h1 or nearedges(h1,vx,tolerance) or nearedges(h0,vx,tolerance)) and (nearedges(hy,vbottom,tolerance) or nearedges(hy,vtop,tolerance) or vtop <= hy <= vbottom):
        
        return True
    return False
    
def nearedges(a,b,tolerance):
    # Checks if both values are within tolerance
    if abs(a-b) <= tolerance:
        return True
    return False

def SnapHEdges(h_edges,hsnappedx0,hsnappedx1,hsnappedtop,tolerance=None):
     # Snaps all Horizontal Edges
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    for edge in h_edges:
        for x0 in hsnappedx0:
            if nearedges(edge["x0"],x0,tolerance):
                edge["x0"] = x0
        for x1 in hsnappedx1:
            if nearedges(edge["x1"],x1,tolerance):
                edge["x1"] = x1
        for top in hsnappedtop:
            if nearedges(edge["top"],top,tolerance):
                edge["top"] = top
                edge["bottom"] = top

def SnapHEdge(h_edges,hsnappedtop,tolerance=None):
    # Snaps Horizontal Edge
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    for edge in h_edges:
        for top in hsnappedtop:
            if nearedges(edge["top"],top,tolerance):
                edge["top"] = top

def SnapVEdges(v_edges,vsnappedx0,vsnappedtop,vsnappedbottom,tolerance=None):
    # Snaps all Vertical Edges
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    for edge in v_edges:
        for x0 in vsnappedx0:
            if nearedges(edge["x0"],x0,tolerance):
                edge["x0"] = x0
                edge["x1"] = x0
        for top in vsnappedtop:
            if nearedges(edge["top"],top,tolerance):
                edge["top"] = top
        for bottom in vsnappedbottom:
            if nearedges(edge["bottom"],bottom,tolerance):
                edge["bottom"] = bottom

def SnapVEdge(v_edges,vsnappedx0,tolerance=None):
    # Snaps Vertical Edge
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    for edge in v_edges:
        for x0 in vsnappedx0:
            if nearedges(edge["x0"],x0,tolerance):
                edge["x0"] = x0

def FindIntersections(h_edges,v_edges,tolerance=None):
    # Finds intersections of Horizontal edges and Vertical edges
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    for h_edge in h_edges:
        for v_edge in v_edges:
            if Intersection(h_edge,v_edge,tolerance):
                x = v_edge["x0"]
                y = h_edge["top"]
                h_edge.setdefault("intersections", []).append({
                    "edge": v_edge,
                    "point": (x, y)
                })
                v_edge.setdefault("intersections", []).append({
                    "edge": h_edge,
                    "point": (x, y)
                })
    return h_edges,v_edges



def RemoveIntersections(h_edges, v_edges):
     # Removes intersections of edges
    for h_edge in h_edges:
        if "intersections" in h_edge:
            del h_edge["intersections"]
    for v_edge in v_edges:
        if "intersections" in v_edge:
            del v_edge["intersections"]
    return h_edges, v_edges

def RemoveIndexing(h_edges, v_edges):
    # Resets indexes of edges
    for h_edge in h_edges:
        if "index" in h_edge:
            del h_edge["index"]
    for v_edge in v_edges:
        if "index" in v_edge:
            del v_edge["index"]
    return h_edges, v_edges


def LeftIntersection(h_edge):
    # Finds Left Intersection of Horizontal edge
    dist = 0
    minDist = float('inf')
    minEdge = None
    v_int_List = h_edge["intersections"]
    #v_int_List = h_edge.get("intersections", [])
    for v_int in v_int_List:
        dist =  abs(h_edge["x0"]- v_int["edge"]["x0"])
        if dist < minDist:
            minDist = dist
            minEdge = v_int["edge"]
    return minEdge


def RightIntersection(h_edge):
    # Finds Right Intersection of Horizontal edge
    dist = 0
    minDist = float('inf')
    minEdge = None
    v_int_List = h_edge.get("intersections", [])
    for v_int in v_int_List:
        dist =  abs(h_edge["x1"]- v_int["edge"]["x1"])
        if dist < minDist:
            minDist = dist
            minEdge = v_int["edge"]

    return minEdge

def TopIntersection(v_edge):
    # Finds Top Intersection of Vertical Edge
    dist = 0
    minDist = float('inf')
    minEdge = None
    h_int_List = v_edge["intersections"]
    for h_int in h_int_List:
        dist =  abs(v_edge["top"]- h_int["edge"]["top"])
        if dist < minDist:
            minDist = dist
            minEdge = h_int["edge"]
    return minEdge

def BottomIntersection(v_edge):
    # Finds Bottom Intersection of Vertical Edge
    dist = 0
    minDist = float('inf')
    minEdge = None
    h_int_List = v_edge["intersections"]
    if h_int_List is not None:
        for h_int in h_int_List:
            dist =  abs(v_edge["bottom"]- h_int["edge"]["bottom"])
            if dist < minDist:
                minDist = dist
                minEdge = h_int["edge"]
    return minEdge

def FindHorizontalEdges(table,h_edges,tolerance1,tolerance2=None):
    # Finds horizontal edges within boundary of table formed
    
    if tolerance2 == None:
        tolerance2 = config.DEFAULT_TOLERANCE
    
    h_edges_tb = []
    left_edge = table[0]
    right_edge = table[1]
    leftPos = table[0]["x0"]
    rightPos = table[1]["x1"]
    topPos = table[2]["top"]
    bottomPos = table[3]["bottom"]
    for h_edge in h_edges:
        if (
            (nearedges(h_edge["x0"], leftPos,tolerance1) or leftPos <= h_edge["x0"] <= rightPos) and
            (nearedges(h_edge["x1"], rightPos,tolerance1) or leftPos <= h_edge["x1"] <= rightPos) and
            (nearedges(h_edge["top"], topPos,tolerance1) or topPos <= h_edge["top"] <= bottomPos) and
            (nearedges(h_edge["bottom"], bottomPos,tolerance1) or topPos <= h_edge["bottom"] <= bottomPos)
        ):
            h_edges_tb.append(h_edge)

    for h_edge in h_edges:
        if(Intersection(h_edge,left_edge,tolerance2) and Intersection(h_edge,right_edge,tolerance2) and topPos <= h_edge["top"] <= bottomPos):
            if h_edge not in h_edges_tb:
                h_edges_tb.append(h_edge)

    for h_edge in h_edges:
        if(leftPos <= h_edge["x0"] <= rightPos and Intersection(h_edge,right_edge,tolerance2) and topPos <= h_edge["top"] <= bottomPos):
            if h_edge not in h_edges_tb:
                h_edges_tb.append(h_edge)

    for h_edge in h_edges:
        if(Intersection(h_edge,left_edge,tolerance2) and leftPos <= h_edge["x1"] <= rightPos and topPos <= h_edge["top"] <= bottomPos):
            if h_edge not in h_edges_tb:
                h_edges_tb.append(h_edge)


    return h_edges_tb

def FindVerticalEdges(table,v_edges,tolerance1,tolerance2=None):
    # Finds vertical edges within boundary of table formed
    if tolerance2 == None:
            tolerance2 = config.DEFAULT_TOLERANCE
    v_edges_tb = []
    top_edge = table[2]
    bottom_edge = table[3]
    leftPos = table[0]["x0"]
    rightPos = table[1]["x1"]
    topPos = table[2]["top"]
    bottomPos = table[3]["bottom"]
    for v_edge in v_edges:
        if (
            (nearedges(v_edge["x0"], leftPos,tolerance1) or leftPos <= v_edge["x0"] <= rightPos) and
            (nearedges(v_edge["x1"], rightPos,tolerance1) or leftPos <= v_edge["x1"] <= rightPos) and
            (nearedges(v_edge["top"], topPos,tolerance1) or  topPos <= v_edge["top"] <= bottomPos) and
            (nearedges(v_edge["bottom"], bottomPos,tolerance1) or topPos <= v_edge["bottom"] <= bottomPos )
        ):
            v_edges_tb.append(v_edge)

    for v_edge in v_edges:
        if(Intersection(top_edge,v_edge,tolerance2) and Intersection(bottom_edge,v_edge,tolerance2) and leftPos <= v_edge["x0"] <= rightPos):
            if v_edge not in v_edges_tb:
                v_edges_tb.append(v_edge)

    for v_edge in v_edges:
        if topPos <= v_edge["top"] <= bottomPos and Intersection(bottom_edge,v_edge,tolerance2) and leftPos <= v_edge["x0"] <= rightPos:
            if v_edge not in v_edges_tb:
                v_edges_tb.append(v_edge)

    for v_edge in v_edges:
        if Intersection(top_edge,v_edge,tolerance2) and topPos <= v_edge["bottom"] <= bottomPos and leftPos <= v_edge["x0"] <= rightPos:
            if v_edge not in v_edges_tb:
                v_edges_tb.append(v_edge)
    return v_edges_tb






def FilterLooseEdges(h_edges, v_edges):
    # Removes edges with less than 2 intersections
    v_edges = [v for v in v_edges if len(v.get("intersections", [])) >= 2]
    h_edges = [h for h in h_edges if len(h.get("intersections", [])) >= 2]
    return h_edges, v_edges 





def ModifyHEdge(edge1,edge2):
    # Modifies Horizontal edge
    edge1["x1"] = edge2["x1"]

def ModifyVEdge(edge1,edge2):
    # Modifies Vertical edge
    edge1["bottom"] = edge2["bottom"]

def FixOverlappingEdges(table,h_edges,v_edges,tolerance=None):
    # Fixes overlapping edges in boundary of table
    if tolerance == None:
        tolerance = config.DEFAULT_TOLERANCE
    v_edges = [
    v for v in v_edges
    if not (
        (nearedges(v["x0"],table[0]["x0"],tolerance)) 
        and
        ((nearedges(v["top"],table[0]["top"],tolerance) and table[0]["top"] < v["bottom"] < table[0]["bottom"])
        or
        (nearedges(v["bottom"],table[0]["bottom"],tolerance) and table[0]["top"] < v["top"] < table[0]["bottom"]))
        
        
    )
    
    ]

    v_edges = [
    v for v in v_edges
    if not (
        (nearedges(v["x0"],table[1]["x0"],tolerance)) 
        and
        ((nearedges(v["top"],table[1]["top"],tolerance) and table[1]["top"] < v["bottom"] < table[1]["bottom"])
        or
        (nearedges(v["bottom"],table[1]["bottom"],tolerance) and table[1]["top"] < v["top"] < table[1]["bottom"]))
    )
    
    ]

    
    v_edges = [
    v for v in v_edges
    if not (
        table[0]["top"] < v["top"] < table[0]["bottom"] and  table[0]["top"] < v["bottom"] < table[0]["bottom"] and nearedges(v["x0"],table[0]["x0"],tolerance)
        or
        table[1]["top"] < v["top"] < table[1]["bottom"] and  table[1]["top"] < v["bottom"] < table[1]["bottom"] and nearedges(v["x0"],table[1]["x0"],tolerance)
    )
    
    ]
   

    h_edges = [
    h for h in h_edges
    if not (
        (nearedges(h["top"],table[2]["top"],tolerance)) 
        and
        ((nearedges(h["x0"],table[2]["x0"],tolerance) and table[2]["x0"] < h["x1"] < table[2]["x1"])
        or
        (nearedges(h["x1"],table[2]["x1"],tolerance) and table[2]["x0"] < h["x0"] < table[2]["x1"]))
    )
    ]

    h_edges = [
    h for h in h_edges
    if not (
        (nearedges(h["top"],table[3]["top"],tolerance)) 
        and
        ((nearedges(h["x0"],table[3]["x0"],tolerance) and table[3]["x0"] < h["x1"] < table[3]["x1"])
        or
        (nearedges(h["x1"],table[3]["x1"],tolerance) and table[3]["x0"] < h["x0"] < table[3]["x1"]))
    )
    ]

    h_edges = [
    h for h in h_edges
    if not (
        table[2]["x0"] < h["x0"] < table[2]["x1"] and  table[2]["x0"] < h["x1"] < table[2]["x1"] and nearedges(h["top"],table[2]["top"],tolerance)
        or
        table[3]["x0"] < h["x0"] < table[3]["x1"] and  table[3]["x0"] < h["x1"] < table[3]["x1"] and nearedges(h["top"],table[3]["top"],tolerance)
    )
    
    ]

    return table,h_edges,v_edges



def FindTable(index,h_edges,v_edges,tolerance=None):
    # Finds table with respect to top horizontal edge
    if tolerance == None:
        tolerance = config.DEFAULT_TOLERANCE

    ed = []
    
    top_edge = h_edges[index]
    left_edge_t = LeftIntersection(top_edge)
    right_edge_t = RightIntersection(top_edge)
    bottom_edge_l = BottomIntersection(left_edge_t)
    bottom_edge_r = BottomIntersection(right_edge_t)
    

    if bottom_edge_r != bottom_edge_l:
        bottom_edge = bottom_edge_l if bottom_edge_l["top"] > bottom_edge_r["top"] else bottom_edge_r
    if bottom_edge_r != bottom_edge_l:
        if nearedges(bottom_edge_l["top"],bottom_edge_r["top"],tolerance):
            ModifyHEdge(bottom_edge_l,bottom_edge_r)
            bottom_edge = bottom_edge_l
    else:
        bottom_edge = bottom_edge_r

   

    left_edge_b = LeftIntersection(bottom_edge)
    left_edge_t = LeftIntersection(top_edge)
    
    if left_edge_t != left_edge_b:
        left_edge = left_edge_t if left_edge_t["x0"] < left_edge_b["x0"] else left_edge_b
    if left_edge_t != left_edge_b:
        if nearedges(left_edge_t["x0"],left_edge_b["x0"],tolerance):
            
            ModifyVEdge(left_edge_t,left_edge_b)
            left_edge = left_edge_t
    else:
        left_edge = left_edge_b

    


    right_edge_b = RightIntersection(bottom_edge)
    right_edge_t = RightIntersection(top_edge)

   

    if right_edge_t != right_edge_b:
        right_edge = right_edge_t if right_edge_t["x0"] > right_edge_b["x0"] else right_edge_b
    if right_edge_t != right_edge_b:
        if nearedges(right_edge_t["x0"],right_edge_b["x0"],tolerance):
            
            ModifyVEdge(right_edge_t,right_edge_b)
            right_edge = right_edge_t
    else:
        right_edge = right_edge_b

    ed.append(left_edge)
    ed.append(right_edge)
    ed.append(top_edge)  
    ed.append(bottom_edge)

    ed,h_edges,v_edges = FixOverlappingEdges(ed,h_edges,v_edges)

    return ed,h_edges,v_edges



def isTableConnected(tb,tolerance = None):
    # Checks if the boundary of table is connected or not
    if tolerance == None:
            tolerance = config.EXTENSION_TOLERANCE
    if (Intersection(tb[2],tb[0],tolerance) and nearedges(tb[2]["top"],tb[0]["top"],tolerance))  and (Intersection(tb[2],tb[1],tolerance) and nearedges(tb[2]["top"],tb[1]["top"],tolerance))  and (Intersection(tb[3],tb[0],tolerance) and nearedges(tb[3]["bottom"],tb[0]["bottom"],tolerance))  and (Intersection(tb[3],tb[1],tolerance) and nearedges(tb[3]["bottom"],tb[1]["bottom"],tolerance)):
        return True
    return False





def snap_edges(edges,x_tolerance,y_tolerance):
    # Snaps horizontal edges within x_tolerance and vertical edges within y_tolerance
    v_edges = [e for e in edges if e["orientation"] == "v"]
    h_edges = [e for e in edges if e["orientation"] == "h"]
    
    h_edges_dict = EdgeIndex(h_edges)

    # Snapping close values of Horizontal edges
    hsnappedtop = snap_close_values(h_edges_dict.top_keys,y_tolerance)
    SnapHEdge(h_edges,hsnappedtop) 

    v_edges_dict = EdgeIndex(v_edges)

    # Snapping close values of Vertical edges
    vsnappedx0 = snap_close_values(v_edges_dict.x0_keys,x_tolerance)
    SnapVEdge(v_edges,vsnappedx0)

    return h_edges+v_edges
   


def ResizeVEdge(last,edge):
    # Resizes Vertical Edge
    last["bottom"] = edge["bottom"]
    return last

def ResizeHEdge(last,edge):
    # Resizes Horizontal Edge
    last["x1"] = edge["x1"]
    return last

def join_edge_group(edges, orientation, threshold):
    """
    Merges nearby edges (lines) on a given axis.
    Example: horizontal lines (axis="y"), vertical lines (axis="x").
    """
   

    if orientation == "h":
        group_key = lambda e: (round(e["top"], 1))
        sort_key = lambda e: e["x0"]
    else:
        group_key = lambda e: (round(e["x0"], 1))
        sort_key = lambda e: e["top"]

    from collections import defaultdict

    grouped = defaultdict(list)
    for edge in edges:
        grouped[group_key(edge)].append(edge)

    merged = []
    for group in grouped.values():
        group = list(sorted(group, key=sort_key))
        
        merged_line = [group[0]]
        
        for edge in group[1:]:
            last = merged_line[-1]
            if orientation == "v":
                if edge["top"] - last["bottom"] < threshold:
                    if edge["bottom"] > last["bottom"]:
                        
                        merged_line[-1] = ResizeVEdge(last,edge)
                        
                else:
                    merged_line.append(edge)
               

                
            else:
                if edge["x0"] - last["x1"] < threshold:
                    if edge["x1"] > last["x1"]:
                        merged_line[-1] = ResizeHEdge(last,edge)
                       
                else :
                    merged_line.append(edge)
                
                
        merged.append(merged_line)
                                     
    flat = [item for sublist in merged for item in sublist]
    return flat



def merge_edges(edges,snap_x_tolerance,snap_y_tolerance,join_x_tolerance,join_y_tolerance):
    """
    Using the `snap_edges` and `join_edge_group` methods above,
    merge a list of edges into a more "seamless" list.
    """

    

    if snap_x_tolerance > 0 or snap_y_tolerance > 0:
        edges = snap_edges(edges, snap_x_tolerance, snap_y_tolerance)

    v_edges = [e for e in edges if e["orientation"] == "v"]
    h_edges = [e for e in edges if e["orientation"] == "h"]

    v_edges = join_edge_group(v_edges,"v",join_y_tolerance)
    h_edges = join_edge_group(h_edges,"h",join_x_tolerance)
    
    
    
    return h_edges+v_edges



def SnapEdges(h_edges,v_edges):
    # Snaps all Horizontal and Vertical Edges
    h_edges_dict = EdgeIndex(h_edges)

    # Snapping close values of Horizontal edges
    hsnappedx0 = snap_close_values(h_edges_dict.x0_keys)
    hsnappedx1 = snap_close_values(h_edges_dict.x1_keys)
    hsnappedtop = snap_close_values(h_edges_dict.top_keys)

    SnapHEdges(h_edges,hsnappedx0,hsnappedx1,hsnappedtop)
    
    v_edges_dict = EdgeIndex(v_edges)

    # Snapping close values of Vertical edges
    vsnappedx0 = snap_close_values(v_edges_dict.x0_keys)
    vsnappedtop = snap_close_values(v_edges_dict.top_keys)
    vsnappedbottom = snap_close_values(v_edges_dict.bottom_keys)

    SnapVEdges(v_edges,vsnappedx0,vsnappedtop,vsnappedbottom)
    return h_edges,v_edges

def PdfCleaner(img,x,tolerance1=None,tolerance2=None,tolerance3=None,tolerance4=None):
    # Merges edges within tolernace and filters edges with less than 2 intersections
    if tolerance1 == None:
        tolerance1 = config.SNAP_X_TOLERANCE
    if tolerance2 == None:
        tolerance2 = config.SNAP_Y_TOLERANCE
    if tolerance3 == None:
        tolerance3 = config.JOIN_X_TOLERANCE
    if tolerance4 == None:
        tolerance4 = config.JOIN_Y_TOLERANCE

    total_edges = merge_edges(x, tolerance1,tolerance2,tolerance3,tolerance4)
    
   
   
    v_edges = [e for e in total_edges if e["orientation"] == "v"]
    h_edges = [e for e in total_edges if e["orientation"] == "h"]
    
    h_edges_o = copy.deepcopy(h_edges)
    v_edges_o = copy.deepcopy(v_edges)
    

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)
    h_edges,v_edges = FilterLooseEdges(h_edges,v_edges)
    
    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)
    h_edges,v_edges = FilterLooseEdges(h_edges,v_edges)
    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)
    
    
    h_edges,v_edges = SnapEdges(h_edges,v_edges)
    
    
   
    return h_edges,v_edges,h_edges_o,v_edges_o


def PdfCleaner_1(img,x,table,tolerance1=None,tolerance2=None,tolerance3=None,tolerance4=None,tolerance5=None):
    """Merge Edges with tolerance and filters edge withn the table boundary formed. This is
       used when the edges are not clear"""
    if tolerance1 == None:
            tolerance1 = config.EDGES_TOLERANCE

    if tolerance2 == None:
            tolerance2 = tableP.TABLE_SNAP_X_TOLERANCE

    if tolerance3 == None:
            tolerance3 = tableP.TABLE_SNAP_Y_TOLERANCE

    if tolerance4 == None:
            tolerance4 = tableP.TABLE_JOIN_X_TOLERANCE

    if tolerance5 == None:
            tolerance5 = tableP.TABLE_JOIN_Y_TOLERANCE

    

    total_edges = merge_edges(x, tolerance2,tolerance3,tolerance4,tolerance5)

    
   
    v_edges = [e for e in total_edges if e["orientation"] == "v"]
    h_edges = [e for e in total_edges if e["orientation"] == "h"]

    
    
    h_edges_o = copy.deepcopy(h_edges)
    v_edges_o = copy.deepcopy(v_edges)   

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    h_edges,v_edges = SnapEdges(h_edges,v_edges)

    

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)
    
    h_edges = FindHorizontalEdges(table,h_edges,tolerance1)
    v_edges = FindVerticalEdges(table,v_edges,tolerance1)
    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    
    
    
   
    
   
    return h_edges,v_edges,h_edges_o,v_edges_o




def FindBottomEdge(top,left,right,h_edges_tb,tolerance=None):
    # Finds the bottom edge of a cell in a table
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    filtered_edges = [h for h in h_edges_tb if h["top"] > top["top"]]
    filtered_edges_1 = [h for h in filtered_edges if Intersection(h,left,tolerance) and Intersection(h,right,tolerance)]
    edges_sorted_by_top = sorted(filtered_edges_1, key=lambda e: e["top"])
    bottom = edges_sorted_by_top[0]
    return bottom

def Indexing(h_edges_tb,v_edges_tb):
    """Indexes the edges if two horizontal edges have same y position they get same index
        Similarly if tow vertical edges have same x position they get same index"""
    h_index = 0
    v_index = 0
    for j,h_edge in enumerate(h_edges_tb):
        if j == 0:
            h_edge["index"] = h_index
            last = h_edge
           
        if j > 0:
            if last["top"] == h_edge["top"]:
                h_edge["index"] = h_index
                last = h_edge
                
            elif last["top"] < h_edge["top"]:
                h_index = h_index + 1
                h_edge["index"] = h_index
                last = h_edge

    for k,v_edge in enumerate(v_edges_tb):
        if k == 0:
            v_edge["index"] = v_index
            last = v_edge
           
        if k > 0:
            if last["x0"] == v_edge["x0"]:
                v_edge["index"] = v_index
                last = v_edge
                
            elif last["x0"] < v_edge["x0"]:
                v_index = v_index + 1
                v_edge["index"] = v_index
                last = v_edge
                
    return h_edges_tb,v_edges_tb

def edge_exists(edge, edge_list):
    # Check if edge exists in a list of edges
    for e in edge_list:
        if is_same_edge(edge, e):
            return True
    return False


def FormCells(h_edges_tb,v_edges_tb,tolerance=None):
    # Constrcuting cells from the table formed
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    v_edges_tb.sort(key=lambda e: e["x0"])
    h_edges_tb.sort(key=lambda e: e["top"])
    h_edges_tb,v_edges_tb = RemoveIndexing(h_edges_tb,v_edges_tb)
    h_edges_tb, v_edges_tb = Indexing(h_edges_tb,v_edges_tb)
    maxRowIndex = h_edges_tb[-1]["index"]
    maxColumnIndex = v_edges_tb[-1]["index"]
    table = []
    pts1 = []
    max_dim = (maxRowIndex,maxColumnIndex)
    
    for j,h_edge in enumerate(h_edges_tb):
        
        if j < len(h_edges_tb) - 1:
            
            v_int = [v.copy() for v in h_edge["intersections"]]

            v_int = [e for e in v_int if not nearedges(e["edge"]["bottom"],h_edge["bottom"],tolerance)]
            filtered_v_int = []
            

            for v in v_int:
                if edge_exists(v["edge"], v_edges_tb):
                    filtered_v_int.append(v)

            for i,v in enumerate(filtered_v_int):
               
                if i < len(filtered_v_int) - 1: 
                    start_row = 0
                    end_row = 0
                    start_col = 0
                    end_col = 0
                    cell = []
                    pts = []
                    a = []
                    b = []
                    c = []
                    d = []
                    left = filtered_v_int[i]["edge"]
                    right = filtered_v_int[i+1]["edge"]
                    bottom = FindBottomEdge(h_edge,left,right,h_edges_tb)
                    
                    start_row = h_edge["index"]
                    end_row = bottom["index"] 
                   
                    start_col = left["index"]
                    end_col = right["index"] 
                    if end_row - start_row == 0:
                        return None,None,None
                    if end_col - start_col == 0:
                        return None,None,None
                   
                    cell.append((start_row,end_row))
                    cell.append((start_col,end_col))
                    cell.append((left["x0"],right["x0"]))
                    cell.append((h_edge["top"],bottom["top"]))
                    table.append(cell)
                    a.append(left["x0"])
                    a.append(h_edge["top"])
                    b.append(right["x0"])
                    b.append(h_edge["top"])
                    c.append(right["x0"])
                    c.append(bottom["top"])
                    d.append(left["x0"])
                    d.append(bottom["top"])
                    pts.append(a)
                    pts.append(b)
                    pts.append(c)
                    pts.append(d)
                    pts1.append(pts)
            
    
    
    return pts1,table,max_dim   

def AddVLeftEdge(edge1,edge2,v_edges,min_prop):
    # Adds Left Vertical Edge
    edge = copy.deepcopy(v_edges[0])
    edge["top"] = edge1["top"]
    edge["bottom"] = edge2["top"]
    
    edge["x0"] = edge2[min_prop] if edge1[min_prop] < edge2[min_prop] else edge1[min_prop]
    edge["x1"] = edge2[min_prop] if edge1[min_prop] < edge2[min_prop] else edge1[min_prop]
    v_edges.append(edge)
    return v_edges,edge

                    
def AddVRightEdge(edge1,edge2,v_edges,min_prop):
    # Adds Right Vertical Edge
    edge = copy.deepcopy(v_edges[0])
    edge["top"] = edge1["top"]
    edge["bottom"] = edge2["top"]
    
    edge["x0"] = edge1[min_prop] if edge1[min_prop] > edge2[min_prop] else edge2[min_prop]
    edge["x1"] = edge1[min_prop] if edge1[min_prop] > edge2[min_prop] else edge2[min_prop]
    v_edges.append(edge)
    return v_edges,edge

def AddHTopEdge(edge1,edge2,h_edges,min_prop):
    # Adds Horizontal Top Edge
    edge = copy.deepcopy(h_edges[0])
    edge["x0"] = edge1["x0"]
    edge["x1"] = edge2["x0"]
    
    
    edge["top"] = edge2[min_prop] if edge1[min_prop] < edge2[min_prop] else edge1[min_prop]
    edge["bottom"] = edge2[min_prop] if edge1[min_prop] < edge2[min_prop] else edge1[min_prop]
    
    h_edges.append(edge)
    return h_edges,edge

def AddHBottomEdge(edge1,edge2,h_edges,min_prop):
    # Add Horizontal Bottom Edge
    edge = copy.deepcopy(h_edges[0])
    edge["x0"] = edge1["x0"]
    edge["x1"] = edge2["x0"]
    
    edge["top"] = edge2[min_prop] if edge1[min_prop] > edge2[min_prop] else edge1[min_prop]
    edge["bottom"] = edge2[min_prop] if edge1[min_prop] > edge2[min_prop] else edge1[min_prop]
    h_edges.append(edge)
    return h_edges,edge

def ReconstructEdges(tb,h_edges,v_edges):
    # Reconstructs boundary edges of table
    if tb[1]["top"] < tb[0]["top"]:
        tb[0]["top"] = tb[1]["top"]
        h_edges,edge = AddHTopEdge(tb[0],tb[1],h_edges,"top")
        tb[2] = edge
    else:
        tb[1]["top"] = tb[0]["top"]
        h_edges,edge = AddHTopEdge(tb[0],tb[1],h_edges,"top")
        tb[2] = edge
    
    if tb[0]["bottom"] < tb[1]["bottom"]:
        tb[0]["bottom"] = tb[1]["bottom"]
        h_edges,edge = AddHBottomEdge(tb[0],tb[1],h_edges,"bottom")
        tb[3] = edge
    else:
        tb[1]["bottom"] = tb[0]["bottom"]
        h_edges,edge = AddHBottomEdge(tb[0],tb[1],h_edges,"bottom")
        tb[3] = edge

    if tb[2]["x1"] > tb[3]["x1"]:
        tb[3]["x1"] = tb[2]['x1']
        v_edges,edge = AddVRightEdge(tb[2],tb[3],v_edges,"x1")
        tb[1] = edge
    else:
        tb[2]["x1"] = tb[3]['x1']
        v_edges,edge = AddVRightEdge(tb[2],tb[3],v_edges,"x1")
        tb[1] = edge

    if tb[2]["x0"] < tb[3]["x0"]:
        tb[3]["x0"] = tb[2]['x0']
        v_edges,edge = AddVLeftEdge(tb[2],tb[3],v_edges,"x0")
        tb[0] = edge
    else:
        tb[2]["x0"] = tb[3]['x0']
        v_edges,edge = AddVLeftEdge(tb[2],tb[3],v_edges,"x0")
        tb[0] = edge
    return h_edges,v_edges

def CheckTableExtended(table,h_edges,v_edges,tolerance = None):
    """Checks if top edge and bottom edge are extended i.e left intersection of 
        both edges does not match with x0 or x1 of those edges
        Simialrly left edge and right edge"""
    if tolerance == None:
        tolerance = config.EXTENSION_TOLERANCE
    isExtended = False
    
    if not (nearedges(table[0]["x0"],table[2]["x0"],tolerance) or nearedges(table[0]["x0"],table[3]["x0"],tolerance)):
        v_edges,edge =  AddVLeftEdge(table[2],table[3],v_edges,"x0")
        table[0] = edge
        isExtended = True
        
    if not (nearedges(table[1]["x1"],table[2]["x1"],tolerance) or nearedges(table[1]["x1"],table[3]["x1"],tolerance)):
        v_edges,edge =  AddVRightEdge(table[2],table[3],v_edges,"x1")
        table[1] = edge
        isExtended = True
        
    if not (nearedges(table[2]["top"],table[0]["top"],tolerance) or nearedges(table[2]["top"],table[1]["top"],tolerance)):
        
        h_edges,edge =  AddHTopEdge(table[0],table[1],h_edges,"top")
        table[2] = edge
        isExtended = True
        
    if not (nearedges(table[3]["bottom"],table[0]["bottom"],tolerance) or nearedges(table[3]["bottom"],table[1]["bottom"],tolerance)):
        h_edges,edge =  AddHBottomEdge(table[0],table[1],h_edges,"bottom")
        table[3] = edge
        isExtended = True
        
    return isExtended,h_edges,v_edges 

def CheckHEdges(h_edges,table,tolerance=None):
    """Checks if more than 3 Horizontal edges exists that do not intersect with boundary of table.
        This is used to check if table exists within a table"""
    if tolerance == None:
        tolerance = config.DEFAULT_TOLERANCE
    h_edges.sort(key=lambda e: (e["top"], e["x0"]))
    h_edges = [h for h in h_edges if h["top"] >= table[2]["top"] and h["top"] <= table[3]["top"]]
    prev_edge = h_edges[0]
    left = table[0]["x0"]
    right = table[1]["x0"]
    
    
    edgeSum = 0
    isSameTop = False
    isFound = False
    for i, h_edge in enumerate(h_edges[1:]):
        
        if i < len(h_edges) - 2:
            if nearedges(prev_edge["top"],h_edge["top"],tolerance):
                
                isSameTop = True
                if nearedges(h_edge["x0"],left,tolerance) or nearedges(h_edge["x1"],right,tolerance):
                    prev_edge = h_edge
                    isFound = True
                    
                else:
                    prev_edge = h_edge
                    
                
            else:
                if isFound == False and isSameTop :
                    edgeSum = edgeSum + 1
                    
                if not nearedges(h_edge["x0"],left,tolerance) and not nearedges(h_edge["x1"],right,tolerance) and not nearedges(h_edges[i+1]["top"],h_edges[i]["top"],tolerance):
                    edgeSum = edgeSum + 1
                isFound = False
                
                prev_edge = h_edge
                isSameTop = False
            
        
        if edgeSum >= 3:
            return True
            
        
    return False
       
        
def CheckVEdges(v_edges,table,tolerance=None):
    """Checks if more than 3 Vertical edges exists that do not intersect with boundary of table.
        This is used to check if table exists within a table"""
    if tolerance == None:
        tolerance = config.DEFAULT_TOLERANCE
    v_edges.sort(key=lambda e: e["x0"])
    v_edges = [v for v in v_edges if v["x0"] >= table[0]["x0"] and v["x0"] <= table[1]["x0"]]
    prev_edge = v_edges[0]
    top = table[2]["top"]
    bottom = table[3]["top"]
    
    
    edgeSum = 0
    isSameTop = False
    isFound = False
    for i, v_edge in enumerate(v_edges[1:]):
        
        if i < len(v_edges) - 2:
            if nearedges(prev_edge["x0"],v_edge["x0"],tolerance):
                
                isSameTop = True
                if nearedges(v_edge["top"],top,tolerance) or nearedges(v_edge["bottom"],bottom,tolerance):
                    prev_edge = v_edge
                    isFound = True
                    
                else:
                    prev_edge = v_edge
                    
                
            else:
                if not isFound and isSameTop :
                    edgeSum = edgeSum + 1
                    
                if not nearedges(v_edge["top"],top,tolerance) and not nearedges(v_edge["bottom"],bottom,tolerance) and not nearedges(v_edges[i+1]["x0"],v_edges[i]["x0"],tolerance):
                    edgeSum = edgeSum + 1
                isFound = False
                
                prev_edge = v_edge
                isSameTop = False
            
        
        if edgeSum >= 3:
            return True
            
        
    return False

def BordersOutside(table,h_edges,v_edges):

    # IF returns True removes invalid table boundaries
    
    if len(table[0]["intersections"]) == 2 and len(table[1]["intersections"]) == 2 and len(table[2]["intersections"]) == 2 and len(table[3]["intersections"]) == 2:
            for h_edge in h_edges:
                if h_edge != table[2] and h_edge != table[3]:
                    if table[0]["x0"] < h_edge["x0"] < table[1]["x0"] and table[2]["top"] < h_edge["top"] < table[3]["top"]:
                        return True
        
    if table[2] == table[3] or table[0] == table[1]:
        return True
   
    if CheckHEdges(h_edges,table) and CheckVEdges(v_edges,table):
        
        return True
    
    return False

def is_same_edge(e1, e2, tol=2):
    # Checks if two edges are same
    return (
        abs(e1["x0"] - e2["x0"]) < tol and
        abs(e1["x1"] - e2["x1"]) < tol and
        abs(e1["top"] - e2["top"]) < tol and
        abs(e1["bottom"] - e2["bottom"]) < tol and
        e1["orientation"] == e2["orientation"]
    )

def LoopFindTable(h_edges,v_edges):
     # Loops through edges and returns a table
     i = 0
     while(i<len(h_edges)-1):
        try:
            table,h_edges,v_edges = FindTable(i,h_edges,v_edges)
            if table:
                return table,h_edges,v_edges
                
        except:
            i = i+1

     return None,h_edges,v_edges

def UnusualTable(table):
    # If returns True removes invalid Boundary table
    left = table[0]["x0"]
    right = table[1]["x0"]
    top = table[2]["top"]
    bottom = table[3]["top"]
    if left > table[2]["x0"] + (table[3]["x1"] - table[2]["x0"])*0.5:
        return True
    if right < table[2]["x0"] + (table[3]["x1"] - table[2]["x0"])*0.5:
        return True
    if top > table[0]["top"] + (table[1]["bottom"] - table[0]["top"])*0.5:
        return True
    if bottom < table[0]["top"] + (table[1]["bottom"] - table[0]["top"])*0.5:
        return True
    return False  

def NearestVEdgeLeft(h_edge, v_edges):
    # Finds nearest Left Vertical Edge for Horizontal Edge
    filtered_v = [v for v in v_edges if v["x0"] <= h_edge["x0"]]
    return filtered_v[-1] if filtered_v else None

def NearestVEdgeRight(h_edge, v_edges):
    # Finds nearest Right Vertical Edge for Horizontal Edge
    filtered_v = [v for v in v_edges if v["x1"] >= h_edge["x1"]]
    return filtered_v[0] if filtered_v else None

def NearestHEdgeTop(v_edge, h_edges):
    # Finds nearest Top Horizontal Edge for Vertical Edge
    filtered_h = [h for h in h_edges if h["top"] <= v_edge["top"]]
    return filtered_h[-1] if filtered_h else None

def NearestHEdgeBottom(v_edge, h_edges):
    # Finds nearest Bottom Horizontal Edge for Vertical Edge
    filtered_h = [h for h in h_edges if h["bottom"] >= v_edge["bottom"]]
    return filtered_h[0] if filtered_h else None

def FixLooseHorizontalEdges(h_edges_tb,v_edges_tb):
    """Checks whether there is no left intersection for start of Horizontal Edge
      then extends the Horizontal edge to nearest left vertical edge similarly 
      end of Horizontal Edge"""
    for h_edge in h_edges_tb:
        if not CheckNoVerticalLeft(h_edge,v_edges_tb):
           v_edge = NearestVEdgeLeft(h_edge,v_edges_tb)
           h_edge["x0"] = v_edge["x0"]

        if not CheckNoVerticalRight(h_edge,v_edges_tb):
           v_edge = NearestVEdgeRight(h_edge,v_edges_tb)
           h_edge["x1"] = v_edge["x1"]

    return h_edges_tb

def FixLooseVerticalEdges(h_edges_tb,v_edges_tb):
    """Checks whether there is no top intersection for start of Vertical Edge
      then extends the Vertical edge to nearest top horizontal edge similarly 
      end of Vertical Edge"""
    for v_edge in v_edges_tb:
        if not CheckNoHorizontalTop(v_edge,h_edges_tb):
           h_edge = NearestHEdgeTop(v_edge,h_edges_tb)
           v_edge["top"] = h_edge["top"]

        if not CheckNoHorizontalBottom(v_edge,h_edges_tb):
           h_edge = NearestHEdgeBottom(v_edge,h_edges_tb)
           v_edge["bottom"] = h_edge["bottom"]

    return v_edges_tb

def Tableextractor(img,x,tolerance=None,tolerance1=None,tolerance2=None,tolerance3=None,tolerance4=None):
    # Extracts table from edges
    if tolerance == None:
            tolerance = config.EDGES_TOLERANCE

    if tolerance1 == None:
        tolerance1 = config.SNAP_X_TOLERANCE
    if tolerance2 == None:
        tolerance2 = config.SNAP_Y_TOLERANCE
    if tolerance3 == None:
        tolerance3 = config.JOIN_X_TOLERANCE
    if tolerance4 == None:
        tolerance4 = config.JOIN_Y_TOLERANCE
    
    try:
        e = []
        h_edges,v_edges,h_edges_o,v_edges_o = PdfCleaner(img,x)
        
        a = 0
        
        table,h_edges,v_edges = LoopFindTable(h_edges,v_edges)
        
        if UnusualTable(table):
            h_edges = [e for e in h_edges if not is_same_edge(e, table[2])]
            h_edges = [e for e in h_edges if not is_same_edge(e, table[3])]
            v_edges = [e for e in v_edges if not is_same_edge(e, table[0])]
            v_edges = [e for e in v_edges if not is_same_edge(e, table[1])]
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)
            table,h_edges,v_edges = LoopFindTable(h_edges,v_edges)
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)
       

        if BordersOutside(table,h_edges,v_edges):
                
            h_edges = [e for e in h_edges if not is_same_edge(e, table[2])]
            h_edges = [e for e in h_edges if not is_same_edge(e, table[3])]
            v_edges = [e for e in v_edges if not is_same_edge(e, table[0])]
            v_edges = [e for e in v_edges if not is_same_edge(e, table[1])]
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)
            table,h_edges,v_edges = LoopFindTable(h_edges,v_edges)
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)
            
        if table:
           
            h_edges_tb = FindHorizontalEdges(table,h_edges,tolerance)
            v_edges_tb = FindVerticalEdges(table,v_edges,tolerance)

            

            isMisAligned,h_edges,v_edges = CheckMisAlignedTableX(h_edges_tb,v_edges_tb,h_edges,v_edges)
            isMisAligned,h_edges,v_edges = CheckMisAlignedTableY(h_edges_tb,v_edges_tb,h_edges,v_edges)

            if isMisAligned:
                table,h_edges,v_edges = LoopFindTable(h_edges,v_edges)
                h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
                h_edges,v_edges = FindIntersections(h_edges,v_edges)
            

        isTableExtended,h_edges,v_edges = CheckTableExtended(table,h_edges,v_edges)

        
        
        if isTableExtended:
            
            v_edges.sort(key=lambda e: e["x0"])
            h_edges.sort(key=lambda e: (e["top"], e["x0"]))
            table,h_edges,v_edges = FixOverlappingEdges(table,h_edges,v_edges)
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)
            table,h_edges,v_edges = LoopFindTable(h_edges,v_edges)
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)

        

        if not isTableConnected(table):
            

            h_edges,v_edges = ReconstructEdges(table,h_edges,v_edges)
            
            v_edges.sort(key=lambda e: e["x0"])
            h_edges.sort(key=lambda e: (e["top"], e["x0"]))
            table, h_edges, v_edges = FixOverlappingEdges(table,h_edges,v_edges)
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)
            table,h_edges,v_edges = LoopFindTable(h_edges,v_edges)
            
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)

       

        table, h_edges, v_edges = FixOverlappingEdges(table,h_edges,v_edges)

        h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
        h_edges,v_edges = FindIntersections(h_edges,v_edges)
        

        

        h_edges_tb = FindHorizontalEdges(table,h_edges,tolerance)
        v_edges_tb = FindVerticalEdges(table,v_edges,tolerance)

        v_edges_tb.sort(key=lambda e: e["x0"])
        h_edges_tb.sort(key=lambda e: e["top"])

        h_edges_tb= FixLooseHorizontalEdges(h_edges_tb,v_edges_tb)
        #v_edges_tb= FixLooseVerticalEdges(h_edges_tb,v_edges_tb)

        total_edges = merge_edges(h_edges+v_edges, tolerance1,tolerance2,tolerance3,tolerance4)
        v_edges = [e for e in total_edges if e["orientation"] == "v"]
        h_edges = [e for e in total_edges if e["orientation"] == "h"]

        h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
        h_edges,v_edges = FindIntersections(h_edges,v_edges)
        
        h_edges_tb = FindHorizontalEdges(table,h_edges,tolerance)
        v_edges_tb = FindVerticalEdges(table,v_edges,tolerance)
       
        
        v_edges_tb.sort(key=lambda e: e["x0"])
        h_edges_tb.sort(key=lambda e: e["top"])
        pts,tableComp,max_dim = FormCells (h_edges_tb,v_edges_tb)
        if pts == None and tableComp == None and max_dim == None:
            return table,None,None,h_edges,v_edges,h_edges_o,v_edges_o,None,None,None

    except:
        return table,None,None,h_edges,v_edges,h_edges_o,v_edges_o,None,None,None

    return table,h_edges_tb,v_edges_tb,h_edges,v_edges,h_edges_o,v_edges_o,pts,tableComp,max_dim




def Tableextractor_1(img,x,table,tolerance=None,tolerance2=None,tolerance3=None,tolerance4=None,tolerance5=None):
        # Extracts table from Edges after edges are connected using tableformation fucntion
        if tolerance == None:
            tolerance = config.EDGES_TOLERANCE

        
        e = []
        h_edges,v_edges,h_edges_o,v_edges_o = PdfCleaner_1(img,x,table)
        
        table,h_edges,v_edges = LoopFindTable(h_edges_o,v_edges_o)
        h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
        h_edges,v_edges = FindIntersections(h_edges,v_edges)

        if UnusualTable(table):
            h_edges = [e for e in h_edges if not is_same_edge(e, table[2])]
            h_edges = [e for e in h_edges if not is_same_edge(e, table[3])]
            v_edges = [e for e in v_edges if not is_same_edge(e, table[0])]
            v_edges = [e for e in v_edges if not is_same_edge(e, table[1])]
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)
            table,h_edges,v_edges = LoopFindTable(h_edges,v_edges)
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)

        if BordersOutside(table,h_edges,v_edges):
           
            h_edges = [e for e in h_edges if not is_same_edge(e, table[2])]
            h_edges = [e for e in h_edges if not is_same_edge(e, table[3])]
            v_edges = [e for e in v_edges if not is_same_edge(e, table[0])]
            v_edges = [e for e in v_edges if not is_same_edge(e, table[1])]
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)
            table,h_edges,v_edges = LoopFindTable(h_edges,v_edges)
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)

        isTableExtended,h_edges,v_edges = CheckTableExtended(table,h_edges,v_edges)

        
        if isTableExtended:
            v_edges.sort(key=lambda e: e["x0"])
            h_edges.sort(key=lambda e: (e["top"], e["x0"]))
           
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)
            table,h_edges,v_edges = LoopFindTable(h_edges,v_edges)
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)

        if not isTableConnected(table):
            h_edges,v_edges = ReconstructEdges(table,h_edges,v_edges)
            
            v_edges.sort(key=lambda e: e["x0"])
            h_edges.sort(key=lambda e: (e["top"], e["x0"]))
           
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)
            table,h_edges,v_edges = LoopFindTable(h_edges,v_edges)
            
            h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
            h_edges,v_edges = FindIntersections(h_edges,v_edges)



        h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
        h_edges,v_edges = FindIntersections(h_edges,v_edges)
        

        h_edges_tb = FindHorizontalEdges(table,h_edges_o,tolerance)
        v_edges_tb = FindVerticalEdges(table,v_edges_o,tolerance)

        
        if table[2] not in h_edges_tb:
            h_edges_tb.append(table[2])
        if table[3] not in h_edges_tb:
            h_edges_tb.append(table[3])

        if table[0] not in v_edges_tb:
            v_edges_tb.append(table[0])
        if table[1] not in v_edges_tb:
            v_edges_tb.append(table[1])
        
        h_edges_n = [e for e in h_edges_o if e not in h_edges_tb]
        v_edges_n = [e for e in v_edges_o if e not in v_edges_tb]
        return h_edges_tb,v_edges_tb,h_edges_n,v_edges_n

def FindBetweenEdges(h_1,h_2,v_edges,tolerance=None):
    # Finds Vertical Edges between two Horizontal Edges
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    v_edges_b = []
    if h_1["x0"] < h_2["x0"]:
        comp = h_2
    else:
        comp = h_1
    for v_edge in v_edges:
        if nearedges(v_edge["x0"],comp["x0"],tolerance) or comp["x0"] < v_edge["x0"] < comp["x1"] or nearedges(v_edge["x1"],comp["x1"],tolerance):
            if (Intersection(h_1,v_edge,tolerance) and Intersection(h_2,v_edge,tolerance)) or h_1["top"] < v_edge["top"] < h_2["top"] or h_1["top"] <  v_edge["bottom"] < h_2["top"] :
                v_edges_b.append(v_edge)
    return v_edges_b

def FindNearBottomEdge(edge,h_edges,tolerance=None):
    # Finds near bottom edge
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    h_edges.sort(key=lambda e: (e["top"], e["x0"]))
    filtered_edges = [h for h in h_edges if h["top"] > edge["top"]]
    for h_edge in filtered_edges:
        if nearedges(h_edge["x0"],edge["x0"],tolerance) and nearedges(h_edge["x1"],edge["x1"],tolerance):
            return h_edge,None,None
        if not nearedges(h_edge["x0"],edge["x0"],tolerance):
            if nearedges(h_edge["x1"],edge["x1"],tolerance):
                min_prop = "x0" 
                max_prop = None
                return h_edge,min_prop,max_prop
        if not nearedges(h_edge["x1"],edge["x1"],tolerance):
            if nearedges(h_edge["x0"],edge["x0"],tolerance):
                min_prop = None
                max_prop = "x1"
                return h_edge,min_prop,max_prop
        if not nearedges(h_edge["x0"],edge["x0"],tolerance) and not nearedges(h_edge["x1"],edge["x1"],tolerance):
            if h_edge["width"] > edge["width"]:
                if h_edge["x0"] < edge["x0"] < h_edge["x1"] and h_edge["x0"] < edge["x1"] < h_edge["x1"]:
                    min_prop = "x0" 
                    max_prop = "x1"
                    return h_edge,min_prop,max_prop
            else:
                if edge["x0"] < h_edge["x0"] < edge["x1"] and edge["x0"] < h_edge["x1"] < edge["x1"]:
                    min_prop = "x0" 
                    max_prop = "x1"
                    return h_edge,min_prop,max_prop
    return filtered_edges[0],None,None
    
def FindNearBottomEdgeLeft(edge,h_edges,tolerance=None):
    # Finds near bottom edge w.r.t start of Horizontal edge
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    h_edges.sort(key=lambda e: (e["top"], e["x0"]))
    filtered_edges = [h for h in h_edges if h["top"] > edge["top"]]
    for h_edge in filtered_edges:
        if nearedges(edge["x0"],h_edge["x0"],tolerance):
            return h_edge
      
    return filtered_edges[0]

def FindNearBottomEdgeRight(edge,h_edges,tolerance=None):
    # # Finds near bottom edge w.r.t end of Horizontal edge
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    h_edges.sort(key=lambda e: (e["top"], e["x0"]))
    filtered_edges = [h for h in h_edges if h["top"] > edge["top"]]
    for h_edge in filtered_edges:
        if nearedges(edge["x1"],h_edge["x1"],tolerance):
            return h_edge
     
    return filtered_edges[0]

def FillGaps(edge1,edge2,v_edges_b,v_edges,tolerance=None):
    # Extends Vertical edges to intersect with Horizontal edges
    if tolerance == None:
            tolerance = config.GAP_TOLERANCE
    for j,v_edge in enumerate(v_edges_b):
        if not Intersection(edge1,v_edge,tolerance):
           
            v_edge["top"] = edge1["top"]
        if not Intersection(edge2,v_edge,tolerance):
            v_edge["bottom"] = edge2["top"]
            
def CheckNoVerticalLeft(edge,v_edges,tolerance=None):
    # Checks if start of Horizontal edge is inersecting with vertical edge
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    for v_edge in v_edges:
        if Intersection(edge,v_edge,tolerance) and nearedges(edge["x0"],v_edge["x0"],tolerance) :
            return True
    return False

def CheckNoVerticalRight(edge,v_edges,tolerance=None):
    # Checks if end of Horizontal edge is inersecting with vertical edge
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    for v_edge in v_edges:
        if Intersection(edge,v_edge,tolerance) and nearedges(edge["x1"],v_edge["x1"],tolerance) :
            return True
    return False

def CheckNoHorizontalTop(edge,h_edges,tolerance=None):
    # Checks if start of Horizontal edge is inersecting with vertical edge
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    for h_edge in h_edges:
        if Intersection(h_edge,edge,tolerance) and nearedges(edge["top"],h_edge["top"],tolerance):
            return True
    return False

def CheckNoHorizontalBottom(edge,h_edges,tolerance=None):
    # Checks if end of Horizontal edge is inersecting with vertical edge
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    for h_edge in h_edges:
        if Intersection(h_edge,edge,tolerance) and nearedges(edge["bottom"],h_edge["bottom"],tolerance):
            return True
    return False

def TopHEdge(h_edge,h_edges):
    # Finds nearest top Horziontal edge for a given Horizontal edge
    dist = 0
    minDist = float('inf')
    minEdge = None
    filtered_edges = [h for h in h_edges if h["top"] < h_edge["top"]]
    for h in filtered_edges:
        dist = abs(h["top"]-h_edge["top"]) 
        if dist<minDist:
            minDist = dist
            minEdge = h
        
    return minEdge

def BottomHEdge(h_edge,h_edges):
    # Finds nearest bottom Horziontal edge for a given Horizontal edge
    dist = 0
    minDist = float('inf')
    minEdge = None
    filtered_edges = [h for h in h_edges if h["top"] > h_edge["top"]]
    for h in filtered_edges:
        dist = abs(h["top"]-h_edge["top"]) 
        if dist<minDist:
            minDist = dist
            minEdge = h
    return minEdge

def isExistsVertical(h_edge,v_edges,min_prop):
    # Checks if a vertical edge exists at start or end of horizotal edge
    for v_edge in v_edges:
        if nearedges(v_edge[min_prop],h_edge[min_prop],config.DEFAULT_TOLERANCE) :
            return True
    return False

def AddVEdgeLeft(top,bottom,h_edge,v_edges):
    # Adds Vertical Edge to left of Horizontal edge
    edge = copy.deepcopy(v_edges[0])
    edge["top"] = top["top"]
    edge["bottom"] = bottom["top"]
    edge["x0"] = h_edge["x0"]
    edge["x1"] = h_edge["x0"]
    v_edges.append(edge)
    return v_edges

def AddVEdgeRight(top,bottom,h_edge,v_edges):
    # Adds Vertical Edge to right of Horizontal edge
    edge = copy.deepcopy(v_edges[0])
    edge["top"] = top["top"]
    edge["bottom"] = bottom["top"]
    edge["x0"] = h_edge["x1"]
    edge["x1"] = h_edge["x1"]
    v_edges.append(edge)
    return v_edges

def SameLineTable(edge,v_edges,tolerance=None):
    # Checks if vertical edge exists in the same line for a given vertical edge and returns that edge
    if tolerance == None:
        tolerance = config.DEFAULT_TOLERANCE
    for v_edge in v_edges:
        if nearedges(edge["x0"],v_edge["x0"],tolerance) and v_edge != edge:
            return v_edge
    return None

def SameLine(h_edges_tb_1,h_edges_tb_2):
    """Checks whether two tables align horizontally """
    top = h_edges_tb_1[0]["top"]
    bottom = h_edges_tb_1[-1]["top"]
    for h_edge in h_edges_tb_2:
        if top <= h_edge["top"] <= bottom:
            return True
    return False


def CheckHLine(v_edge,h_edges,tolerance=None):
    # Checks if horizontal edge exists in the same line for a given horizontal edge
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    for h_edge in h_edges:
        if nearedges(h_edge["x0"],v_edge["x0"],tolerance) or nearedges(h_edge["x1"],v_edge["x1"],tolerance) :
            return True
    return False    
    

def TableFormation(img,x,table,tolerance1=None,tolerance2=None,tolerance3=None,tolerance4=None,tolerance5=None):
    # Forms table from broken edges
    if tolerance1 == None:
            tolerance1 = config.DEFAULT_TOLERANCE

    if tolerance2 == None:
            tolerance2 = tableP.EDGE_SNAP_X_TOLERANCE

    if tolerance3 == None:
            tolerance3 = tableP.EDGE_SNAP_Y_TOLERANCE

    if tolerance4 == None:
            tolerance4 = tableP.EDGE_JOIN_X_TOLERANCE

    if tolerance5 == None:
            tolerance5 = tableP.EDGE_JOIN_Y_TOLERANCE

    #img1 = copy.deepcopy(img)
      
    h_edges,v_edges,h_edges_n,v_edges_n = Tableextractor_1(img,x,table)
    
    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)

    
    
    
    

    v_edges.sort(key=lambda e: e["x0"])
    h_edges.sort(key=lambda e: (e["top"], e["x0"]))

    mergeR = []
   
    edgeC_1 = []
    edgeC_2 = []
    
    for i,h_edge in enumerate(h_edges):
        if not nearedges(h_edge["top"],h_edges[-1]["top"],tolerance1):
            
            t1 = h_edges[i]
            t2,min_prop,max_prop = FindNearBottomEdge(t1,h_edges)
            mergeR.append((t2,t1,min_prop,max_prop))
            v_edges_b = FindBetweenEdges(t1,t2,v_edges)
            FillGaps(t1,t2,v_edges_b,v_edges)
    
    
                    
   
            
    for t in mergeR:
        (a,b,c,d) = t
        if c == None and d == None:
            continue
        if c == "x0" and d == None:
            e = FindNearBottomEdgeLeft(b,h_edges)
            v_edges_b = FindBetweenEdges(b,e,v_edges)
            if b["width"] > a["width"]:
                v_edges_b = [v for v in v_edges_b if v["x0"] < a["x0"]]
                FillGaps(b,e,v_edges_b,v_edges)
            else:
                v_edges_b = [v for v in v_edges_b if v["x0"] < b["x0"]]
                FillGaps(b,e,v_edges_b,v_edges)
        if c == None and d == "x1":
            e = FindNearBottomEdgeRight(b,h_edges)
            v_edges_b = FindBetweenEdges(b,e,v_edges)
            if b["width"] > a["width"]:
                v_edges_b = [v for v in v_edges_b if v["x1"] > a["x1"]]
                FillGaps(b,e,v_edges_b,v_edges)
            else:
                v_edges_b = [v for v in v_edges_b if v["x1"] > b["x1"]]
                FillGaps(b,e,v_edges_b,v_edges)
        if c == "x0" and d == "x1":
            e = FindNearBottomEdgeLeft(b,h_edges)
            v_edges_b = FindBetweenEdges(b,e,v_edges)
            if b["width"] > a["width"]:
                v_edges_b = [v for v in v_edges_b if v["x0"] < a["x0"]]
                FillGaps(b,e,v_edges_b,v_edges)
            else:
                v_edges_b = [v for v in v_edges_b if v["x0"] < b["x0"]]
                FillGaps(b,e,v_edges_b,v_edges)
            e = FindNearBottomEdgeRight(b,h_edges)
            v_edges_b = FindBetweenEdges(b,e,v_edges)
            if b["width"] > a["width"]:
                v_edges_b = [v for v in v_edges_b if v["x1"] > a["x1"]]
                FillGaps(b,e,v_edges_b,v_edges)
            else:
                v_edges_b = [v for v in v_edges_b if v["x1"] > b["x1"]]
                FillGaps(b,e,v_edges_b,v_edges)

    
    
    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)
    

    for i,h_edge in enumerate(h_edges):
        
        isLeft = CheckNoVerticalLeft(h_edge,v_edges)
        isRight = CheckNoVerticalRight(h_edge,v_edges)

        if not isLeft :
            edgeC_1.append(h_edge)
        if not isRight:
            edgeC_2.append(h_edge)

    v_edges.sort(key=lambda e: e["x0"])
    h_edges.sort(key=lambda e: (e["top"], e["x0"]))

    

    for i,h_edge in enumerate(h_edges):
        notFoundL = False
        notFoundR = False
        if i < len(h_edges) - 1 and i > 0:
            
            isLeft = CheckNoVerticalLeft(h_edge,v_edges)
            isRight = CheckNoVerticalRight(h_edge,v_edges)
            if not isLeft:
                if isExistsVertical(h_edge,v_edges,"x0"):
                    top = TopHEdge(h_edge,h_edges)
                    bottom = BottomHEdge(h_edge,h_edges)
                    v_edges = AddVEdgeLeft(top,bottom,h_edge,v_edges)
                    
                else:
                   
                        h_edge["x0"] = h_edges[i-1]["x1"]

        
            
            if not isRight:
                if isExistsVertical(h_edge,v_edges,"x1"):
                    top = TopHEdge(h_edge,h_edges)
                    bottom = BottomHEdge(h_edge,h_edges)
                    v_edges = AddVEdgeRight(top,bottom,h_edge,v_edges)
                    
                else:
                    
                        h_edge["x1"] = h_edges[i+1]["x0"]

    v_edges.sort(key=lambda e: e["x0"])
    h_edges.sort(key=lambda e: (e["top"], e["x0"]))

    

    
    for edgeCC in edgeC_1:
        bottom = BottomHEdge(edgeCC,h_edges)
        if bottom != None:
            if nearedges(bottom["top"],h_edges[-1]["top"],tolerance1):
                v_edges = AddVEdgeLeft(edgeCC,bottom,edgeCC,v_edges)
    
    for edgeCC in edgeC_2:
        bottom = BottomHEdge(edgeCC,h_edges)
        if bottom != None:
            if nearedges(bottom["top"],h_edges[-1]["top"],tolerance1):
                v_edges = AddVEdgeRight(edgeCC,bottom,edgeCC,v_edges)
   

    
    h_edges,v_edges = SnapEdges(h_edges,v_edges)          

    total_edges = merge_edges(h_edges+v_edges, tolerance2,tolerance3,tolerance4,tolerance5)
    v_edges = [e for e in total_edges if e["orientation"] == "v"]
    h_edges = [e for e in total_edges if e["orientation"] == "h"]
   
    
    
    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)
    
    
       
     

   

    for v_edge in v_edges:
    
        isTop = CheckNoHorizontalTop(v_edge,h_edges)
        isBottom = CheckNoHorizontalBottom(v_edge,h_edges)
        
        if not isBottom:
            v_edge_1 = SameLineTable(v_edge,v_edges)
            
            if v_edge_1:
                v_edge["bottom"] = v_edge_1["top"]
            else:
                v_edge["bottom"] = h_edges[-1]["top"]

        if not isTop:
            v_edge_1 = SameLineTable(v_edge,v_edges)
            
            if v_edge_1:
                v_edge["top"] = v_edge_1["bottom"]
            else:
                v_edge["top"] = h_edges[0]["top"]
        
      
        
    total_edges = merge_edges(h_edges+v_edges,tolerance2,tolerance3,tolerance4,tolerance5)
    v_edges = [e for e in total_edges if e["orientation"] == "v"]
    h_edges = [e for e in total_edges if e["orientation"] == "h"]

    v_edges.sort(key=lambda e: e["x0"])
    h_edges.sort(key=lambda e: (e["top"], e["x0"]))

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)
    h_edges,v_edges = SnapEdges(h_edges,v_edges)

    
    for i,v_edge in enumerate(v_edges):
        if i < len(v_edges) - 1:
            if nearedges(v_edges[i]["x0"],v_edges[i+1]["x0"],tolerance1):
                if CheckHLine(v_edges[i+1],h_edges) and CheckHLine(v_edges[i],h_edges):
                    v_edges[i]["bottom"] = v_edges[i+1]["bottom"]
                    

    total_edges = merge_edges(h_edges+v_edges,tolerance2,tolerance3,tolerance4,tolerance5)
    v_edges = [e for e in total_edges if e["orientation"] == "v"]
    h_edges = [e for e in total_edges if e["orientation"] == "h"]

    

    h_edges,v_edges = RemoveIntersections(h_edges,v_edges)
    h_edges,v_edges = FindIntersections(h_edges,v_edges)
    
    
   
    


    return h_edges,v_edges,h_edges_n,v_edges_n


# Step 1: Convert scanned PDF to image


#print(np.__version__)



def extract_edges(thresh_img):
    # Detect edges from image converted from scanned pdf
    horizontal = thresh_img.copy()
    vertical = thresh_img.copy()

    # Horizontal structure
    horiz_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
    detect_horizontal = cv2.morphologyEx(horizontal, cv2.MORPH_OPEN, horiz_kernel)

    # Vertical structure
    vert_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))
    detect_vertical = cv2.morphologyEx(vertical, cv2.MORPH_OPEN, vert_kernel)

    return detect_horizontal, detect_vertical


def EdgeDetection(img):
    # Detect edges and returns edges as dictioanry for table extraction
    h_edges = []
    v_edges = []
    # Step 2: Preprocess image
    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
    blurred = cv2.GaussianBlur(gray, (3, 3), 0)
    thresh = cv2.adaptiveThreshold(
        blurred, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 15, 4
    )
    

    horizontal_edges, vertical_edges = extract_edges(thresh)

    
    # Find contours of table cells


    contours_h, _ = cv2.findContours(horizontal_edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contours_v, _ = cv2.findContours(vertical_edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)



    # Draw contours on image


    for cnt in contours_h:
        x, y, w, h = cv2.boundingRect(cnt)
        x0, top, x1, bottom =  x, y, x + w, y 
        edge = {
            "x0": x0,
            "x1": x1,
            "top": top,
            "bottom": bottom,
            "orientation": "h",
            "width": x1 - x0,
            "height": bottom - top
        }
        h_edges.append(edge)

    for cnt in contours_v:
        x, y, w, h = cv2.boundingRect(cnt)
        x0, top, x1, bottom  = x, y , x, y +h
        edge = {
            "x0": x0,
            "x1": x1,
            "top": top,
            "bottom": bottom,
            "orientation": "v",
            "width": x1 - x0,
            "height": bottom - top
        }
        v_edges.append(edge)

    v_edges.sort(key=lambda e: e["x0"])
    h_edges.sort(key=lambda e: (e["top"], e["x0"]))

    return img,h_edges,v_edges




def filter_text_by_x(j, x0, x1):
    # Returns text which lies between x0 and x1
    char_w = j["charWidth"]
    text = j["text"]
    textx0 = j["x0"]

    result = []
    first_idx = None
    last_idx = None

    for i, ch in enumerate(text):
        char_x = textx0 + i * char_w

        if x0 <= char_x <= x1:
            result.append(ch)

            if first_idx is None:
                first_idx = i
            last_idx = i

    if first_idx is None:
        return "", None, None

    last_x0 = textx0 + first_idx * char_w
    last_x1 = textx0 + (last_idx + 1) * char_w

    return "".join(result), last_x0, last_x1





def FindTextinCells(table,textGrouped):
    # returns text in each cell of table formed
    textCells = []
    last = []
    
    for i, tb in enumerate(table):
        textCell = []
        x0 , x1 = table[i][2]
        top, bottom = table[i][3]
        start_row , end_row = table[i][0]
        start_col, end_col = table[i][1]
        for t in textGrouped:
            if top <= t <= bottom:
                to_remove = []
                for j in textGrouped[t]:
                   if x0 <= j["x0"] + (j["x1"]-j["x0"])/2 <= x1 and x0 <= j["x0"] <= x1 and x0 <= j["x1"] <= x1:
                       textCell.append(j["text"])
                       textCell.append(end_row-start_row)
                       textCell.append(end_col-start_col)
                       textCell.append(j["top"]+(j["bottom"]-j["top"])/2)
                       textCell.append(j["x0"] + (j["x1"]-j["x0"])/2)
                       to_remove.append(j)
                   
                   elif x0 <= j["x0"] <= x1:
                       
                       
                       truncated_text,last_x0,last_x1 = filter_text_by_x(j, x0, x1)
                       if truncated_text:
                            textCell.append(truncated_text)
                            textCell.append(end_row-start_row)
                            textCell.append(end_col-start_col)
                            textCell.append(j["top"]+(j["bottom"]-j["top"])/2)
                            textCell.append(last_x0 + (last_x1-last_x0)/2)

                   elif  x0 <= j["x0"] + (j["x1"]-j["x0"])/2 <= x1:
                       
                       truncated_text,last_x0,last_x1 = filter_text_by_x(j, x0, x1)
                       if truncated_text:
                            textCell.append(truncated_text)
                            textCell.append(end_row-start_row)
                            textCell.append(end_col-start_col)
                            textCell.append(j["top"]+(j["bottom"]-j["top"])/2)
                            textCell.append(last_x0 + (last_x1-last_x0)/2)     
                   
                       
                   elif x0 <= j["x1"] <= x1:
                       
                       
                        truncated_text,last_x0,last_x1 = filter_text_by_x(j, x0, x1)
                        if truncated_text:
                            textCell.append(truncated_text)
                            textCell.append(end_row-start_row)
                            textCell.append(end_col-start_col)
                            textCell.append(j["top"]+(j["bottom"]-j["top"])/2)
                            textCell.append(last_x0 + (last_x1-last_x0)/2)
                        to_remove.append(j)
                
                for k in to_remove:
                    textGrouped[t].remove(k)
        
        if textCell == []:
            textCell.append("")
            textCell.append(end_row-start_row)
            textCell.append(end_col-start_col)
            
        textCells.append(textCell)

    return textCells,textGrouped   

def FindTextClusters(textCells,max_dim):
    # returns text of each row of table formed
    maxRowIndex , maxColumnIndex = max_dim
  
    textRows = []
    totalCol = 0
   
    total = 0
    offset = {}
    start = 0
    end = 0
    isOffset = False
    colsCheck = []
    row = 1
    addT = 0
    col = 0
    isrowStart = False
    
    for textCell in textCells:
       
        rows = textCell[1]
        columns = textCell[2]
        totalCol = columns + totalCol
        
        end = end + 1
        if columns == 1:
            col = col+1
        else:
            col = col+columns
       
         
        if isrowStart == True:
            for i in range(1,maxColumnIndex+1):
                
                if i in offset:
                    for l in offset[i]:
                        (r,c,m) = l
                        
                        addT = c + addT
                    
            isrowStart = False
       
        
        if rows > 1:
            
            
            if columns == 1:
                colCheck = totalCol
            elif columns > 1:
                colCheck = totalCol - columns
            

            if colCheck not in offset:
                offset[colCheck] = []

            offset[colCheck].append((rows - 1, columns, row))
            
                
        
        if maxColumnIndex == totalCol + addT:
            
           
            # Decrease first element of tuple by 1
            for k in list(offset.keys()):
                updated_tuples = []
                for v in offset[k]:
                    if v[2] < row:
                        new_v0 = v[0] - 1
                        if new_v0 != 0:
                            updated_tuples.append((new_v0, v[1], v[2]))
                    else:
                        updated_tuples.append(v)

                if updated_tuples:
                    offset[k] = updated_tuples
                else:
                    del offset[k]
                    
            totalCol = 0
            row = row + 1
            addT = 0      
            textRows.append(textCells[start:end])
            start = end
            col = 0
            isrowStart = True
       
    return textRows



def FindDataFrame(textClusters):
    # returns textrows with text and centered x and y positions
    textRows = []
    for cluster in textClusters:
        row = []
        for group in cluster:
            cell = []
            for i in range(0, len(group), 5):
                if i + 5 <= len(group):
                    text, rspan, cspan, top, center = group[i:i+5]
                    cell.append((str(text), round(top, 1), round(center, 1)))
                else:
                   # print("⚠️ Skipped incomplete group:", group[i:])
                   pass
            row.append(cell)
        textRows.append(row)
    return textRows


def ModifyDataFrame(textDataFrame):
    """Arrange and join cell texts based on their positions.
    If x positions differ, separate by space.
    If y positions differ, separate by newline.
    """
    processed_rows = []
    tolerance = 2

    for row in textDataFrame:
        new_row = []
        for cell in row:
            # Sort by y (top), then by x (center)
            sorted_cell = sorted(cell, key=lambda t: (t[1], (round(t[2] / tolerance) * tolerance)))
            
            output = []
            current_y = None
            line = []

            for i, (text, y, x) in enumerate(sorted_cell):
                y_group = round(y / tolerance) * tolerance
                if current_y is None:
                    current_y = y_group
                if y_group != current_y:
                    output.append("".join(line))
                    line = [text]
                    current_y = y_group
                else:
                    if line:
                        prev_x = sorted_cell[i - 1][2]
                        space_count = max(1, round((x - prev_x) / 25))  # adjustable spacing factor
                        line.append(" " * space_count + text)
                    else:
                        line.append(text)
            if line:
                output.append("".join(line))

            new_row.append("\n".join(output))
        processed_rows.append(new_row)

    return processed_rows



def FindMergeData(textClusters,max_dim):
   # Extract (second, third) from each cell, preserving row structure
    merge_data = {}
    offset = {}
    maxRow, maxCol = max_dim
    
    colIndex = 0
    for row in range(1,maxRow+1):
        colsSkip = []
        colPassed = 0
        rowIndex = row-1
        for col in range(1,maxCol+1):
            isCol = False
            if col == 1:
                colIndex = col-1
            
            if col in colsSkip:
                continue
           
            if row in offset:
                columns = offset[row]
               
                if col in columns:
                    isCol = True
                    continue
            
            if isCol == False:
                cell = textClusters[rowIndex][colIndex]
                rowSpan = cell[1]
                colSpan = cell[2]
                
                if not (rowSpan == 1 and colSpan == 1):
                    if row not in merge_data:
                        merge_data[row] = []
                    merge_data[row].append((col,rowSpan,colSpan))
                
                if rowSpan > 1:
                    for ii in range(row+1,rowSpan+row):
                        if ii not in offset:
                            offset[ii] = []
                        for jj in range(col,col+colSpan):
                            offset[ii].append(jj)
                if colSpan > 1:
                    for jk in range(col,col+colSpan):
                        colsSkip.append(jk)

                colIndex = colIndex+1
                
    return merge_data


   
   
def mergeCells(index1,index2,mergeData,ws,maxRow,maxCol):
    # merge cells and writes test value in top left cell
    for row in range(index1,maxRow+index1-1):
        if not row in mergeData:
            continue
        else:
            start_row = row
            row1 = row
            for col in range(index2,maxCol+index2-1):
                for m in mergeData[row]:
                    column, rowSpan, colSpan = m
                    col1 = col
                    
                    start_row = row1
                    if col == column:
                       start_col = col1
                       end_row = start_row + rowSpan-1
                       end_col = start_col + colSpan-1
                       if rowSpan > 1 and colSpan > 1:
                            
                            ws.cell(row = start_row, column = start_col, value = "Bhanu")
                            ws.merge_cells(start_row = start_row, start_column = start_col, end_row = end_row, end_column = end_col) 
                       elif rowSpan > 1 :
                            
                            ws.cell(row = start_row, column = start_col, value = "Bhanu")
                            ws.merge_cells(start_row = start_row, start_column = start_col, end_row = end_row, end_column = start_col)
                       elif colSpan > 1:
                            
                            ws.cell(row = start_row, column = start_col, value = "Bhanu")
                            ws.merge_cells(start_row = start_row, start_column = start_col, end_row = start_row, end_column = end_col)
    return ws
    
   

def get_merged_cell_map(ws):
    # Returns merged_cells and top_left_cells in merged cell range to write the merged data to top_left_cells 
    merged_cells = set()
    top_left_cells = set()
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_cells.add((r, c))
        top_left_cells.add((min_row, min_col))
    return merged_cells, top_left_cells   



def write_to_excel(index1,index2,wb,ws,data,mergeRowsData,max_dim):
    # writing cell data to excel using openpyxl
    maxRow,maxColumn = max_dim
    maxRow = maxRow + 1
    maxColumn = maxColumn + 1
    
    ws = mergeCells(index1,index2,mergeRowsData,ws,maxRow,maxColumn)
    values = []
    valueTo = 0
    # Write each cell and apply wrap_text where needed
    merged_cells, top_left_cells = get_merged_cell_map(ws)
    
    
        
    for row_idx in range(index1,maxRow+index1-1):
        rowData = data[row_idx-index1]
        if rowData:
            i = 0
            for col_idx in range(index2,maxColumn+index2-1):
                if i < len(rowData):
                    colValue = rowData[i]
                
                if (row_idx, col_idx) in merged_cells:
                        if (row_idx, col_idx) in top_left_cells:
                            start_row, start_col = row_idx, col_idx
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.value = colValue
                            cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                        
                            i = i + 1
                        continue
                        
                else:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = colValue
                        cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                        i = i + 1
                
    
    thin_border = Border(
        left = Side(style='thin'),
        right = Side(style='thin'),
        top = Side(style='thin'),
        bottom = Side(style='thin')
    )

    for row in ws.iter_rows(min_row=index1,
    max_row=maxRow+index1-2, min_col=index2, max_col=maxColumn+index2-2):
        for cell in row:
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                
                cell_text = str(cell.value)

                lines = cell_text.split("\n")
                max_len = max(len(line) for line in lines)

                max_length = max(max_length, max_len)

        ws.column_dimensions[col_letter].width = max_length + 2

    




def IndexCalculation1(present_edge,tableFactor,base_edge_t,base_edge_v,tolerance=None):
    # Caluclates index based on horizontal edge alignment
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    tableFactor1 = tableFactor[:]
    tableFactor1.reverse()
    index1 = 0
    index2 = 0
    for table in tableFactor1:
        isHFound = False
        isVFound = False
        (h_edges_tb,v_edges_tb,max_dim) = table
        maxRow, lindex1,maxColumn, lindex2 = max_dim
        for i, h_edge in enumerate(h_edges_tb):
            if nearedges(h_edge["top"],present_edge["top"],tolerance):
                
                index2 = maxColumn + 3 
                index1 = lindex1 + i
                isHFound = True
                break
        if isHFound == True:
            break

    if index1 == 0 and index2 == 0:
        return -1,-1
    else:
        return index1,index2

def IndexCalculation2(present_edge,tableFactor,base_edge_t,base_edge_v,tolerance=None):
    # Caluclates index based on vertical edge alignment
    if tolerance == None:
            tolerance = config.DEFAULT_TOLERANCE
    tableFactor1 = tableFactor[:]
    tableFactor1.reverse()
    index1 = 0
    index2 = 0
    for table in tableFactor1:
        isHFound = False
        isVFound = False
        (h_edges_tb,v_edges_tb,max_dim) = table
        maxRow, lindex1,maxColumn, lindex2 = max_dim
        for j, v_edge in enumerate(v_edges_tb):
            if nearedges(v_edge["x0"],present_edge["x0"],tolerance):
                index1 = maxRow + 3
                
                index2 = lindex2 + j 
                isVFound = True
                break
        if isVFound == True:
            break

    if index1 == 0 and index2 == 0:
        return -1,-1
    else:
        return index1,index2
    
def IndexCalculation3(present_edge,tableFactor,base_edge_t,base_edge_v):
    # Calculates index if none of horizontal edges or vertical edges align
    tableFactor1 = tableFactor[:]
    tableFactor1.reverse()
    index1 = 0
    index2 = 0
    for table in tableFactor1:
        isHFound = False
        isVFound = False
        (h_edges_tb,v_edges_tb,max_dim) = table
        maxRow, lindex1,maxColumn, lindex2 = max_dim
        index1 = maxRow + 3
        if v_edges_tb[-1]["x0"] < base_edge_v["x0"]:
            index2 = lindex2 + 2
            break
        else:
            (h_edges_tb,v_edges_tb,max_dim) = tableFactor[0]
            maxRow, lindex1,maxColumn, lindex2 = max_dim
            index2 = lindex2 - 2
            break
    return index1,index2


def IndexCalculation(h_edges_tb,v_edges_tb,h_edges,v_edges,tableFactor,first):
   # Calculates row-index and column index of table to be placed in excel
   if first:
     present_left = v_edges_tb[0]
     v_edges_left = [e for e in v_edges if e["x0"] < present_left["x0"]]
     unique_x0 = []
     for e in v_edges_left:
        if e["x0"] not in unique_x0:
            unique_x0.append(e["x0"])
     spaces = len(unique_x0)
    
     index1 = 1
     index2 = 1+spaces
    
     return index1,index2

   else:
    base_edge_t = h_edges_tb[0]
    base_edge_v = v_edges_tb[0]
    for i, h_edge in enumerate(h_edges_tb):
        index1, index2 = IndexCalculation1(h_edge,tableFactor,base_edge_t,base_edge_v)
        if index1 != -1 and index2 != -1:
            return index1+i , index2
    for j, v_edge in enumerate(v_edges_tb):
        index1, index2 = IndexCalculation2(v_edge,tableFactor,base_edge_t,base_edge_v)
        if index1 != -1 and index2 != -1:
            return index1, index2+j 
        
    if index1 == -1 and index2 == -1:
        index1, index2 = IndexCalculation3(h_edges_tb[0],tableFactor,base_edge_t,base_edge_v)
        return index1, index2
        
    

        
    


def FindMergeDataModified(mergeRowsData,index1,index2):
    # updated merged cells data of tables except first table
    updated_data = {}
    for key,tuples in mergeRowsData.items():
        for t in tuples:
            new_key = key + index1-1 # increase key by index1 value
            new_tuple = (t[0] + index2-1, t[1], t[2])  # increase first value by index2 value
            updated_data.setdefault(new_key, []).append(new_tuple)

    return updated_data






def CheckRearrangeTables(tables):
    """As the tables are extracted based on h_edge["top"]
       it will rearrange tables based on table position in pdf"""
    tablesOrdered = []
    if len(tables) == 1:
        return tables
    else:
        for i,table in enumerate(tables):
            if i < len(tables)-1:
            
                t1 = tables[i]
                t2 = tables[i+1]
                (h_edges_tb_1, v_edges_tb_1, h_edges_1,v_edges_1, tableComp_1,table_1, max_dim_1) = t1
                (h_edges_tb_2, v_edges_tb_2, h_edges_2,v_edges_2, tableComp_2, table_2,max_dim_2) = t2
                
                
                if h_edges_tb_1[0]["top"] < h_edges_tb_2[0]["top"] and SameLine(h_edges_tb_1,h_edges_tb_2)  :
                    if v_edges_tb_2[0]["x0"] < v_edges_tb_1[0]["x0"] :
                        
                        tablesOrdered.append(t2)
                        if i == len(tables) - 2:
                            tablesOrdered.append(t1)
                        
                        
                    else:
                        
                        tablesOrdered.append(t1)
                        if i == len(tables) - 2:
                            tablesOrdered.append(t2)
                    
                        
                        
                else:
                    tablesOrdered.append(t1)
                    if i == len(tables) - 2:
                            tablesOrdered.append(t2)
                
        return tablesOrdered
    
   


def ArrangeTables(tables):
    """Arranges tables based on their positions in pdf"""
    tableFactor = []
    sheetData = []
    lindex1 = 1
    lindex2 = 1
    try:
        for i,table in enumerate(tables):
            (h_edges_tb, v_edges_tb, h_edges, v_edges, tableComp,table_o, max_dim_o) = table
            max_dim = max_dim_o
            if i==0:
                index1, index2 = IndexCalculation(h_edges_tb,v_edges_tb,h_edges,v_edges,tableFactor,True)
            
                maxRow,maxColumn = max_dim
                max_dim_1 = (maxRow+index1,index1,maxColumn+index2,index2)
            
                tableFactor.append((h_edges_tb,v_edges_tb,max_dim_1))
                sheetData.append((index1,index2,tableComp,table_o,max_dim_o))
                lindex1 = index1
                lindex2 = index2
                
            else:
                index1, index2 = IndexCalculation(h_edges_tb,v_edges_tb,h_edges,v_edges,tableFactor,False)
                maxRow,maxColumn = max_dim
                max_dim_1 = (maxRow+index1,index1,maxColumn+index2,index2)
                
                tableFactor.append((h_edges_tb,v_edges_tb,max_dim_1))
                sheetData.append((index1,index2,tableComp,table_o,max_dim_o))
                lindex1 = index1
                lindex2 = index2
    except:
          sheetData,lindex1+maxRow+1,lindex2,False

    return sheetData,index1+maxRow+1,index2,True

                
   
            


def TableWriter(h_edges,v_edges,output_image):

    # Returns s valid table
    
    table = None
    tableData = None
    
   
    # Save the result
        
    try: 
        noTableFound = False
        isTableFormedMain = False
                    
        
                    
        output_image_1 = output_image.copy()
                    
        table,h_edges_tb,v_edges_tb,h_edges,v_edges,h_edges_o,v_edges_o,pts,tableComp,max_dim = Tableextractor(output_image_1,h_edges+v_edges)
        
       
            
                
        if pts:  
        
            tableData = (h_edges_tb,v_edges_tb,h_edges,v_edges,tableComp,table,max_dim)      
                
            h_edges_t1 = [e for e in h_edges if e not in h_edges_tb]
            v_edges_t1 = [e for e in v_edges if e not in v_edges_tb]
                
            isTableFormedMain = True
                    
    except:
        
        noTableFound = True
        if table == None:
            return isTableFormedMain,noTableFound,{},{},None,None,True
        
        return isTableFormedMain,noTableFound,h_edges_o,v_edges_o,None,table,False
            
    
    if tableData:           
        return True,False,h_edges_t1,v_edges_t1,tableData,table,True
    else:
        return True,False,h_edges,v_edges,None,None,False
               
    


def CountNonEmpty(data):
    # Counts Non-empty cells in the table formed
    for i, row in enumerate(data):
        total = sum(1 for row in data for item in row if item != '')
    return total


def ExcelWriter(tableComp,max_dim,textGrouped):
    # Returns excelData to be written to openpyxl
    textCells,textGrouped = FindTextinCells(tableComp,textGrouped)
    textClusters = FindTextClusters(textCells,max_dim)
   
    mergeRowsData = FindMergeData(textClusters,max_dim)
            
    textDataFrame = FindDataFrame(textClusters)
    excelData = ModifyDataFrame(textDataFrame)
    
    count  = CountNonEmpty(excelData)
    return mergeRowsData,excelData,textGrouped,count


def ReturnText(j):
    # Strips the text and returns it
    text = (j.get("text") or "").strip()
    return text

def isValidTable(table,textGrouped,max_dim):
    # Checks if text is as high as the table 
   
    height = table[3]["top"] - table[2]["top"]
    top = table[2]["top"]
    bottom = table[3]["top"]

    for t in textGrouped:
        if top <= t <= bottom:
            for j in textGrouped[t]:
                heightText = j["height"]
                text = ReturnText(j)
                if text!= "" and text!=None:
                    if heightText >= 0.8*height:
                        return False
    
    return True  






def WriteText(sheetData,textGrouped,wb,ws):
    
    # Writes table data to openpyxl
    sum = 0
    count = 0
    notValid = []
    
    for i,sheet in enumerate(sheetData):
        (index1, index2, tableComp, table_o,max_dim) = sheet
        if not isValidTable(table_o,textGrouped,max_dim):
            notValid.append(sheet)
    
    if notValid: 
        return wb,sum,count,notValid
    
    for i,sheet in enumerate(sheetData):
        (index1, index2, tableComp, table_o, max_dim) = sheet
        sum = sum + len(tableComp)
        
        mergeRowsData,excelData,textGrouped,countTable = ExcelWriter(tableComp,max_dim,textGrouped)
        count = count + countTable
        mergeRowsModified = FindMergeDataModified(mergeRowsData,index1,index2)
        write_to_excel(index1,index2,wb,ws,excelData,mergeRowsModified,max_dim)
        
    return wb,sum,count,notValid,textGrouped



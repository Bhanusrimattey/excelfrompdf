# Copyright (c) 2024 Bhanusri mattey
# Licensed under the Business Source License 1.1
# See LICENSE file in the project root for details.
# Commercial use prohibited until 2030-03-01

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import range_boundaries
from openpyxl.styles import Border, Side

import cv2
import numpy as np
import os

from core.deskew import*
from paddleocr import PaddleOCR
import paddle
import bisect
import copy
import statistics

from collections import defaultdict
from openpyxl.utils import get_column_letter

def deskew_box(pts):
    """
    pts: list or array of 4 points [[x,y], ...] in clockwise order
         (top-left, top-right, bottom-right, bottom-left)

    returns:
        rotated_pts : points rotated to be straight
        aligned_box : perfectly axis-aligned rectangle
        angle_deg   : detected rotation angle (degrees)
    """

    pts = np.array(pts, dtype=float)

    # 1️⃣ compute angle from top edge (p0 -> p1)
    dx = pts[1][0] - pts[0][0]
    dy = pts[1][1] - pts[0][1]
    angle = np.arctan2(dy, dx)
    angle_deg = np.degrees(angle)

    # 2️⃣ rotation matrix (deskew → rotate back)
    cos_a = np.cos(-angle)
    sin_a = np.sin(-angle)
    R = np.array([
        [cos_a, -sin_a],
        [sin_a,  cos_a]
    ])

    # 3️⃣ rotate around center
    center = pts.mean(axis=0)
    rotated_pts = (pts - center) @ R.T + center

    # 4️⃣ axis-aligned bounding box
    xmin, ymin = rotated_pts.min(axis=0)
    xmax, ymax = rotated_pts.max(axis=0)

    aligned_box = np.array([
        [xmin, ymin],
        [xmax, ymin],
        [xmax, ymax],
        [xmin, ymax]
    ])

    return rotated_pts,aligned_box,-angle

def nearedges(a,b,tolerance):
  
    if abs(a-b) <= tolerance:
        return True
    return False

def snap_close_values(values, tolerance):
    
    values = sorted(values)
    snapped = []
    group = []

    for i, val in enumerate(values):
        if not group:
            group.append(val)
        elif abs(val - group[-1]) <= tolerance:
            group.append(val)
        else:
            avg = sum(group) / len(group)
            snapped.append(avg)
            group = [val]

    # handle last group
    if group:
        avg = sum(group) / len(group)
        snapped.append(avg)

    return snapped

class EdgeIndex:
    def __init__(self, edges):
        # Store lists of edges per value
        self.by_x1 = defaultdict(list)
        self.by_top = defaultdict(list)
        self.by_x0 = defaultdict(list)
        self.by_bottom = defaultdict(list)

        for edge in edges:
            self.by_x1[edge["x1"]].append(edge)
            self.by_top[edge["top"]].append(edge)
            self.by_x0[edge["x0"]].append(edge)
            self.by_bottom[edge["bottom"]].append(edge)

        # Sorted unique keys for bisect
        self.x1_keys = sorted(self.by_x1.keys())
        self.top_keys = sorted(self.by_top.keys())
        self.x0_keys = sorted(self.by_x0.keys())
        self.bottom_keys = sorted(self.by_bottom.keys())

    def get_closest_key(self, val, sorted_keys, tolerance=10):
        
    # Use bisect to find the insertion point
        idx = bisect.bisect_left(sorted_keys, val)

        candidates = []
        if idx > 0:
            candidates.append(sorted_keys[idx - 1])
        if idx < len(sorted_keys):
            candidates.append(sorted_keys[idx])

    # Find the closest candidate within tolerance
        closest = min(
        (key for key in candidates if abs(key - val) <= tolerance),
        key=lambda k: abs(k - val),
        default=None
        )
        return closest
    # These now return a LIST of matching edges
    def find_by_x1(self, x1_val, tolerance=10):
        
        key = self.get_closest_key(x1_val, self.x1_keys, tolerance)
        return self.by_x1.get(key, [])

    def find_by_x0(self, x0_val, tolerance=10):
       
        key = self.get_closest_key(x0_val, self.x0_keys, tolerance)
        return self.by_x0.get(key, [])

    def find_by_top(self, top_val, tolerance=10):
        
        key = self.get_closest_key(top_val, self.top_keys, tolerance)
        return self.by_top.get(key, [])

    def find_by_bottom(self, bottom_val, tolerance=10):
        
        key = self.get_closest_key(bottom_val, self.bottom_keys, tolerance)
        return self.by_bottom.get(key, [])
    
    def find_by_x1_and_bottom(self, x1_val, bottom_val, x1_tol=10, bottom_tol=10):
        
        x1_key = self.get_closest_key(x1_val, self.x1_keys, x1_tol)
        x1_matches = self.by_x1.get(x1_key, [])

    # Now filter x1_matches by top tolerance manually
        filtered = [
            edge for edge in x1_matches
            if abs(edge["bottom"] - bottom_val) <= bottom_tol
        ]

        return filtered
    
    def find_by_x0_and_top(self, x0_val, top_val, x0_tol=10, top_tol=10):
        
        x0_key = self.get_closest_key(x0_val, self.x0_keys, x0_tol)
        x0_matches = self.by_x0.get(x0_key, [])

    # Now filter x1_matches by top tolerance manually
        filtered = [
            edge for edge in x0_matches
            if abs(edge["top"] - top_val) <= top_tol
        ]

        return filtered

def average_width(text_list):
    widths = [item["width"] for item in text_list if "width" in item]
    return sum(widths) / len(widths) if widths else 0

def shift_list(vals, delta=10):
    return [v - delta for v in vals[0:]] + [vals[-1] + delta]

def TextDetection_text(allText,tolerance=10):
    # Detects text from paddleocr
    def SnapRows_top(text,textsnapped):
        for t in text:
            for top in textsnapped:
                if nearedges(t["top"],top,10):
                    t["top"] = top
                    t["bottom"] = top
    
    def SnapRows_x0(text,textsnapped):
        for t in text:
            for x0 in textsnapped:
                if nearedges(t["x0"],x0,10):
                    t["x0"] = x0
                   
    def SnapRows_x1(text,textsnapped):
        for t in text:
            for x1 in textsnapped:
                if nearedges(t["x1"],x1,10):
                    t["x1"] = x1

    
    text_dict = EdgeIndex(allText)
    textsnapped_top = snap_close_values(text_dict.top_keys,10)
    SnapRows_top(allText,textsnapped_top)
    textsnapped_x0 = snap_close_values(text_dict.x0_keys,10)
    SnapRows_x0(allText,textsnapped_x0)
    textsnapped_x1 = snap_close_values(text_dict.x1_keys,10)
    SnapRows_x1(allText,textsnapped_x1)
    rawText = copy.deepcopy(allText)
    allText = sorted(allText, key=lambda e: e["top"])
    
    def group_text_by_top(text: list, tolerance: int = 10):
        text_group = defaultdict(list)
        for t in text:
            top = t["top"]
            matched = False
            for key in text_group:
                if abs(top - key) <= tolerance:
                    text_group[key].append(t)
                    matched = True
                    break
            if not matched:
                text_group[top].append(t)
        return text_group

    textGroupedY = group_text_by_top(allText)

    allText = sorted(allText, key=lambda e: e["x0"])
    

    def group_text_by_xm(text: list, tolerance: int = 10):
        text_group = defaultdict(list)
        for t in text:
            top =  t["x0"] + (t["x1"] - t["x0"])/2
            matched = False
            for key in text_group:
                if abs(top - key) <= tolerance:
                    text_group[key].append(t)
                    matched = True
                    break
            if not matched:
                text_group[top].append(t)
        return text_group
    
    
    
    def group_text_by_xs(text: list, tolerance: int = 10):
        text_group = defaultdict(list)
        for t in text:
            top =  t["x0"]
            matched = False
            for key in text_group:
                if abs(top - key) <= tolerance:
                    text_group[key].append(t)
                    matched = True
                    break
            if not matched:
                text_group[top].append(t)
        return text_group
    
    
    
    def group_text_by_xe(text: list, tolerance: int = 10):
        text_group = defaultdict(list)
        for t in text:
            top =  t["x1"]
            matched = False
            for key in text_group:
                if abs(top - key) <= tolerance:
                    text_group[key].append(t)
                    matched = True
                    break
            if not matched:
                text_group[top].append(t)
        return text_group

    textGroupedXm = group_text_by_xm(allText[:])
    allText = sorted(allText, key=lambda e: e["x0"])
    textGroupedXs = group_text_by_xs(allText[:])
    allText = sorted(allText, key=lambda e: e["x0"])
    textGroupedXe = group_text_by_xe(allText[:])
    allText = sorted(allText, key=lambda e: e["x0"])
    
    return allText,textGroupedXm,textGroupedXs,textGroupedXe,textGroupedY


   

def get_max_list_info_x(data_dict_1, data_dict_2, data_dict_3):
    """
    Finds the key with the longest list value across three dictionaries
    and returns the key, list, and length.
    """

    results = []

    for idx, d in enumerate([data_dict_1, data_dict_2, data_dict_3]):
        if d:
            key = max(d, key=lambda k: len(d[k]))
            lst = d[key]
            results.append((key, lst, len(lst)))

    if not results:
        return None, None, 0

    # pick the tuple with maximum length
    #key_with_max_len, max_list, max_length = max(results, key=lambda x: x[2])
   

    return results


def get_max_list_info(data_dict):
    """
    Finds the key with the longest list value in a dictionary
    and returns the key, the list, and its length.
    """

    if not data_dict:
        return None, None, 0

    key_with_max_len = max(data_dict, key=lambda k: len(data_dict[k]))
    max_list = data_dict[key_with_max_len]
    max_length = len(max_list)

    return key_with_max_len, max_list, max_length


def extract_y_values(elements):
    """
    Extracts the 'top' values from a list of dictionaries.

    Args:
        elements (list[dict]): List of dictionaries containing a 'top' key.

    Returns:
        list[float]: List of all 'top' values found.
    """
    if not isinstance(elements, list):
        raise ValueError("Input must be a list of dictionaries.")

    top_values = [item['top'] for item in elements if 'top' in item]
    return top_values

def extract_x_values(elements):
    """
    Extracts x-values from a list of dictionaries:
    - For the first item: uses 'x0'
    - For the remaining items: uses 'x1'

    Args:
        elements (list[dict]): List of dictionaries containing 'x0' and 'x1' keys.

    Returns:
        list[float]: List of extracted x-values.
    """
    if not isinstance(elements, list):
        raise ValueError("Input must be a list of dictionaries.")

    if not elements:
        return []

    x_values = []

    for idx, item in enumerate(elements):
        if idx == 0 and 'x0' in item and "x1" in item:
            x_values.append(item['x0'])
            x_values.append(item["x1"])
        elif 'x1' in item:
            x_values.append(item['x1'])

    return x_values

def sort_values_by_top_inplace(data):
    # Sorts values by top
    for v in data.values():
        v.sort(key=lambda item: item.get('top', float('inf')))
    return data

def group_text_excel(allText,xPositions,yPositions,maxRow,maxCol):
    # Groups text for writing to openpyxl
    excelData = [['' for _ in range(maxCol-1)] for _ in range(maxRow-1)]
    for text in allText:
        
        for i in range(maxRow-1):
            isFound = False
            for j in range(maxCol-1):
                left = xPositions[j]
                right = xPositions[j+1]
                top = yPositions[i]
                bottom = yPositions[i+1]
                if left <= text["x0"]  <= right and top <= text["top"] <= bottom:
                    
                    if not excelData[i][j]:
                        excelData[i][j] = text["text"]
                    else:
                        # If text already exists, append new line + new text
                        excelData[i][j] = str(excelData[i][j]) + "\n" + text["text"]
                    isFound = True
                    break
                
            if isFound == True:
                break
    return excelData

def AppendLast(values):
    diffs = [values[i+1] - values[i] for i in range(len(values)-1)]

    # Step 2: compute average difference
    avg_diff = sum(diffs) / len(diffs)

    # Step 3: add average difference to the last value
    next_val = values[-1] + avg_diff

    # Step 4: append it to the list
    values.append(next_val)
    return values


def write_to_excel_text(index1,index2,wb,ws,data,max_dim):
    # writes to openpyxl
    maxRow,maxColumn = max_dim
    
    
    for row_idx in range(index1,maxRow+index1-1):
        rowData = data[row_idx-index1]
        if rowData:
            i = 0
            for col_idx in range(index2,maxColumn+index2-1):
                if i < len(rowData):
                    colValue = rowData[i]
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = colValue
                    cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    i = i + 1
                
    
    thin_border = Border(
        left = Side(style='thin'),
        right = Side(style='thin'),
        top = Side(style='thin'),
        bottom = Side(style='thin')
    )

    for row in ws.iter_rows(min_row=index1,
    max_row=maxRow+index1-2, min_col=index2, max_col=maxColumn+index2-2):
        for cell in row:
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                
                cell_text = str(cell.value)

                lines = cell_text.split("\n")
                max_len = max(len(line) for line in lines)

                max_length = max(max_length, max_len)

        ws.column_dimensions[col_letter].width = max_length + 2

    

def ExpandList(results,allText):
    # Expands the List
    resultsF = copy.deepcopy(results)
    for i, result in enumerate(results):
        if i < len(results) - 1:
            top = results[i]["top"]
            bottom = results[i+1]["top"]
            left = max(results[i]["x0"],results[i+1]["x0"])
            right = max(results[i]["x1"],results[i+1]["x1"])
            for t in allText:
                xPos = t["x0"] + (t["x1"] - t["x0"])/2
                if top < t["top"] < bottom and left < xPos < right:
                    resultsF.append(t)
    return resultsF

def sort_values_by_x(data):
    filtered = defaultdict(list)

    for y, items in data.items():
        
        if len(items) > 1:
            filtered[y] = sorted(
                items,
                key=lambda item: float(item.get("x0", 0))
            )

    return filtered

def NewText(allText,textGroupedY):
    # returns text that is in textGrouped
   newText = []

   for items in textGroupedY.values():
        for item in items:
            for text in allText:
                if item == text:
                    newText.append(item)
            
   return newText

def group_text_by_xm_text(text: list, tolerance: int = 10):
        # Gorups text by center
        text_group = defaultdict(list)
        for t in text:
            top =  t["x0"] + (t["x1"] - t["x0"])/2
            matched = False
            for key in text_group:
                if abs(top - key) <= tolerance:
                    text_group[key].append(t)
                    matched = True
                    break
            if not matched:
                text_group[top].append(t)
        return text_group
    
    
    
def group_text_by_xs_text(text: list, tolerance: int = 10):
        # Gorups text by x0
        text_group = defaultdict(list)
        for t in text:
            top =  t["x0"]
            matched = False
            for key in text_group:
                if abs(top - key) <= tolerance:
                    text_group[key].append(t)
                    matched = True
                    break
            if not matched:
                text_group[top].append(t)
        return text_group
    
    
    
def group_text_by_xe_text(text: list, tolerance: int = 10):
        # Groups text by x1
        text_group = defaultdict(list)
        for t in text:
            top =  t["x1"]
            matched = False
            for key in text_group:
                if abs(top - key) <= tolerance:
                    text_group[key].append(t)
                    matched = True
                    break
            if not matched:
                text_group[top].append(t)
        return text_group  



def newValueX(textGroupedY):
    # Find the maximum rowList
    keyY, rowList, maxCol = get_max_list_info(textGroupedY)
    return rowList,maxCol


def newValue(results,allText):
    # Find the maximum colList comparing with left, right and centered text
    resultsF = []
    
    (key_0,resultsF_0,length_0) = results[0]
    resultsF_0 = ExpandList(results[0][1],allText)
    resultsF_0.sort(key=lambda e: e["top"])
    length_0 = len(resultsF_0)
    resultsF.append((key_0,resultsF_0,length_0))
    
    
    (key_1,resultsF_1,length_1) = results[1]
    resultsF_1 = ExpandList(results[1][1],allText)
    resultsF_1.sort(key=lambda e: e["top"])
    length_1 = len(resultsF_1)
    resultsF.append((key_1,resultsF_1,length_1))

    
    (key_2,resultsF_2,length_2) = results[2]
    resultsF_2 = ExpandList(results[2][1],allText)
    resultsF_2.sort(key=lambda e: e["top"])
    length_2 = len(resultsF_2)
    resultsF.append((key_2,resultsF_2,length_2))

    key_with_max_len, max_list, max_length = max(resultsF, key=lambda x: x[2])
    return key_with_max_len, max_list, max_length

def get_key_index_by_top(textGrouped, target_top, tolerance=1e-6):
    # returns key index by top
    for idx, items in enumerate(textGrouped.values()):
        for item in items:
            if abs(item["top"] - target_top) < tolerance:
                return idx
    return -1


def split_textgrouped_by_index(textGrouped, idx):
    # splits text by index
    items = list(textGrouped.items())

    before = defaultdict(list, items[:idx+1])
    
    after = defaultdict(list, items[idx+1:])

    return before,after


def DetectTable(textGroupedY,colList):
    # Returns text of the table detected and the remaining text
    tableEnd = colList[-1]["top"]
    idx = get_key_index_by_top(textGroupedY,tableEnd)
    textGroupedY_p,textGroupedY_r = split_textgrouped_by_index(textGroupedY,idx)
    return textGroupedY_p,textGroupedY_r

def crop_at_large_jump(data):
    # crops the data if it crooses certain threshold
    gaps = [data[i + 1]['top'] - data[i]['top'] for i in range(len(data) - 1)]
    if not gaps:
        return data
    
    mean_gap = statistics.mean(gaps)
    stdev_gap = statistics.stdev(gaps) if len(gaps) > 1 else 0
    
    for i, gap in enumerate(gaps):
        if gap > mean_gap + 3.5 * stdev_gap:  # true outlier
            return data[:i + 1]
    
    return data  # no abrupt jump found, return full list


def crop_at_large_jump_1(data):
    # crops the data if it crooses certain threshold
    gaps = [data[i + 1]['x0'] - data[i]['x0'] for i in range(len(data) - 1)]
    if not gaps:
        return data
    
    mean_gap = statistics.mean(gaps)
    stdev_gap = statistics.stdev(gaps) if len(gaps) > 1 else 0
    
    for i, gap in enumerate(gaps):
        if gap > mean_gap + 3.5 * stdev_gap:  # true outlier
            return data[i:]
    
    return data  

def AddLeftOrRight(rowList,allText):
    # Adds LEft or right text to rowList
    left = rowList[0]["x0"]
    right = rowList[-1]["x1"]
    leftTo = []
    rightTo = []
    leftNo = 0
    rightNo = 0
    idx1 = 1
    idx2 = 1
    firstTo2 = None
    firstTo1 = None
    for text in allText:
        if text["x0"] < left and text["x1"] < left:
            leftTo.append(text)
           
    for text in allText:
        if text["x0"] > right and text["x1"] > right:
            rightTo.append(text)
   
    if leftTo:
        leftNo = 1
        firstTo1 = leftTo[0]
        rowList.insert(0, firstTo1)
        while idx1 < len(leftTo):
            if firstTo1:
                if leftTo[idx1]["x0"] < firstTo1["x0"] and leftTo[idx1]["x1"] < firstTo1["x0"]:
                    firstTo1 = leftTo[idx1]
                    leftNo = leftNo + 1
                    rowList.insert(0, firstTo1)
            idx1 = idx1 + 1
    if rightTo:
        rightNo = 1
        firstTo2 = rightTo[0]
        rowList.append(firstTo2)
        while idx2 < len(rightTo):
            if firstTo2:
                if rightTo[idx2]["x0"] > firstTo2["x1"] and rightTo[idx2]["x1"] > firstTo2["x1"]:
                    firstTo2 = rightTo[idx2]
                    rightNo = rightNo + 1
                    rowList.append(firstTo2)
            idx2 = idx2 + 1
    return rowList,leftNo,rightNo
            
def AddTopOrBottom(colList,allText):
    # Adds top or bottom text to colList
    top = colList[0]["top"]
    bottom = colList[-1]["top"]
    topTo = []
    bottomTo = []
    topNo = 0
    bottomNo = 0
    idx1 = 1
    idx2 = 1
    firstTo2 = None
    firstTo1 = None
    for text in allText:
        if text["top"] < top and text["bottom"] < top:
            topTo.append(text)
         
    for text in allText:
        if text["top"] > bottom and text["bottom"] > bottom:
            bottomTo.append(text)
    
    if topTo:
        topNo = 1
        firstTo1 = topTo[0]
        colList.insert(0, firstTo1)
        while idx1 < len(topTo):
            if firstTo1:
                if topTo[idx1]["top"] < firstTo1["top"] and topTo[idx1]["bottom"] < firstTo1["top"]:
                    firstTo1 = topTo[idx1]
                    topNo = topNo + 1
                    colList.insert(0, firstTo1)
            idx1 = idx1 + 1
    if bottomTo:
        bottomNo = 1
        firstTo2 = bottomTo[0]
        colList.append(firstTo2)
        while idx2 < len(bottomTo):
            if firstTo2:
                if bottomTo[idx2]["top"] > firstTo2["bottom"] and bottomTo[idx2]["bottom"] > firstTo2["bottom"]:
                    firstTo2 = bottomTo[idx2]
                    bottomNo = bottomNo + 1
                    colList.append(firstTo2)
            idx2 = idx2 + 1
    return colList,topNo,bottomNo

def ArrangeTablesText(tables,lindex1,lindex2):
    # Arranges tables based on its positions in pdf
    tableData = []
    if len(tables) == 1:
        (excelData_1,max_dim_1,last_1,right_1,first_1,left_1) = tables[0]
        tableData.append((lindex1,lindex2,excelData_1,max_dim_1))
    else:
        for i,table in enumerate(tables):
            if i < len(tables) - 1:
                (excelData_1,max_dim_1,bottom_1,right_1,top_1,left_1) = tables[i]
                (excelData_2,max_dim_2,bottom_2,right_2,top_2,left_2) = tables[i+1]
                if i==0:
                    tableData.append((lindex1,lindex2,excelData_1,max_dim_1))
                lindex1, lindex2, excelData_1, max_dim_1 = tableData[i]
                if left_2 < left_1 and bottom_2 > bottom_1 and top_2 > bottom_1:
                    (maxRow_1,maxColumn_1) = max_dim_1
                    (maxRow_2,maxColumn_2) = max_dim_2
                    for j, t in enumerate(tableData):
                        if t[2] == excelData_1:
                            lindex3, lindex4, excelData, max_dim = tableData[j]
                            # update what you need
                            tableData[i] = (lindex3+maxRow_2+1, lindex4+maxColumn_2+1, excelData, max_dim)
                            break
                    tableData.append((lindex1+maxRow_1+1,lindex2,excelData_2,max_dim_2))
                    
                elif top_2 > bottom_1 and bottom_2 > bottom_1:
                    (maxRow_1,maxColumn_1) = max_dim_1
                    (maxRow_2,maxColumn_2) = max_dim_2
                    lindex1, lindex2, excelData_1, max_dim_1 = tableData[i]
                    tableData.append((maxRow_1+lindex1+2,lindex2,excelData_2,max_dim_2))
                
                elif top_2 > bottom_1 and left_2 > right_1 and right_2 > right_1:
                    (maxRow_1,maxColumn_1) = max_dim_1
                    (maxRow_2,maxColumn_2) = max_dim_2
                    lindex1, lindex2, excelData_1, max_dim_1 = tableData[i]
                    tableData.append((maxRow_1+lindex1+2,maxColumn_1+lindex2+2,excelData_2,max_dim_2))
                
                elif left_2 > right_1 and right_2 > right_1:
                    (maxRow_1,maxColumn_1) = max_dim_1
                    (maxRow_2,maxColumn_2) = max_dim_2
                    lindex1, lindex2, excelData_1, max_dim_1 = tableData[i]
                    tableData.append((lindex1,maxColumn_1+lindex2+2,excelData_2,max_dim_2))


    return tableData           

def TableDetectionText(allText):

    # Detects tables if there is unique text which spans 2 or more columns in a row

    allText,textGroupedXm,textGroupedXs,textGroupedXe,textGroupedY =TextDetection_text(allText)
    tables = []
    
    while allText:
        colList = []
        rowList = []
        maxRow = 0
        maxCol = 0
        results = []
        textGroupedY = sort_values_by_x(textGroupedY)
        #print(textGroupedY)
        
        allText = NewText(allText,textGroupedY)
        
    
        if allText:
            allText = sorted(allText, key=lambda e: e["x0"])
            textGroupedXm = group_text_by_xm_text(allText[:])
            allText = sorted(allText, key=lambda e: e["x0"])
            textGroupedXs = group_text_by_xs_text(allText[:])
            allText = sorted(allText, key=lambda e: e["x0"])
            textGroupedXe = group_text_by_xe_text(allText[:])
            allText = sorted(allText, key=lambda e: e["x0"])

            textGroupedXm = sort_values_by_top_inplace(textGroupedXm)
            textGroupedXs = sort_values_by_top_inplace(textGroupedXs)
            textGroupedXe = sort_values_by_top_inplace(textGroupedXe)

            results = get_max_list_info_x(textGroupedXm,textGroupedXs,textGroupedXe)
            
            
            if results:
                keyX, colList, maxRow = newValue(results,allText)
                
                if not colList:
                    break
                colList = crop_at_large_jump(colList)
                
                maxRow = len(colList)
                
                textGroupedY_p,textGroupedY_r = DetectTable(textGroupedY,colList)
                
                allText_p = NewText(allText,textGroupedY_p)
                allText_r = NewText(allText,textGroupedY_r)
                
                rowList,maxCol = newValueX(textGroupedY_p)
                
                if not rowList:
                    break
                
                
                rowList = crop_at_large_jump_1(rowList)
                
                maxCol = len(rowList)
                
                
            
                if rowList and colList:
                    rowList,leftNo,rightNo = AddLeftOrRight(rowList,allText_p)
                    colList,topNo,bottomNo = AddTopOrBottom(colList,allText_p)

                    xPositions = extract_x_values(rowList)
                    yPositions = extract_y_values(colList)
                    yPositions = shift_list(yPositions)
                    
                    
                    
                    
                    textGroupedY = copy.deepcopy(textGroupedY_r)

                    allText = sorted(allText, key=lambda e: e["top"])


                    max_dim = (maxRow+topNo+bottomNo+1,maxCol+leftNo+rightNo+1)
                    excelTextData = group_text_excel(allText_p,xPositions,yPositions,maxRow+topNo+bottomNo+1,maxCol+leftNo+rightNo+1)
                    tables.append((excelTextData,max_dim,colList[-1]["top"],rowList[-1]["x1"],colList[0]["top"],rowList[0]["x0"]))
       
    return tables

def TableWriterText(tables,wb,ws,lindex1,lindex2):
    tableData = ArrangeTablesText(tables,lindex1,lindex2)

    for tableD in tableData:
        (index1,index2,excelTextData,max_dim) = tableD
        write_to_excel_text(index1,index2,wb,ws,excelTextData,max_dim)

